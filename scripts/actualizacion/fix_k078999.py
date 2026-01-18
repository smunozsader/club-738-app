#!/usr/bin/env python3
import openpyxl
import shutil
from datetime import datetime

archivo = 'data/socios/Copy of 2026.31.01_RELACION_SOCIOS_ARMAS_SEPARADO_verified.xlsx'

# Backup
backup = archivo.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
shutil.copy(archivo, backup)
print(f"Backup: {backup}")

# Cargar
wb = openpyxl.load_workbook(archivo)
ws = wb.active

# Buscar datos de Arechiga
arechiga_data = None
for row in ws.iter_rows(min_row=2):
    if row[13].value == 'arechiga@jogarplastics.com':
        arechiga_data = row
        break

# Reasignar K078999
if arechiga_data:
    for row in ws.iter_rows(min_row=2):
        if row[18].value == 'K078999' and row[13].value == 'jrgardoni@gmail.com':
            row[13].value = 'arechiga@jogarplastics.com'
            row[3].value = arechiga_data[3].value
            row[4].value = arechiga_data[4].value
            row[2].value = arechiga_data[2].value
            row[5].value = arechiga_data[5].value
            row[12].value = arechiga_data[12].value
            wb.save(archivo)
            print("✅ K078999 reasignada a Arechiga")
            print("💾 Guardado")
            break
