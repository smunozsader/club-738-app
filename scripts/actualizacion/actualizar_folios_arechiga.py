#!/usr/bin/env python3
"""
Actualizar folios y correcciones finales
17 Enero 2026

CAMBIOS:
1. K078999 → Agregar FOLIO: A3601943 (sin espacios)
2. K084328 → Corregir modelo: P380 → LP380
3. C647155 → Folio pendiente (no encontrado en históricos)
"""

import openpyxl
import shutil
from datetime import datetime

archivo = 'data/socios/Copy of 2026.31.01_RELACION_SOCIOS_ARMAS_SEPARADO_verified.xlsx'

# Backup
backup = archivo.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
shutil.copy(archivo, backup)
print(f"✓ Backup: {backup}\n")

# Cargar
wb = openpyxl.load_workbook(archivo)
ws = wb.active

cambios = []

print("=" * 80)
print("ACTUALIZACIONES FINALES - ARECHIGA")
print("=" * 80)

for row in ws.iter_rows(min_row=2):
    matricula = row[18].value
    email = row[13].value
    
    # Solo armas de Arechiga
    if email == 'arechiga@jogarplastics.com':
        
        # K078999 - Agregar folio
        if matricula == 'K078999':
            row[19].value = 'A3601943'
            cambios.append(f"✅ K078999 - FOLIO agregado: A3601943")
        
        # K084328 - Corregir modelo
        elif matricula == 'K084328':
            if row[17].value == 'P380':
                row[17].value = 'LP380'
                cambios.append(f"✅ K084328 - Modelo corregido: P380 → LP380")

# Guardar
wb.save(archivo)

print("\n📝 CAMBIOS APLICADOS:")
for cambio in cambios:
    print(f"   {cambio}")

# Verificación final
print("\n" + "=" * 80)
print("🔍 VERIFICACIÓN FINAL - ARMAS DE ARECHIGA")
print("=" * 80)

wb = openpyxl.load_workbook(archivo)
ws = wb.active

print("\n✅ MARIA FERNANDA ARECHIGA:")
for row in ws.iter_rows(min_row=2):
    if row[13].value == 'arechiga@jogarplastics.com' and row[18].value:
        clase = row[14].value
        marca = row[16].value
        modelo = row[17].value
        matricula = row[18].value
        folio = row[19].value if row[19].value else '⚠️ PENDIENTE'
        
        print(f"   • {clase:20} {marca:20} {modelo:10} MAT: {matricula:15} FOLIO: {folio}")

print("\n" + "=" * 80)
print("✅ ACTUALIZACIÓN COMPLETADA")
print("=" * 80)
print("\n⚠️  C647155 (CZ P07) - FOLIO aún pendiente")
print("    Este folio deberá obtenerse del registro físico RFA")
print("=" * 80)
