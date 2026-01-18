#!/usr/bin/env python3
"""
Crear NUEVA FUENTE DE VERDAD combinando:
1. Base normalizada diciembre 2025 (276 armas)
2. + 4 armas nuevas enero 2026 (Iván Cabo, Gardoni, Arechiga)
3. + Direcciones estructuradas del archivo actual
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# Archivos
archivo_dic_normalizado = "2025-dic-usb-738/CLUB 738-31-DE-DICIEMBRE-2025_RELACION_SOCIOS_ARMAS_NORMALIZADO.xlsx"
archivo_actual = "data/socios/Copy of 2026.31.01_RELACION_SOCIOS_ARMAS_SEPARADO_verified.xlsx"
archivo_salida = "data/socios/2026_ENERO_RELACION_SOCIOS_ARMAS_MASTER.xlsx"

print("=" * 80)
print("CREANDO NUEVA FUENTE DE VERDAD - ENERO 2026")
print("=" * 80)

# 1. LEER ARCHIVO DICIEMBRE NORMALIZADO (base)
print("\n📖 Leyendo base diciembre normalizado...")
wb_dic = openpyxl.load_workbook(archivo_dic_normalizado)
ws_dic = wb_dic.active

# 2. LEER ARCHIVO ACTUAL (para direcciones estructuradas)
print("📖 Leyendo archivo actual (direcciones estructuradas)...")
wb_actual = openpyxl.load_workbook(archivo_actual)
ws_actual = wb_actual.active

# Crear diccionario de direcciones por CURP
direcciones_por_curp = {}
for i in range(2, ws_actual.max_row + 1):
    curp = ws_actual.cell(i, 5).value  # Col 5 = CURP
    if curp:
        direcciones_por_curp[curp] = {
            'calle': ws_actual.cell(i, 7).value,      # Col 7
            'colonia': ws_actual.cell(i, 8).value,    # Col 8
            'ciudad': ws_actual.cell(i, 9).value,     # Col 9
            'estado': ws_actual.cell(i, 10).value,    # Col 10
            'cp': ws_actual.cell(i, 11).value         # Col 11
        }

print(f"   ✅ {len(direcciones_por_curp)} CURPs con direcciones estructuradas")

# 3. CREAR NUEVO WORKBOOK
wb_nuevo = openpyxl.Workbook()
ws_nuevo = wb_nuevo.active
ws_nuevo.title = "ENERO_2026"

# 4. COPIAR HEADERS DEL ARCHIVO ACTUAL (tiene más columnas)
print("\n📋 Copiando headers...")
for j in range(1, ws_actual.max_column + 1):
    cell_val = ws_actual.cell(1, j).value
    ws_nuevo.cell(1, j, cell_val)
    # Formatear header
    cell = ws_nuevo.cell(1, j)
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 5. COPIAR DATOS DE DICIEMBRE + AGREGAR DIRECCIONES ESTRUCTURADAS
print("\n📝 Procesando armas de diciembre + direcciones estructuradas...")
fila_nueva = 2
armas_diciembre = 0
direcciones_agregadas = 0

# Primero, identificar la estructura del archivo diciembre normalizado
# Leer headers para mapear columnas
headers_dic = {}
for j in range(1, ws_dic.max_column + 1):
    val = ws_dic.cell(1, j).value
    if val:
        headers_dic[j] = str(val).strip()

print(f"   📋 Headers diciembre: {list(headers_dic.values())[:10]}...")

# Identificar columnas clave
col_nombre_dic = None
col_curp_dic = None
col_clase_dic = None

for j, header in headers_dic.items():
    if 'NOMBRE' in header.upper() and 'SOCIO' in header.upper():
        col_nombre_dic = j
    elif 'CURP' in header.upper():
        col_curp_dic = j
    elif 'CLASE' in header.upper():
        col_clase_dic = j

print(f"   🔍 CURP en columna: {col_curp_dic}")
print(f"   🔍 NOMBRE en columna: {col_nombre_dic}")
print(f"   🔍 CLASE en columna: {col_clase_dic}")

for i in range(2, ws_dic.max_row + 1):
    # Leer CURP del diciembre
    curp = ws_dic.cell(i, col_curp_dic).value if col_curp_dic else None
    
    # Buscar datos completos del socio en archivo actual por CURP
    datos_encontrados = False
    if curp:
        for k in range(2, ws_actual.max_row + 1):
            curp_actual = ws_actual.cell(k, 5).value
            if curp_actual == curp:
                # Copiar datos base del socio (cols 1-6)
                for j in range(1, 7):
                    ws_nuevo.cell(fila_nueva, j, ws_actual.cell(k, j).value)
                
                # Copiar dirección estructurada (cols 7-11)
                for j in range(7, 12):
                    ws_nuevo.cell(fila_nueva, j, ws_actual.cell(k, j).value)
                
                datos_encontrados = True
                direcciones_agregadas += 1
                break
    
    # Si no se encontró en archivo actual, copiar del diciembre
    if not datos_encontrados:
        for j in range(1, min(7, ws_dic.max_column + 1)):
            ws_nuevo.cell(fila_nueva, j, ws_dic.cell(i, j).value)
    
    # Copiar datos de arma del diciembre
    if col_clase_dic:
        # Mapear las columnas de arma del diciembre al nuevo formato
        # Diciembre tiene: CLASE, CALIBRE, MARCA, MODELO, MATRICULA, FOLIO
        # Nuevo tiene estas columnas en: 12, 13, 14, 15, 16, 17
        
        ws_nuevo.cell(fila_nueva, 12, ws_dic.cell(i, col_clase_dic).value)
        ws_nuevo.cell(fila_nueva, 13, ws_dic.cell(i, col_clase_dic + 1).value)  # CALIBRE
        ws_nuevo.cell(fila_nueva, 14, ws_dic.cell(i, col_clase_dic + 2).value)  # MARCA
        ws_nuevo.cell(fila_nueva, 15, ws_dic.cell(i, col_clase_dic + 3).value)  # MODELO
        ws_nuevo.cell(fila_nueva, 16, ws_dic.cell(i, col_clase_dic + 4).value)  # MATRICULA
        ws_nuevo.cell(fila_nueva, 17, ws_dic.cell(i, col_clase_dic + 5).value)  # FOLIO
    
    armas_diciembre += 1
    fila_nueva += 1

print(f"   ✅ {armas_diciembre} armas de diciembre procesadas")
print(f"   ✅ {direcciones_agregadas} direcciones estructuradas agregadas")

print(f"   ✅ {armas_diciembre} armas de diciembre procesadas")

# 6. AGREGAR 4 ARMAS NUEVAS DE ENERO
print("\n➕ Agregando 4 armas nuevas de enero...")

armas_nuevas = [
    # IVÁN CABO - 2 nuevas
    {
        'curp': 'CAVI880605HYSBVS09',
        'nombre': 'IVÁN ALEJANDRO CABO VELÁZQUEZ',
        'clase': 'PISTOLA',
        'calibre': '9 mm',
        'marca': 'RETAY',
        'modelo': 'GORDION',
        'matricula': '73-H21YT-001717',
        'folio': 'POR ASIGNAR'
    },
    {
        'curp': 'CAVI880605HYSBVS09',
        'nombre': 'IVÁN ALEJANDRO CABO VELÁZQUEZ',
        'clase': 'PISTOLA',
        'calibre': '9 mm',
        'marca': 'CZ',
        'modelo': 'SHADOW 2',
        'matricula': 'FP40104',
        'folio': 'POR ASIGNAR'
    },
    # GARDONI - 1 faltante
    {
        'curp': 'GAPJ650918HYNRRQ00',
        'nombre': 'JOAQUÍN MARÍA GARDONI PÉREZ',
        'clase': 'PISTOLA',
        'calibre': '9 mm',
        'marca': 'CZ',
        'modelo': 'SHADOW 2',
        'matricula': 'DP25087',
        'folio': 'POR ASIGNAR'
    },
    # ARECHIGA - 1 nueva (las otras 2 son transferencias de Gardoni)
    {
        'curp': 'AECL650813HYNHRR06',
        'nombre': 'LUIS FERNANDO ARECHIGA HERRERA',
        'clase': 'PISTOLA',
        'calibre': '9 mm',
        'marca': 'CZ',
        'modelo': 'P07',
        'matricula': 'C647155',
        'folio': 'POR ASIGNAR'
    }
]

for arma in armas_nuevas:
    curp = arma['curp']
    
    # Buscar datos del socio en archivo actual
    for i in range(2, ws_actual.max_row + 1):
        if ws_actual.cell(i, 5).value == curp:
            # Copiar datos base del socio
            ws_nuevo.cell(fila_nueva, 1, ws_actual.cell(i, 1).value)  # No. REGISTRO
            ws_nuevo.cell(fila_nueva, 2, ws_actual.cell(i, 2).value)  # DOMICILIO CLUB
            ws_nuevo.cell(fila_nueva, 3, ws_actual.cell(i, 3).value)  # No. CREDENCIAL
            ws_nuevo.cell(fila_nueva, 4, arma['nombre'])              # NOMBRE
            ws_nuevo.cell(fila_nueva, 5, curp)                        # CURP
            ws_nuevo.cell(fila_nueva, 6, ws_actual.cell(i, 6).value)  # No. CONSEC.
            
            # Dirección estructurada
            if curp in direcciones_por_curp:
                dir_data = direcciones_por_curp[curp]
                ws_nuevo.cell(fila_nueva, 7, dir_data['calle'])
                ws_nuevo.cell(fila_nueva, 8, dir_data['colonia'])
                ws_nuevo.cell(fila_nueva, 9, dir_data['ciudad'])
                ws_nuevo.cell(fila_nueva, 10, dir_data['estado'])
                ws_nuevo.cell(fila_nueva, 11, dir_data['cp'])
            
            break
    
    # Datos del arma
    ws_nuevo.cell(fila_nueva, 12, arma['clase'])
    ws_nuevo.cell(fila_nueva, 13, arma['calibre'])
    ws_nuevo.cell(fila_nueva, 14, arma['marca'])
    ws_nuevo.cell(fila_nueva, 15, arma['modelo'])
    ws_nuevo.cell(fila_nueva, 16, arma['matricula'])
    ws_nuevo.cell(fila_nueva, 17, arma['folio'])
    
    print(f"   ✅ {arma['nombre']}: {arma['marca']} {arma['modelo']} ({arma['matricula']})")
    fila_nueva += 1

# 7. AJUSTAR ANCHOS DE COLUMNA
print("\n🎨 Ajustando formato...")
ws_nuevo.column_dimensions['A'].width = 12
ws_nuevo.column_dimensions['B'].width = 50
ws_nuevo.column_dimensions['C'].width = 12
ws_nuevo.column_dimensions['D'].width = 40
ws_nuevo.column_dimensions['E'].width = 20
ws_nuevo.column_dimensions['F'].width = 12
ws_nuevo.column_dimensions['G'].width = 35
ws_nuevo.column_dimensions['H'].width = 25
ws_nuevo.column_dimensions['I'].width = 20
ws_nuevo.column_dimensions['J'].width = 20
ws_nuevo.column_dimensions['K'].width = 10
ws_nuevo.column_dimensions['L'].width = 15
ws_nuevo.column_dimensions['M'].width = 12
ws_nuevo.column_dimensions['N'].width = 20
ws_nuevo.column_dimensions['O'].width = 20
ws_nuevo.column_dimensions['P'].width = 20
ws_nuevo.column_dimensions['Q'].width = 15

# 8. GUARDAR
print(f"\n💾 Guardando archivo: {archivo_salida}")
wb_nuevo.save(archivo_salida)

# 9. ESTADÍSTICAS FINALES
print("\n" + "=" * 80)
print("✅ NUEVA FUENTE DE VERDAD CREADA")
print("=" * 80)
print(f"📅 Fecha: Enero 2026")
print(f"📊 Total armas: {armas_diciembre + len(armas_nuevas)}")
print(f"   - Base diciembre: {armas_diciembre}")
print(f"   - Nuevas enero: {len(armas_nuevas)}")
print(f"\n📂 Archivo guardado en:")
print(f"   {archivo_salida}")
print(f"\n✨ Características:")
print(f"   ✅ Direcciones estructuradas (calle, colonia, ciudad, estado, CP)")
print(f"   ✅ 4 armas nuevas agregadas")
print(f"   ✅ Base verificada de diciembre 2025")
print(f"   ✅ Headers formateados")

wb_dic.close()
wb_actual.close()
wb_nuevo.close()
