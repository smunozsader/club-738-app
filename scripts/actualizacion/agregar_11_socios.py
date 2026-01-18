#!/usr/bin/env python3
"""
Agregar 11 socios faltantes a la Nueva Fuente de Verdad
- 10 socios sin armas (poner "0" en columna de armas)
- 1 socio (Roger) con 5 armas (buscar en diciembre normalizado)
"""
import openpyxl
from openpyxl.styles import Font, Alignment

# Archivos
anexo_a = "2025-dic-usb-738/CLUB 738-31-DE-DICIEMBRE-2025_ANEXOS A, B Y C.xlsx"
dic_normalizado = "2025-dic-usb-738/CLUB 738-31-DE-DICIEMBRE-2025_RELACION_SOCIOS_ARMAS_NORMALIZADO.xlsx"
archivo_actual = "data/socios/Copy of 2026.31.01_RELACION_SOCIOS_ARMAS_SEPARADO_verified.xlsx"
nueva_fuente = "data/socios/2026_ENERO_RELACION_SOCIOS_ARMAS_MASTER.xlsx"
archivo_salida = "data/socios/2026_ENERO_RELACION_SOCIOS_ARMAS_MASTER_76.xlsx"

print("=" * 80)
print("AGREGANDO 11 SOCIOS FALTANTES A NUEVA FUENTE DE VERDAD")
print("=" * 80)

# 1. LEER ANEXO A para datos de los 11 socios
print("\n📖 Leyendo Anexo A...")
wb_anexo = openpyxl.load_workbook(anexo_a)
ws_anexo = wb_anexo["Anexo A"]

faltantes_curps = [
    " MECR871030HYNNRG05",  # Roger - CON armas
    "AERF781023MDFRMR09",   # Sin armas
    "CAPG891110HYNMCD00",   # Sin armas
    "CECR890104HYNRBL02",   # Sin armas
    "FEPR920403HYNRRC06",   # Sin armas
    "GOAE840623HYNMRD00",   # Sin armas
    "GOMA940118MVZMNM00",   # Sin armas
    "GOXK740906HNERXR09",   # Sin armas
    "PUSJ000131HYNCSSA4",   # Sin armas
    "RODY940625HYNMSL01",   # Sin armas
    "RUES971109HCCDSN07"    # Sin armas
]

socios_anexo = {}
for i in range(7, ws_anexo.max_row + 1):
    curp = ws_anexo.cell(i, 3).value
    if curp and curp.strip() in [c.strip() for c in faltantes_curps]:
        socios_anexo[curp.strip()] = {
            'registro': ws_anexo.cell(i, 1).value,
            'nombre': ws_anexo.cell(i, 2).value,
            'curp': curp.strip(),
            'telefono': ws_anexo.cell(i, 4).value,
            'email': ws_anexo.cell(i, 5).value,
            'num_socio': ws_anexo.cell(i, 6).value,
            'armas_cortas': ws_anexo.cell(i, 7).value or 0,
            'armas_largas': ws_anexo.cell(i, 8).value or 0,
            'fecha_alta': ws_anexo.cell(i, 9).value
        }

print(f"   ✅ {len(socios_anexo)} socios encontrados en Anexo A")

# 2. LEER DICIEMBRE NORMALIZADO para armas de Roger
print("\n📖 Buscando armas de Roger en diciembre normalizado...")
wb_dic = openpyxl.load_workbook(dic_normalizado)
ws_dic = wb_dic.active

armas_roger = []
for i in range(2, ws_dic.max_row + 1):
    curp = ws_dic.cell(i, 4).value  # Col 4 = CURP en normalizado
    if curp and curp.strip() == "MECR871030HYNNRG05":
        arma = {
            'nombre': ws_dic.cell(i, 3).value,
            'curp': curp.strip(),
            'clase': ws_dic.cell(i, 9).value,
            'calibre': ws_dic.cell(i, 10).value,
            'marca': ws_dic.cell(i, 11).value,
            'modelo': ws_dic.cell(i, 12).value,
            'matricula': ws_dic.cell(i, 13).value,
            'folio': ws_dic.cell(i, 14).value
        }
        armas_roger.append(arma)

print(f"   ✅ {len(armas_roger)} armas de Roger encontradas")

# 3. LEER ARCHIVO ACTUAL para direcciones
print("\n📖 Leyendo direcciones del archivo actual...")
wb_actual = openpyxl.load_workbook(archivo_actual)
ws_actual = wb_actual.active

direcciones = {}
for i in range(2, ws_actual.max_row + 1):
    curp = ws_actual.cell(i, 5).value
    if curp:
        direcciones[curp] = {
            'calle': ws_actual.cell(i, 7).value,
            'colonia': ws_actual.cell(i, 8).value,
            'ciudad': ws_actual.cell(i, 9).value,
            'estado': ws_actual.cell(i, 10).value,
            'cp': ws_actual.cell(i, 11).value
        }

# 4. CARGAR NUEVA FUENTE ACTUAL
print("\n📖 Cargando nueva fuente de verdad actual...")
wb_nueva = openpyxl.load_workbook(nueva_fuente)
ws_nueva = wb_nueva.active

fila_actual = ws_nueva.max_row
print(f"   Última fila actual: {fila_actual}")

# 5. AGREGAR LOS 11 SOCIOS
print("\n➕ Agregando socios faltantes...")
socios_agregados = 0
armas_agregadas = 0

for curp_stripped in faltantes_curps:
    curp = curp_stripped.strip()
    
    if curp not in socios_anexo:
        print(f"   ⚠️  {curp} no encontrado en Anexo A")
        continue
    
    socio = socios_anexo[curp]
    
    # CASO 1: Roger Méndez - tiene 5 armas
    if curp == "MECR871030HYNNRG05":
        print(f"\n   📌 {socio['nombre']} - CON {len(armas_roger)} armas")
        
        # Primero, buscar y corregir si ya existe con error
        for i in range(2, fila_actual + 1):
            curp_celda = ws_nueva.cell(i, 5).value
            if curp_celda and '=E248+1' in str(curp_celda):
                print(f"      🔧 Corrigiendo CURP en fila {i}")
                ws_nueva.cell(i, 5, curp)
                ws_nueva.cell(i, 4, socio['nombre'])
        
        # Agregar sus armas
        for arma in armas_roger:
            fila_actual += 1
            
            # Datos base del socio
            ws_nueva.cell(fila_actual, 1, socio['registro'])
            ws_nueva.cell(fila_actual, 2, "CALLE 50 No. 531-E x 69 y 71, CENTRO, 97000 MÉRIDA, YUC.")
            ws_nueva.cell(fila_actual, 3, socio['num_socio'])
            ws_nueva.cell(fila_actual, 4, socio['nombre'])
            ws_nueva.cell(fila_actual, 5, curp)
            ws_nueva.cell(fila_actual, 6, socio['num_socio'])
            
            # Dirección estructurada (si existe)
            if curp in direcciones:
                dir_data = direcciones[curp]
                ws_nueva.cell(fila_actual, 7, dir_data['calle'])
                ws_nueva.cell(fila_actual, 8, dir_data['colonia'])
                ws_nueva.cell(fila_actual, 9, dir_data['ciudad'])
                ws_nueva.cell(fila_actual, 10, dir_data['estado'])
                ws_nueva.cell(fila_actual, 11, dir_data['cp'])
            
            # Datos del arma
            ws_nueva.cell(fila_actual, 12, arma['clase'])
            ws_nueva.cell(fila_actual, 13, arma['calibre'])
            ws_nueva.cell(fila_actual, 14, arma['marca'])
            ws_nueva.cell(fila_actual, 15, arma['modelo'])
            ws_nueva.cell(fila_actual, 16, arma['matricula'])
            ws_nueva.cell(fila_actual, 17, arma['folio'])
            
            armas_agregadas += 1
        
        socios_agregados += 1
    
    # CASO 2: Socios SIN armas - 1 fila con "0"
    else:
        print(f"\n   📌 {socio['nombre']} - SIN armas (0)")
        fila_actual += 1
        
        # Datos base del socio
        ws_nueva.cell(fila_actual, 1, socio['registro'])
        ws_nueva.cell(fila_actual, 2, "CALLE 50 No. 531-E x 69 y 71, CENTRO, 97000 MÉRIDA, YUC.")
        ws_nueva.cell(fila_actual, 3, socio['num_socio'])
        ws_nueva.cell(fila_actual, 4, socio['nombre'])
        ws_nueva.cell(fila_actual, 5, curp)
        ws_nueva.cell(fila_actual, 6, socio['num_socio'])
        
        # Dirección estructurada (si existe)
        if curp in direcciones:
            dir_data = direcciones[curp]
            ws_nueva.cell(fila_actual, 7, dir_data['calle'])
            ws_nueva.cell(fila_actual, 8, dir_data['colonia'])
            ws_nueva.cell(fila_actual, 9, dir_data['ciudad'])
            ws_nueva.cell(fila_actual, 10, dir_data['estado'])
            ws_nueva.cell(fila_actual, 11, dir_data['cp'])
        
        # Columnas de armas: "0" o vacío
        ws_nueva.cell(fila_actual, 12, "0")  # CLASE = "0"
        # Dejar vacías las demás columnas de arma
        
        socios_agregados += 1

# 6. GUARDAR
print(f"\n💾 Guardando: {archivo_salida}")
wb_nueva.save(archivo_salida)

print("\n" + "=" * 80)
print("✅ NUEVA FUENTE DE VERDAD ACTUALIZADA")
print("=" * 80)
print(f"👥 Socios agregados: {socios_agregados}/11")
print(f"🔫 Armas agregadas: {armas_agregadas}")
print(f"📊 Total filas: {ws_nueva.max_row}")
print(f"\n📂 Archivo guardado: {archivo_salida}")

wb_anexo.close()
wb_dic.close()
wb_actual.close()
wb_nueva.close()
