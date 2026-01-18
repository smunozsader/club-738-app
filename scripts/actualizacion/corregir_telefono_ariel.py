#!/usr/bin/env python3
"""
Corregir teléfono de Ariel Baltazar Córdoba Wilson y regenerar fuente de verdad
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import shutil

# Archivos
anexo_a_corregido = "2025-dic-usb-738/CLUB 738-31-DE-DICIEMBRE-2025_ANEXOS A, B Y C_CORREGIDO.xlsx"
anexo_a_final = "2025-dic-usb-738/CLUB 738-31-DE-DICIEMBRE-2025_ANEXOS A, B Y C_CORREGIDO_FINAL.xlsx"

print("=" * 80)
print("CORRIGIENDO TELÉFONO DE ARIEL BALTAZAR CÓRDOBA WILSON")
print("=" * 80)

# Leer anexo corregido
wb_anexo = openpyxl.load_workbook(anexo_a_corregido)
ws_anexo = wb_anexo["Anexo A"]

# Buscar y corregir Ariel Baltazar Córdoba Wilson
print("\n🔍 Buscando Ariel Baltazar Córdoba Wilson...")
for i in range(7, ws_anexo.max_row + 1):
    nombre = ws_anexo.cell(i, 2).value
    if nombre and "ARIEL BALTAZAR" in str(nombre).upper() and "CÓRDOBA" in str(nombre).upper():
        print(f"\n✅ Encontrado en fila {i}:")
        print(f"   Nombre: {nombre}")
        print(f"   CURP: {ws_anexo.cell(i, 3).value}")
        print(f"   Teléfono anterior: {ws_anexo.cell(i, 4).value}")
        
        # Corregir teléfono
        ws_anexo.cell(i, 4, "+52 999 200 3314")
        print(f"   Teléfono nuevo: +52 999 200 3314")
        break

# Guardar anexo corregido
print(f"\n💾 Guardando: {anexo_a_final}")
wb_anexo.save(anexo_a_final)

# REGENERAR FUENTE DE VERDAD
print("\n" + "=" * 80)
print("REGENERANDO FUENTE DE VERDAD COMPLETA")
print("=" * 80)

dic_normalizado = "2025-dic-usb-738/CLUB 738-31-DE-DICIEMBRE-2025_RELACION_SOCIOS_ARMAS_NORMALIZADO.xlsx"
archivo_actual = "data/socios/Copy of 2026.31.01_RELACION_SOCIOS_ARMAS_SEPARADO_verified.xlsx"
archivo_salida = "data/socios/2026_ENERO_FUENTE_VERDAD_COMPLETA_76_SOCIOS.xlsx"

# Leer Anexo A FINAL
print("\n📋 Leyendo Anexo A final...")
socios_anexo = {}
for i in range(7, ws_anexo.max_row + 1):
    curp = ws_anexo.cell(i, 3).value
    nombre = ws_anexo.cell(i, 2).value
    
    if curp and nombre and "TOTAL" not in str(nombre).upper():
        curp_clean = str(curp).strip()
        socios_anexo[curp_clean] = {
            'registro': ws_anexo.cell(i, 1).value,
            'nombre': ws_anexo.cell(i, 2).value,
            'curp': curp_clean,
            'telefono': ws_anexo.cell(i, 4).value,
            'email': ws_anexo.cell(i, 5).value,
            'num_credencial': ws_anexo.cell(i, 6).value,
            'armas_cortas': int(ws_anexo.cell(i, 7).value or 0),
            'armas_largas': int(ws_anexo.cell(i, 8).value or 0),
            'fecha_alta': ws_anexo.cell(i, 9).value
        }

print(f"   ✅ {len(socios_anexo)} socios")

# Leer armas
print("\n📋 Leyendo armas...")
wb_dic = openpyxl.load_workbook(dic_normalizado)
ws_dic = wb_dic.active

armas_por_curp = {}
for i in range(2, ws_dic.max_row + 1):
    curp = ws_dic.cell(i, 4).value
    if curp:
        curp_clean = str(curp).strip()
        if curp_clean not in armas_por_curp:
            armas_por_curp[curp_clean] = []
        
        arma = {
            'clase': ws_dic.cell(i, 9).value,
            'calibre': ws_dic.cell(i, 10).value,
            'marca': ws_dic.cell(i, 11).value,
            'modelo': ws_dic.cell(i, 12).value,
            'matricula': ws_dic.cell(i, 13).value,
            'folio': ws_dic.cell(i, 14).value
        }
        armas_por_curp[curp_clean].append(arma)

print(f"   ✅ {sum(len(v) for v in armas_por_curp.values())} armas")

# Leer direcciones
print("\n📋 Leyendo direcciones...")
wb_actual = openpyxl.load_workbook(archivo_actual)
ws_actual = wb_actual.active

direcciones = {}
for i in range(2, ws_actual.max_row + 1):
    curp = ws_actual.cell(i, 5).value
    if curp:
        direcciones[curp] = {
            'calle': ws_actual.cell(i, 7).value,
            'colonia': ws_actual.cell(i, 8).value,
            'ciudad': ws_actual.cell(i, 9).value,
            'estado': ws_actual.cell(i, 10).value,
            'cp': ws_actual.cell(i, 11).value
        }

print(f"   ✅ {len(direcciones)} direcciones")

# CREAR NUEVO WORKBOOK
print("\n📝 Creando nueva fuente de verdad...")
wb_nuevo = openpyxl.Workbook()
ws_nuevo = wb_nuevo.active
ws_nuevo.title = "SOCIOS_ENERO_2026"

# HEADERS
headers = [
    "No. REGISTRO", "DOMICILIO CLUB", "No. CREDENCIAL", "NOMBRE SOCIO", "CURP",
    "TELEFONO", "EMAIL", "FECHA ALTA", "CALLE", "COLONIA", "CIUDAD", "ESTADO",
    "CP", "CLASE", "CALIBRE", "MARCA", "MODELO", "MATRÍCULA", "FOLIO"
]

for j, header in enumerate(headers, 1):
    cell = ws_nuevo.cell(1, j, header)
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# AGREGAR DATOS
fila = 2
total_armas = 0
socios_con_armas = 0
socios_sin_armas = 0

for curp, datos_socio in sorted(socios_anexo.items(), key=lambda x: x[1]['num_credencial'] or 0):
    armas = armas_por_curp.get(curp, [])
    
    if armas:
        socios_con_armas += 1
        for arma in armas:
            ws_nuevo.cell(fila, 1, datos_socio['registro'])
            ws_nuevo.cell(fila, 2, "CALLE 50 No. 531-E x 69 y 71, CENTRO, 97000 MÉRIDA, YUC.")
            ws_nuevo.cell(fila, 3, datos_socio['num_credencial'])
            ws_nuevo.cell(fila, 4, datos_socio['nombre'])
            ws_nuevo.cell(fila, 5, curp)
            ws_nuevo.cell(fila, 6, datos_socio['telefono'])
            ws_nuevo.cell(fila, 7, datos_socio['email'])
            ws_nuevo.cell(fila, 8, datos_socio['fecha_alta'])
            
            if curp in direcciones:
                dir_data = direcciones[curp]
                ws_nuevo.cell(fila, 9, dir_data['calle'])
                ws_nuevo.cell(fila, 10, dir_data['colonia'])
                ws_nuevo.cell(fila, 11, dir_data['ciudad'])
                ws_nuevo.cell(fila, 12, dir_data['estado'])
                ws_nuevo.cell(fila, 13, dir_data['cp'])
            
            ws_nuevo.cell(fila, 14, arma['clase'])
            ws_nuevo.cell(fila, 15, arma['calibre'])
            ws_nuevo.cell(fila, 16, arma['marca'])
            ws_nuevo.cell(fila, 17, arma['modelo'])
            ws_nuevo.cell(fila, 18, arma['matricula'])
            ws_nuevo.cell(fila, 19, arma['folio'])
            
            total_armas += 1
            fila += 1
    else:
        socios_sin_armas += 1
        ws_nuevo.cell(fila, 1, datos_socio['registro'])
        ws_nuevo.cell(fila, 2, "CALLE 50 No. 531-E x 69 y 71, CENTRO, 97000 MÉRIDA, YUC.")
        ws_nuevo.cell(fila, 3, datos_socio['num_credencial'])
        ws_nuevo.cell(fila, 4, datos_socio['nombre'])
        ws_nuevo.cell(fila, 5, curp)
        ws_nuevo.cell(fila, 6, datos_socio['telefono'])
        ws_nuevo.cell(fila, 7, datos_socio['email'])
        ws_nuevo.cell(fila, 8, datos_socio['fecha_alta'])
        
        if curp in direcciones:
            dir_data = direcciones[curp]
            ws_nuevo.cell(fila, 9, dir_data['calle'])
            ws_nuevo.cell(fila, 10, dir_data['colonia'])
            ws_nuevo.cell(fila, 11, dir_data['ciudad'])
            ws_nuevo.cell(fila, 12, dir_data['estado'])
            ws_nuevo.cell(fila, 13, dir_data['cp'])
        
        ws_nuevo.cell(fila, 14, "0")
        fila += 1

# Ajustar anchos
anchos = {
    'A': 12, 'B': 50, 'C': 12, 'D': 40, 'E': 20, 'F': 15,
    'G': 35, 'H': 20, 'I': 35, 'J': 25, 'K': 20, 'L': 20,
    'M': 10, 'N': 15, 'O': 12, 'P': 20, 'Q': 20, 'R': 20, 'S': 15
}

for col, ancho in anchos.items():
    ws_nuevo.column_dimensions[col].width = ancho

print(f"\n💾 Guardando: {archivo_salida}")
wb_nuevo.save(archivo_salida)

print("\n" + "=" * 80)
print("✅ CORRECCIÓN COMPLETADA")
print("=" * 80)
print(f"👥 Total socios: {len(socios_anexo)}")
print(f"   - Con armas: {socios_con_armas}")
print(f"   - Sin armas: {socios_sin_armas}")
print(f"🔫 Total armas: {total_armas}")
print(f"\n📂 Archivos:")
print(f"   Anexo A final: {anexo_a_final}")
print(f"   Fuente verdad: {archivo_salida}")
print(f"\n✨ Correcciones totales:")
print(f"   ✅ Agustín Moreno → Email: agus_tin1_@hotmail.com, Tel: +52 999 278 0476")
print(f"   ✅ Ariel B. Córdoba Wilson → Tel: +52 999 200 3314")

wb_anexo.close()
wb_dic.close()
wb_actual.close()
wb_nuevo.close()
