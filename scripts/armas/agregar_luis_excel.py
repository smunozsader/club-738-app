#!/usr/bin/env python3

import openpyxl
from datetime import datetime

archivo = '/Applications/club-738-web/socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx'
wb = openpyxl.load_workbook(archivo)
ws = wb.active

# Obtener headers
headers = [cell.value for cell in ws[1]]

print("=" * 100)
print("AGREGANDO SOCIO A FUENTE DE VERDAD")
print("=" * 100)

print(f"\nHeaders en Excel: {len(headers)} columnas")
for i, h in enumerate(headers):
    print(f"  [{i}] {h}")

# Crear datos para la nueva fila (Luis Fernando)
datos_luis = {
    'No. REGISTRO': 738,
    'DOMICILIO CLUB': 'CALLE 50 No. 531-E x 69 y 71, CENTRO, 97000 MÉRIDA, YUC.',
    'No. CREDENCIAL': 236,
    'NOMBRE SOCIO': 'LUIS FERNANDO GUILLERMO GAMBOA',
    'CURP': 'GUGL750204HYNLMS04',
    'TELEFONO': '9992420621',
    'EMAIL': 'oso.guigam@gmail.com',
    'FECHA ALTA': datetime(2026, 1, 8),
    'CALLE': 'Calle 32 x 9 Cedro, Tablaje 23222',
    'COLONIA': 'Loc. Tixcuytun',
    'CIUDAD': 'Mérida',
    'ESTADO': 'YUCATÁN',
    'CP': '97305',
    'CLASE': 0,  # Sin armas
    'CALIBRE': None,
    'MARCA': None,
    'MODELO': None,
    'MATRÍCULA': None,
    'FOLIO': None
}

# Crear nueva fila con los datos
nueva_fila = []
for header in headers:
    valor = datos_luis.get(header, None)
    nueva_fila.append(valor)

# Agregar la fila al final del Excel
ws.append(nueva_fila)

print(f"\n✅ Fila agregada para:")
print(f"   Credencial: 236")
print(f"   Nombre: LUIS FERNANDO GUILLERMO GAMBOA")
print(f"   Email: oso.guigam@gmail.com")
print(f"   CURP: GUGL750204HYNLMS04")
print(f"   Domicilio: Calle 32 x 9 Cedro, Tablaje 23222, Loc. Tixcuytun, Mérida, YUCATÁN 97305")
print(f"   Fecha de Alta: 2026-01-08")
print(f"   Armas: NO (CLASE = 0)")

# Guardar el archivo
wb.save(archivo)

print(f"\n✅ FUENTE DE VERDAD ACTUALIZADA")
print(f"   Archivo: {archivo}")
print(f"   Total de socios: {ws.max_row - 1}")

print("\n" + "=" * 100)
