#!/usr/bin/env python3

import openpyxl
from datetime import datetime

archivo = '/Applications/club-738-web/socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx'
wb = openpyxl.load_workbook(archivo)
ws = wb.active

# Obtener headers
headers = [cell.value for cell in ws[1]]

print("=" * 100)
print("AGREGANDO LUIS FERNANDO GUILLERMO GAMBOA AL EXCEL")
print("=" * 100)

# Mapear índices de columnas
indice_credencial = headers.index('No. CREDENCIAL') if 'No. CREDENCIAL' in headers else 2
indice_nombre = headers.index('NOMBRE SOCIO') if 'NOMBRE SOCIO' in headers else 3
indice_curp = headers.index('CURP') if 'CURP' in headers else 4
indice_domicilio = headers.index('DOMICILIO CLUB') if 'DOMICILIO CLUB' in headers else 1
indice_telefono = headers.index('TELEFONO') if 'TELEFONO' in headers else 5
indice_email = headers.index('EMAIL') if 'EMAIL' in headers else 6
indice_fecha_alta = headers.index('FECHA ALTA') if 'FECHA ALTA' in headers else 7
indice_clase = headers.index('CLASE') if 'CLASE' in headers else 8

print(f"\nÍndices de columnas:")
print(f"  Credencial: {indice_credencial} ({headers[indice_credencial]})")
print(f"  Nombre: {indice_nombre} ({headers[indice_nombre]})")
print(f"  CURP: {indice_curp} ({headers[indice_curp]})")
print(f"  Domicilio: {indice_domicilio} ({headers[indice_domicilio]})")
print(f"  Teléfono: {indice_telefono} ({headers[indice_telefono]})")
print(f"  Email: {indice_email} ({headers[indice_email]})")
print(f"  Fecha Alta: {indice_fecha_alta} ({headers[indice_fecha_alta]})")
print(f"  Clase: {indice_clase} ({headers[indice_clase]})")

# Obtener última fila
ultima_fila_numero = ws.max_row
nueva_fila_numero = ultima_fila_numero + 1

print(f"\nÚltima fila: {ultima_fila_numero}")
print(f"Nueva fila será: {nueva_fila_numero}")

# Crear lista con los datos
nueva_fila_datos = [None] * len(headers)
nueva_fila_datos[indice_credencial] = 236
nueva_fila_datos[indice_nombre] = "LUIS FERNANDO GUILLERMO GAMBOA"
nueva_fila_datos[indice_curp] = "GUGL750204HYNLMS04"
nueva_fila_datos[indice_domicilio] = "Calle 32 x 9 Cedro, Tablaje 23222, Loc. Tixcuytun, Mérida, YUCATÁN CP97305"
nueva_fila_datos[indice_telefono] = "9992420621"
nueva_fila_datos[indice_email] = "oso.guigam@gmail.com"
nueva_fila_datos[indice_fecha_alta] = datetime(2026, 1, 8)  # Fecha de alta en Firestore
nueva_fila_datos[indice_clase] = 0  # Sin armas

# Agregar la fila al worksheet
for col_idx, valor in enumerate(nueva_fila_datos, 1):
    ws.cell(row=nueva_fila_numero, column=col_idx, value=valor)

# Guardar el archivo
wb.save(archivo)

print("\n" + "=" * 100)
print("✅ AGREGADO EXITOSAMENTE")
print("=" * 100)
print(f"\nCredencial: 236")
print(f"Nombre: LUIS FERNANDO GUILLERMO GAMBOA")
print(f"Email: oso.guigam@gmail.com")
print(f"CURP: GUGL750204HYNLMS04")
print(f"Teléfono: 9992420621")
print(f"Domicilio: Calle 32 x 9 Cedro, Tablaje 23222, Loc. Tixcuytun, Mérida, YUCATÁN CP97305")
print(f"Fecha Alta: 2026-01-08")
print(f"Clase (Armas): 0 (SIN ARMAS)")
print(f"\nFila agregada en: {nueva_fila_numero}")
print(f"Archivo guardado: {archivo}")
print("=" * 100)
