#!/usr/bin/env python3
"""
CORRECCIÓN CRÍTICA: CZ P-10 C de Ricardo Soberanis
Calibre CORRECTO: .380" ACP (NO .40 S&W)
Cumplimiento SEDENA: Art. 50 LFAFE - máximo permitido para civiles es .380 ACP
"""

import openpyxl
from openpyxl.utils import get_column_letter
import json

# Rutas
EXCEL_PATH = "/Applications/club-738-web/socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx"

# Datos CORRECTOS del arma
DATOS_CORRECTOS = {
    "marca": "CESKA ZBROJOVKA",
    "modelo": "CZ P-10 C", 
    "calibre": ".380\"",  # ✅ CORRECTO - NO .40 S&W
    "matricula": "EP29710",
    "folio": "A3912487",
    "clase": "PISTOLA"
}

print("=" * 80)
print("🚨 CORRECCIÓN CRÍTICA - CZ P-10 C DE RICARDO SOBERANIS")
print("=" * 80)
print("\n⚠️  ERROR ANTERIOR: Calibre registrado como .40 S&W (ILEGAL para civiles)")
print("✅ CALIBRE CORRECTO: .380\" ACP (máximo permitido por SEDENA Art. 50)")
print()

# 1. CORREGIR EXCEL
print("1️⃣ CORRIGIENDO EXCEL (fila 283)...")
try:
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    
    # Buscar fila con Ricardo Soberanis y EP29710
    for row_num in range(2, ws.max_row + 1):
        row_data = [cell.value for cell in ws[row_num]]
        
        # Buscar por matrícula EP29710
        matricula_col = None
        for col_idx, cell_value in enumerate(row_data):
            if cell_value == "EP29710":
                matricula_col = col_idx + 1
                break
        
        if matricula_col:
            print(f"   ✅ Encontrada fila {row_num} con matrícula EP29710")
            
            # Buscar columnas por encabezado (fila 1)
            headers = [cell.value for cell in ws[1]]
            
            # Corregir MARCA
            if "marca" in [str(h).lower() if h else "" for h in headers]:
                marca_col = [i+1 for i, h in enumerate(headers) if h and "marca" in str(h).lower()][0]
                ws[f'{get_column_letter(marca_col)}{row_num}'] = DATOS_CORRECTOS["marca"]
                print(f"   ✏️  Marca → {DATOS_CORRECTOS['marca']}")
            
            # Corregir CALIBRE (crítico)
            if "calibre" in [str(h).lower() if h else "" for h in headers]:
                calibre_col = [i+1 for i, h in enumerate(headers) if h and "calibre" in str(h).lower()][0]
                ws[f'{get_column_letter(calibre_col)}{row_num}'] = DATOS_CORRECTOS["calibre"]
                print(f"   ✏️  Calibre → {DATOS_CORRECTOS['calibre']} (✅ CORRECTO)")
            
            break
    
    # Guardar Excel corregido
    wb.save(EXCEL_PATH)
    print("   ✅ Excel guardado con correcciones")
    
except Exception as e:
    print(f"   ❌ Error al corregir Excel: {e}")
    exit(1)

# 2. INSTRUCCIONES PARA FIRESTORE
print("\n2️⃣ CORRECCIONES PENDIENTES EN FIRESTORE (ID: 8d1f8140):")
print("   Documento: socios/rsoberanis11@hotmail.com/armas/8d1f8140")
print("   Campos a actualizar:")
print(f"     • calibre: '{DATOS_CORRECTOS['calibre']}'")
print(f"     • marca: '{DATOS_CORRECTOS['marca']}'")

print("\n" + "=" * 80)
print("✅ EXCEL CORREGIDO")
print("=" * 80)
print("""
ACCIÓN REQUERIDA:
- Actualizar Firestore manualmente o con Firebase CLI
- Comando:
  npx firebase firestore set-document \
    "socios/rsoberanis11@hotmail.com/armas/8d1f8140" \
    "calibre=.380\\"" \
    "marca=CESKA ZBROJOVKA"
""")
