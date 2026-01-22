#!/usr/bin/env python3
"""
Corrección de transferencias de armas entre Gardoni y Arechiga
Basado en commit d5c3733: "v1.23.0 - Sincronización completa Excel + Transferencias de arsenal"
"""

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

EXCEL_PATH = "/Applications/club-738-web/socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx"

# Datos de los socios
GARDONI = {
    "nombre": "JOAQUIN RODOLFO GARDONI NUÑEZ",
    "email": "jrgardoni@gmail.com",
    "credencial": 199,
    "armas_correctas": [
        {"clase": "PISTOLA", "calibre": ".22\" L.R.", "marca": "GRAND POWER", "modelo": "K22 X-TRIM", "matricula": "K078928", "folio": "A 3601944"},
        {"clase": "RIFLE", "calibre": ".22\" L.R.", "marca": "KRISS", "modelo": "DMK22C", "matricula": "22C002369", "folio": "3722287"},
        {"clase": "RIFLE", "calibre": ".22\" L.R.", "marca": "RUGER", "modelo": "10/22", "matricula": "0008-32069", "folio": "A 3722288"},
        {"clase": "RIFLE", "calibre": ".22\" L.R.", "marca": "RUGER", "modelo": "10/22", "matricula": "0013-82505", "folio": "A 3605099"},
        {"clase": "PISTOLA", "calibre": ".380\"", "marca": "CESKA ZBROJOVKA", "modelo": "CZ SHADOW 2", "matricula": "DP25086", "folio": "A 3782099"},
        {"clase": "PISTOLA", "calibre": ".380\"", "marca": "CESKA ZBROJOVKA", "modelo": "CZ SHADOW 2", "matricula": "DP25246", "folio": "A 3792515"},
        {"clase": "PISTOLA", "calibre": ".380\"", "marca": "CESKA ZBROJOVKA", "modelo": "CZ SHADOW 2", "matricula": "DP25087", "folio": "A 3782098"},
    ]
}

ARECHIGA = {
    "nombre": "MARIA FERNANDA GUADALUPE ARECHIGA RAMOS",
    "email": "arechiga@jogarplastics.com",
    "credencial": 212,
    "armas_correctas": [
        {"clase": "PISTOLA", "calibre": ".380\" AUTO", "marca": "GRAND POWER", "modelo": "LP380", "matricula": "K084328", "folio": "A3714371"},
        {"clase": "PISTOLA", "calibre": ".380\" AUTO", "marca": "GRAND POWER", "modelo": "LP380", "matricula": "K078999", "folio": "A 3601943"},
        {"clase": "PISTOLA", "calibre": ".380\"", "marca": "CESKA ZBROJOVKA", "modelo": "CZ P-07", "matricula": "C647155", "folio": "B 611940"},
    ]
}

print("=" * 100)
print("🔄 CORRECCIÓN DE TRANSFERENCIAS: GARDONI ↔ ARECHIGA")
print("=" * 100)
print(f"\nBasado en commit d5c3733 - v1.23.0")
print(f"\nGARDONI debe tener: {len(GARDONI['armas_correctas'])} armas")
print(f"ARECHIGA debe tener: {len(ARECHIGA['armas_correctas'])} armas")

# 1. Cargar Excel
print("\n1️⃣ CARGANDO EXCEL...")
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

# Obtener encabezados
headers = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}

print(f"   ✅ Excel cargado: {EXCEL_PATH}")

# 2. Limpiar filas actuales de ambos socios
print("\n2️⃣ LIMPIANDO FILAS ACTUALES...")

rows_to_delete = []
for row_num in range(2, ws.max_row + 1):
    nombre_cell = ws[f'D{row_num}'].value  # Columna NOMBRE SOCIO
    if nombre_cell:
        nombre_str = str(nombre_cell).upper()
        if "GARDONI" in nombre_str or "ARECHIGA" in nombre_str:
            rows_to_delete.append(row_num)

# Eliminar filas en orden inverso para no afectar índices
for row_num in sorted(rows_to_delete, reverse=True):
    ws.delete_rows(row_num)
    print(f"   ❌ Eliminada fila {row_num}")

print(f"   ✅ Total de filas eliminadas: {len(rows_to_delete)}")

# 3. Agregar filas correctas de Gardoni
print("\n3️⃣ AGREGANDO {len(GARDONI['armas_correctas'])} ARMAS DE GARDONI...")

nueva_fila = ws.max_row + 1
for arma in GARDONI['armas_correctas']:
    ws[f'C{nueva_fila}'] = GARDONI['credencial']
    ws[f'D{nueva_fila}'] = GARDONI['nombre']
    ws[f'E{nueva_fila}'] = GARDONI['email']
    ws[f'N{nueva_fila}'] = arma['clase']
    ws[f'O{nueva_fila}'] = arma['calibre']
    ws[f'P{nueva_fila}'] = arma['marca']
    ws[f'Q{nueva_fila}'] = arma['modelo']
    ws[f'R{nueva_fila}'] = arma['matricula']
    ws[f'S{nueva_fila}'] = arma['folio']
    
    print(f"   ✅ Fila {nueva_fila}: {arma['marca']} {arma['modelo']} ({arma['calibre']})")
    nueva_fila += 1

# 4. Agregar filas correctas de Arechiga
print("\n4️⃣ AGREGANDO {len(ARECHIGA['armas_correctas'])} ARMAS DE ARECHIGA...")

for arma in ARECHIGA['armas_correctas']:
    ws[f'C{nueva_fila}'] = ARECHIGA['credencial']
    ws[f'D{nueva_fila}'] = ARECHIGA['nombre']
    ws[f'E{nueva_fila}'] = ARECHIGA['email']
    ws[f'N{nueva_fila}'] = arma['clase']
    ws[f'O{nueva_fila}'] = arma['calibre']
    ws[f'P{nueva_fila}'] = arma['marca']
    ws[f'Q{nueva_fila}'] = arma['modelo']
    ws[f'R{nueva_fila}'] = arma['matricula']
    ws[f'S{nueva_fila}'] = arma['folio']
    
    print(f"   ✅ Fila {nueva_fila}: {arma['marca']} {arma['modelo']} ({arma['calibre']})")
    nueva_fila += 1

# 5. Guardar Excel
print("\n5️⃣ GUARDANDO EXCEL...")
wb.save(EXCEL_PATH)
print(f"   ✅ Excel guardado: {EXCEL_PATH}")

# 6. Verificar con pandas
print("\n6️⃣ VERIFICACIÓN...")
df = pd.read_excel(EXCEL_PATH)

gardoni_check = df[df['NOMBRE SOCIO'].astype(str).str.upper().str.contains('GARDONI', na=False)]
arechiga_check = df[df['NOMBRE SOCIO'].astype(str).str.upper().str.contains('ARECHIGA', na=False)]

print(f"\n   GARDONI en Excel: {len(gardoni_check)} armas")
print(f"   ARECHIGA en Excel: {len(arechiga_check)} armas")

if len(gardoni_check) == len(GARDONI['armas_correctas']) and len(arechiga_check) == len(ARECHIGA['armas_correctas']):
    print("\n   ✅ EXCEL CORREGIDO CORRECTAMENTE")
else:
    print("\n   ⚠️  VERIFICAR MANUALMENTE")

print("\n" + "=" * 100)
print("📝 ACCIÓN PENDIENTE: Actualizar Firestore")
print("=" * 100)
print(f"""
Necesitas actualizar Firestore para:

GARDONI (jrgardoni@gmail.com):
- ELIMINAR: K078999 (GRAND POWER LP380), K084328 (GRAND POWER LP380)
- AGREGAR: DP25087 (CZ SHADOW 2)

ARECHIGA (arechiga@jogarplastics.com):
- AGREGAR: K078999 (GRAND POWER LP380), K084328 (GRAND POWER LP380)
- AGREGAR: C647155 (CZ P-07)

Usa Firebase Console o firebase CLI para actualizar las colecciones:
- socios/jrgardoni@gmail.com/armas/
- socios/arechiga@jogarplastics.com/armas/
""")
