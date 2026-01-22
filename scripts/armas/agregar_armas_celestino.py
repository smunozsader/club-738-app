#!/usr/bin/env python3
"""
Agregar 4 armas nuevas al arsenal de Celestino Sánchez Fernández (credencial 183)
"""

import openpyxl
import pandas as pd

EXCEL_PATH = "/Applications/club-738-web/socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx"

# Datos del socio
SOCIO = {
    "credencial": 183,
    "nombre": "CELESTINO SANCHEZ FERNANDEZ",
    "email": "tinosanchezf@yahoo.com.mx"
}

# Armas nuevas a agregar
ARMAS_NUEVAS = [
    {"clase": "RIFLE", "calibre": ".22\" L.R.", "marca": "WINCHESTER", "modelo": "9422", "matricula": "F11281", "folio": "A3917581"},
    {"clase": "PISTOLA", "calibre": ".380\" AUTO", "marca": "CESKA ZBROJOVKA", "modelo": "CZ P-07", "matricula": "D207727", "folio": "A3747924"},
    {"clase": "PISTOLA", "calibre": ".380\" AUTO", "marca": "CESKA ZBROJOVKA", "modelo": "CZ P-10 C", "matricula": "CP18665", "folio": "A3747922"},
    {"clase": "PISTOLA", "calibre": ".380\" AUTO", "marca": "SIG SAUER", "modelo": "P250", "matricula": "57C048858", "folio": "B596607"}
]

print("=" * 120)
print(f"🔫 AGREGAR {len(ARMAS_NUEVAS)} ARMAS A CELESTINO SÁNCHEZ (Credencial {SOCIO['credencial']})")
print("=" * 120)

# 1. Cargar Excel
print("\n1️⃣ CARGANDO EXCEL...")
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

# Encontrar última fila ocupada
ultima_fila = ws.max_row

print(f"   ✅ Excel cargado: {ultima_fila} filas")

# 2. Agregar nuevas filas
print(f"\n2️⃣ AGREGANDO {len(ARMAS_NUEVAS)} ARMAS...")

nueva_fila = ultima_fila + 1
for arma in ARMAS_NUEVAS:
    ws[f'C{nueva_fila}'] = SOCIO['credencial']
    ws[f'D{nueva_fila}'] = SOCIO['nombre']
    ws[f'G{nueva_fila}'] = SOCIO['email']
    ws[f'N{nueva_fila}'] = arma['clase']
    ws[f'O{nueva_fila}'] = arma['calibre']
    ws[f'P{nueva_fila}'] = arma['marca']
    ws[f'Q{nueva_fila}'] = arma['modelo']
    ws[f'R{nueva_fila}'] = arma['matricula']
    ws[f'S{nueva_fila}'] = arma['folio']
    
    print(f"   ✅ Fila {nueva_fila}: {arma['marca']} {arma['modelo']} ({arma['calibre']})")
    nueva_fila += 1

# 3. Guardar Excel
print("\n3️⃣ GUARDANDO EXCEL...")
wb.save(EXCEL_PATH)
print(f"   ✅ Excel guardado")

# 4. Verificar con pandas
print("\n4️⃣ VERIFICACIÓN...")
df = pd.read_excel(EXCEL_PATH)
celestino_check = df[(df['No. CREDENCIAL'] == 183)]
print(f"\n   CELESTINO en Excel: {len(celestino_check)} armas (antes: 1, ahora debe ser: 5)")

print("\n" + "=" * 120)
print("✅ EXCEL ACTUALIZADO - Próximo: Actualizar Firestore")
print("=" * 120)
