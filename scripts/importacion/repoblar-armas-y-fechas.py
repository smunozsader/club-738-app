#!/usr/bin/env python3

"""
Script: Repoblar ARMAS y FECHA DE INGRESO desde Excel maestro

Fuente: /Applications/club-738-web/data/socios/2025.31.12_RELACION_SOCIOS_ARMAS.xlsx

Acciones:
1. Lee el Excel (CLUB 738. RELACION SOCIOS + Anexo A)
2. Limpia socios/{email}/armas/{armaId}
3. Repuebla con datos correctos
4. Actualiza socios/{email}.fechaAlta

Uso:
   python3 scripts/repoblar-armas-y-fechas.py
"""

import openpyxl
import os
import sys
from pathlib import Path
from datetime import datetime
import firebase_admin
from firebase_admin import credentials, firestore
import uuid

# ============================================================================
# INICIALIZACIÓN FIREBASE
# ============================================================================

script_dir = Path(__file__).parent.absolute()
service_account_path = script_dir / 'serviceAccountKey.json'

if not service_account_path.exists():
    print(f"❌ ERROR: No encontrado {service_account_path}")
    sys.exit(1)

cred = credentials.Certificate(str(service_account_path))
firebase_admin.initialize_app(cred)
db = firestore.client()

# ============================================================================
# LEER EXCEL
# ============================================================================

def leer_excel():
    """Lee el Excel y retorna armas por socio y fechas de ingreso."""
    print('📖 Leyendo Excel...')
    
    excel_path = '/Applications/club-738-web/data/socios/2025.31.12_RELACION_SOCIOS_ARMAS.xlsx'
    wb = openpyxl.load_workbook(excel_path)
    
    # Leer hoja principal
    ws_main = wb['CLUB 738. RELACION SOCIOS 31 DI']
    
    # Leer Anexo A para fechas
    ws_anexo = wb['Anexo A']
    
    # Crear mapa de fechas: email -> fecha
    fechas_map = {}
    for row_num, row in enumerate(ws_anexo.iter_rows(min_row=2, max_row=ws_anexo.max_row, values_only=False), 2):
        email_cell = row[4]  # Columna E (INDEX 4)
        fecha_cell = row[8]  # Columna I (INDEX 8)
        
        email = email_cell.value
        fecha = fecha_cell.value
        
        if email and fecha:
            email_lower = str(email).lower().strip()
            # Si es datetime, convertir a fecha
            if isinstance(fecha, datetime):
                fecha_obj = fecha
            else:
                try:
                    fecha_obj = datetime.fromisoformat(str(fecha).split()[0])
                except:
                    fecha_obj = None
            
            if fecha_obj:
                fechas_map[email_lower] = fecha_obj
    
    print(f'✅ Fechas cargadas: {len(fechas_map)} socios')
    
    # Agrupar armas por socio
    armas_por_socio = {}
    for row_num, row in enumerate(ws_main.iter_rows(min_row=2, max_row=ws_main.max_row, values_only=False), 2):
        # Extraer datos
        email_cell = row[7]  # Columna H (INDEX 7) - EMAIL
        clase_cell = row[8]  # Columna I (INDEX 8) - CLASE
        calibre_cell = row[9]  # Columna J (INDEX 9) - CALIBRE
        marca_cell = row[10]  # Columna K (INDEX 10) - MARCA
        modelo_cell = row[11]  # Columna L (INDEX 11) - MODELO
        matricula_cell = row[12]  # Columna M (INDEX 12) - MATRÍCULA
        folio_cell = row[13]  # Columna N (INDEX 13) - FOLIO
        
        email = email_cell.value
        clase = clase_cell.value
        calibre = calibre_cell.value
        marca = marca_cell.value
        modelo = modelo_cell.value
        matricula = matricula_cell.value
        folio = folio_cell.value
        
        # Validar
        if not all([email, clase, calibre, matricula]):
            continue
        
        email = str(email).lower().strip()
        clase = str(clase).strip()
        calibre = str(calibre).strip()
        marca = str(marca).strip() if marca else ''
        modelo = str(modelo).strip() if modelo else ''
        matricula = str(matricula).strip()
        folio = str(folio).strip() if folio else ''
        
        if email not in armas_por_socio:
            armas_por_socio[email] = []
        
        armas_por_socio[email].append({
            'clase': clase,
            'calibre': calibre,
            'marca': marca,
            'modelo': modelo,
            'matricula': matricula,
            'folio': folio,
            'row': row_num
        })
    
    print(f'✅ Armas cargadas: {len(armas_por_socio)} socios')
    
    return armas_por_socio, fechas_map

# ============================================================================
# VALIDAR SOCIOS
# ============================================================================

async def validar_socios(emails):
    """Valida que los socios existan en Firestore."""
    print('\n🔍 Validando socios en Firestore...')
    
    socios_ref = db.collection('socios')
    validos = []
    invalidos = []
    
    for email in emails:
        doc = socios_ref.document(email).get()
        if doc.exists:
            validos.append(email)
        else:
            invalidos.append(email)
    
    if invalidos:
        print(f'⚠️  Socios NO encontrados: {len(invalidos)}')
        for e in invalidos[:5]:
            print(f'   - {e}')
    
    print(f'✅ Socios válidos: {len(validos)}/{len(emails)}')
    return validos

def validar_socios_sync(emails):
    """Versión síncrona de validar_socios."""
    print('\n🔍 Validando socios en Firestore...')
    
    socios_ref = db.collection('socios')
    validos = []
    invalidos = []
    
    for email in emails:
        doc = socios_ref.document(email).get()
        if doc.exists:
            validos.append(email)
        else:
            invalidos.append(email)
    
    if invalidos:
        print(f'⚠️  Socios NO encontrados: {len(invalidos)}')
        for e in invalidos[:5]:
            print(f'   - {e}')
    
    print(f'✅ Socios válidos: {len(validos)}/{len(emails)}')
    return validos

# ============================================================================
# LIMPIAR ARMAS
# ============================================================================

def limpiar_armas(emails):
    """Elimina todas las armas actuales."""
    print('\n🗑️  Limpiando colecciones de armas...')
    
    total_eliminadas = 0
    
    for email in emails:
        armas_ref = db.collection('socios').document(email).collection('armas')
        docs = armas_ref.stream()
        
        for doc in docs:
            doc.reference.delete()
            total_eliminadas += 1
    
    print(f'✅ Armas eliminadas: {total_eliminadas}')

# ============================================================================
# REPOBLAR ARMAS
# ============================================================================

def determinar_modalidad(clase):
    """Determina si es 'caza' o 'tiro' basado en la clase."""
    clase_upper = clase.upper()
    
    if any(x in clase_upper for x in ['RIFLE', 'CARABINA', 'ESCOPETA', 'SHOTGUN']):
        return 'tiro'
    
    if any(x in clase_upper for x in ['PISTOLA', 'REVOLVER']):
        return 'tiro'
    
    return 'tiro'

def repoblar_armas(armas_por_socio, socios_validos):
    """Repuebla la colección de armas."""
    print('\n📝 Repoblando armas...')
    
    total_insertadas = 0
    batch_count = 0
    batch = db.batch()
    BATCH_SIZE = 100
    
    for email in socios_validos:
        if email not in armas_por_socio:
            continue
        
        armas = armas_por_socio[email]
        
        for arma in armas:
            # Usar UUID para evitar problemas con caracteres especiales en la matrícula
            arma_id = str(uuid.uuid4())
            arma_ref = db.collection('socios').document(email).collection('armas').document(arma_id)
            
            batch.set(arma_ref, {
                'clase': arma['clase'],
                'calibre': arma['calibre'],
                'marca': arma['marca'],
                'modelo': arma['modelo'],
                'matricula': arma['matricula'],
                'folio': arma['folio'],
                'modalidad': determinar_modalidad(arma['clase']),
                'fechaActualizacion': firestore.SERVER_TIMESTAMP
            })
            
            batch_count += 1
            total_insertadas += 1
            
            if batch_count >= BATCH_SIZE:
                batch.commit()
                print(f'   ✓ {total_insertadas} armas insertadas...')
                batch = db.batch()
                batch_count = 0
    
    if batch_count > 0:
        batch.commit()
    
    print(f'✅ Total de armas insertadas: {total_insertadas}')

# ============================================================================
# ACTUALIZAR FECHAS
# ============================================================================

def actualizar_fechas(fechas_map, socios_validos):
    """Actualiza las fechas de ingreso."""
    print('\n📅 Actualizando fechas de ingreso...')
    
    actualizadas = 0
    batch = db.batch()
    batch_count = 0
    BATCH_SIZE = 100
    
    for email in socios_validos:
        if email not in fechas_map:
            print(f'   ⚠️  Sin fecha para: {email}')
            continue
        
        socio_ref = db.collection('socios').document(email)
        fecha = fechas_map[email]
        
        batch.update(socio_ref, {
            'fechaAlta': fecha,
            'fechaActualizacionFecha': firestore.SERVER_TIMESTAMP
        })
        
        batch_count += 1
        actualizadas += 1
        
        if batch_count >= BATCH_SIZE:
            batch.commit()
            print(f'   ✓ {actualizadas} fechas actualizadas...')
            batch = db.batch()
            batch_count = 0
    
    if batch_count > 0:
        batch.commit()
    
    print(f'✅ Total de fechas actualizadas: {actualizadas}')

# ============================================================================
# MAIN
# ============================================================================

def main():
    try:
        print('🚀 REPOBLACIÓN DE ARMAS Y FECHAS DE INGRESO')
        print('=' * 50 + '\n')
        
        # 1. Leer Excel
        armas_por_socio, fechas_map = leer_excel()
        
        # 2. Validar socios
        emails = list(armas_por_socio.keys())
        socios_validos = validar_socios_sync(emails)
        
        # 3. Limpiar armas
        limpiar_armas(socios_validos)
        
        # 4. Repoblar armas
        repoblar_armas(armas_por_socio, socios_validos)
        
        # 5. Actualizar fechas
        actualizar_fechas(fechas_map, socios_validos)
        
        print('\n✅ ¡REPOBLACIÓN COMPLETADA!\n')
        
    except Exception as err:
        print(f'\n❌ ERROR: {err}')
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    main()
