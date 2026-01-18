#!/usr/bin/env python3
"""
Normalizar archivo de diciembre 2025 - Duplicar datos de socio en cada fila de arma
"""
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime

archivo_original = '/Applications/club-738-web/2025-dic-usb-738/CLUB 738-31-DE-DICIEMBRE-2025_RELACION_SOCIOS_ARMAS.xlsx'
archivo_normalizado = '/Applications/club-738-web/2025-dic-usb-738/CLUB 738-31-DE-DICIEMBRE-2025_RELACION_SOCIOS_ARMAS_NORMALIZADO.xlsx'

print("=" * 100)
print("NORMALIZACIÓN DE ARCHIVO DICIEMBRE 2025")
print("=" * 100)

# Cargar archivo original
wb_orig = openpyxl.load_workbook(archivo_original)
ws_orig = wb_orig['CLUB 738. RELACION SOCIOS 31 DI']

print(f"\n📂 Archivo original: {archivo_original.split('/')[-1]}")
print(f"   Filas totales: {ws_orig.max_row}")

# Crear nuevo workbook normalizado
wb_nuevo = openpyxl.Workbook()
ws_nuevo = wb_nuevo.active
ws_nuevo.title = "CLUB 738 NORMALIZADO"

# Copiar encabezados de la fila 7 (donde están los headers reales)
headers = []
for i, cell in enumerate(ws_orig[7], 1):
    if cell.value:
        headers.append(cell.value)
        ws_nuevo.cell(row=1, column=i, value=cell.value)
        ws_nuevo.cell(row=1, column=i).font = Font(bold=True)
        ws_nuevo.cell(row=1, column=i).alignment = Alignment(horizontal='center')

print(f"\n📋 Headers detectados: {len(headers)} columnas")
for i, h in enumerate(headers[:15], 1):
    print(f"   Col {i}: {h}")

# Variables para trackear
datos_socio_actual = {}
fila_nueva = 2
total_armas = 0
total_socios = 0
armas_por_socio = []

print("\n🔄 Procesando filas...")

# Procesar desde fila 8 (datos)
for i in range(8, ws_orig.max_row + 1):
    row = list(ws_orig[i])
    
    # Obtener valores
    col1 = row[0].value  # No. REGISTRO DEL CLUB
    col2 = row[1].value  # DOMICILIO
    col3 = row[2].value  # NOMBRE DEL SOCIO
    col4 = row[3].value  # CURP
    col5 = row[4].value  # No. CONSEC.
    col6 = row[5].value  # Posible "TOTAL POR PERSONA"
    col9 = row[8].value  # CLASE
    
    # Detectar fila "TOTAL POR PERSONA" (skip)
    if col6 and 'TOTAL POR PERSONA' in str(col6).upper():
        if datos_socio_actual:
            armas_count = len([a for a in armas_por_socio if a == datos_socio_actual.get('nombre')])
            print(f"   ✅ {datos_socio_actual.get('nombre')}: {armas_count} armas")
        continue
    
    # Detectar nueva fila de SOCIO (tiene col3=nombre Y NO tiene col9=CLASE)
    # Puede o no tener col1=738
    if col3 and not col9:
        # Guardar datos del socio
        if datos_socio_actual:
            total_socios += 1
        
        datos_socio_actual = {
            'registro': col1 if col1 else '738',
            'domicilio': col2,
            'nombre': str(col3).strip() if col3 else '',
            'curp': str(col4).strip() if col4 else '',
            'consec': col5
        }
        continue
    
    # Detectar fila de ARMA (tiene col9=CLASE)
    if col9 and datos_socio_actual:
        # Copiar datos del socio + datos del arma
        ws_nuevo.cell(row=fila_nueva, column=1, value=datos_socio_actual['registro'])
        ws_nuevo.cell(row=fila_nueva, column=2, value=datos_socio_actual['domicilio'])
        ws_nuevo.cell(row=fila_nueva, column=3, value=datos_socio_actual['nombre'])
        ws_nuevo.cell(row=fila_nueva, column=4, value=datos_socio_actual['curp'])
        ws_nuevo.cell(row=fila_nueva, column=5, value=datos_socio_actual['consec'])
        
        # Copiar resto de columnas (6 en adelante son datos del arma)
        for j in range(5, len(row)):
            ws_nuevo.cell(row=fila_nueva, column=j+1, value=row[j].value)
        
        total_armas += 1
        armas_por_socio.append(datos_socio_actual['nombre'])
        fila_nueva += 1

# Último socio
if datos_socio_actual:
    total_socios += 1
    armas_count = len([a for a in armas_por_socio if a == datos_socio_actual.get('nombre')])
    print(f"   ✅ {datos_socio_actual.get('nombre')}: {armas_count} armas")

print(f"\n📊 RESUMEN DE NORMALIZACIÓN:")
print(f"   Total socios procesados: {total_socios}")
print(f"   Total armas procesadas: {total_armas}")
print(f"   Filas en nuevo archivo: {fila_nueva - 1}")

# Guardar archivo normalizado
wb_nuevo.save(archivo_normalizado)
print(f"\n✅ Archivo normalizado guardado:")
print(f"   {archivo_normalizado}")

print("\n" + "=" * 100)
print("AUDITORÍA DE NORMALIZACIÓN")
print("=" * 100)

# Auditoría: verificar que cada fila tiene todos los datos
wb_check = openpyxl.load_workbook(archivo_normalizado)
ws_check = wb_check.active

filas_completas = 0
filas_incompletas = 0
problemas = []

for i in range(2, ws_check.max_row + 1):
    row = list(ws_check[i])
    
    # Verificar que tenga nombre, CURP, y CLASE
    nombre = row[2].value if len(row) > 2 else None
    curp = row[3].value if len(row) > 3 else None
    clase = row[8].value if len(row) > 8 else None
    matricula = row[12].value if len(row) > 12 else None
    
    if nombre and curp and clase and matricula:
        filas_completas += 1
    else:
        filas_incompletas += 1
        problemas.append({
            'fila': i,
            'nombre': nombre,
            'curp': curp,
            'clase': clase,
            'matricula': matricula
        })

print(f"\n✅ Filas completas: {filas_completas}")
print(f"❌ Filas incompletas: {filas_incompletas}")

if problemas:
    print(f"\n⚠️ PROBLEMAS DETECTADOS ({len(problemas)} filas):")
    for p in problemas[:10]:  # Mostrar primeros 10
        print(f"   Fila {p['fila']}: Nombre={p['nombre']}, CURP={p['curp']}, Clase={p['clase']}, Mat={p['matricula']}")
else:
    print(f"\n✅ AUDITORÍA EXITOSA - Todos los datos están completos")

# Estadísticas por socio
from collections import Counter
socios_contador = Counter()

for i in range(2, ws_check.max_row + 1):
    row = list(ws_check[i])
    nombre = row[2].value if len(row) > 2 else None
    if nombre:
        socios_contador[nombre] += 1

print(f"\n📊 DISTRIBUCIÓN DE ARMAS POR SOCIO:")
print(f"   Total socios: {len(socios_contador)}")
print(f"   Total armas: {sum(socios_contador.values())}")
print(f"\n   Top 10 socios con más armas:")
for socio, count in socios_contador.most_common(10):
    print(f"     • {socio[:40]:40s} → {count} armas")

print("\n" + "=" * 100)
