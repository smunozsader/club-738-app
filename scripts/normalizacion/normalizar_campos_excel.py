#!/usr/bin/env python3
"""
Normalizar campos en Excel:
1. Matrículas: números como texto SIN comas
2. Teléfonos: sin comas
3. Modelos: sin comas (1,100 → 1100)
4. Corregir REMIGIO: 19/22 → 10/22
"""

from openpyxl import load_workbook
from openpyxl.styles import Alignment

archivo = 'socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx'

print(f"📂 Abriendo: {archivo}")
wb = load_workbook(archivo)
ws = wb.active

# Headers en fila 1
headers = [cell.value for cell in ws[1]]
print(f"\n📋 Headers encontrados:")
for i, h in enumerate(headers, 1):
    print(f"   Col {i}: {h}")

# Encontrar columnas
col_nombre = headers.index("NOMBRE SOCIO") + 1 if "NOMBRE SOCIO" in headers else None
col_telefono = headers.index("TELEFONO") + 1 if "TELEFONO" in headers else None
col_modelo = headers.index("MODELO") + 1 if "MODELO" in headers else None
col_matricula = headers.index("MATRÍCULA") + 1 if "MATRÍCULA" in headers else None

print(f"\n🔍 Columnas a normalizar:")
print(f"   NOMBRE SOCIO: Col {col_nombre}")
print(f"   TELEFONO: Col {col_telefono}")
print(f"   MODELO: Col {col_modelo}")
print(f"   MATRÍCULA: Col {col_matricula}")

# Contadores
matriculas_normalizadas = 0
telefonos_normalizados = 0
modelos_normalizados = 0
remigio_corregido = False

print(f"\n🔧 Procesando {ws.max_row - 1} filas...")

for row in range(2, ws.max_row + 1):
    # 1. NORMALIZAR MATRÍCULAS (números como texto SIN comas)
    if col_matricula:
        matricula = ws.cell(row, col_matricula).value
        if matricula and matricula != "0":
            # Convertir a string y quitar comas
            matricula_str = str(matricula).replace(',', '').strip()
            
            # Si cambió, actualizar
            if str(matricula) != matricula_str:
                ws.cell(row, col_matricula).value = matricula_str
                ws.cell(row, col_matricula).number_format = '@'  # Formato texto
                ws.cell(row, col_matricula).alignment = Alignment(horizontal='left')
                matriculas_normalizadas += 1
    
    # 2. NORMALIZAR TELÉFONOS (sin comas)
    if col_telefono:
        telefono = ws.cell(row, col_telefono).value
        if telefono:
            telefono_str = str(telefono).replace(',', '').strip()
            if str(telefono) != telefono_str:
                ws.cell(row, col_telefono).value = telefono_str
                ws.cell(row, col_telefono).number_format = '@'
                telefonos_normalizados += 1
    
    # 3. NORMALIZAR MODELOS (sin comas: 1,100 → 1100)
    if col_modelo:
        modelo = ws.cell(row, col_modelo).value
        if modelo and modelo != "0":
            modelo_str = str(modelo).replace(',', '').strip()
            if str(modelo) != modelo_str:
                ws.cell(row, col_modelo).value = modelo_str
                modelos_normalizados += 1
    
    # 4. CORREGIR REMIGIO: 19/22 → 10/22
    if col_nombre and col_modelo:
        nombre = ws.cell(row, col_nombre).value
        modelo = ws.cell(row, col_modelo).value
        
        if nombre and "REMIGIO" in nombre.upper() and modelo == "19/22":
            ws.cell(row, col_modelo).value = "10/22"
            print(f"   ✅ REMIGIO corregido en fila {row}: 19/22 → 10/22")
            remigio_corregido = True

print(f"\n📊 Resumen de cambios:")
print(f"   Matrículas normalizadas: {matriculas_normalizadas}")
print(f"   Teléfonos normalizados: {telefonos_normalizados}")
print(f"   Modelos normalizados: {modelos_normalizados}")
print(f"   REMIGIO corregido: {'✅ Sí' if remigio_corregido else '❌ No encontrado'}")

# Guardar
print(f"\n💾 Guardando cambios...")
wb.save(archivo)
print(f"✅ Archivo actualizado: {archivo}")

# Verificación
print(f"\n🔍 Verificando REMIGIO...")
wb2 = load_workbook(archivo)
ws2 = wb2.active
for row in range(2, ws2.max_row + 1):
    nombre = ws2.cell(row, col_nombre).value
    if nombre and "REMIGIO" in nombre.upper():
        modelo = ws2.cell(row, col_modelo).value
        matricula = ws2.cell(row, col_matricula).value
        print(f"   Fila {row}: {nombre}")
        print(f"      Modelo: {modelo}")
        print(f"      Matrícula: {matricula} (tipo: {type(matricula).__name__})")

print(f"\n✅ Normalización completada!")
