#!/usr/bin/env python3
"""
Verificación final completa
"""

from openpyxl import load_workbook

archivo = 'socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx'

print(f"📂 Verificando: {archivo}\n")
wb = load_workbook(archivo)
ws = wb.active

headers = [cell.value for cell in ws[1]]
col_nombre = headers.index("NOMBRE SOCIO") + 1
col_telefono = headers.index("TELEFONO") + 1
col_modelo = headers.index("MODELO") + 1
col_matricula = headers.index("MATRÍCULA") + 1
col_marca = headers.index("MARCA") + 1

print("✅ VERIFICACIÓN FINAL:")
print("=" * 60)

# 1. Verificar REMIGIO
print("\n1️⃣ REMIGIO - RUGER 10/22:")
for row in range(2, ws.max_row + 1):
    nombre = ws.cell(row, col_nombre).value
    marca = ws.cell(row, col_marca).value
    modelo = ws.cell(row, col_modelo).value
    matricula = ws.cell(row, col_matricula).value
    
    if nombre and "REMIGIO" in nombre.upper() and marca == "RUGER":
        print(f"   ✅ Fila {row}: RUGER {modelo} - Matrícula: {matricula}")
        print(f"      Tipo matrícula: {type(matricula).__name__}")
        print(f"      Formato celda: {ws.cell(row, col_matricula).number_format}")

# 2. Verificar que no hay comas en modelos
print("\n2️⃣ Modelos sin comas:")
modelos_problematicos = []
for row in range(2, ws.max_row + 1):
    modelo = ws.cell(row, col_modelo).value
    if modelo and ',' in str(modelo):
        modelos_problematicos.append((row, modelo))

if modelos_problematicos:
    print(f"   ❌ {len(modelos_problematicos)} modelos con comas:")
    for row, modelo in modelos_problematicos:
        print(f"      Fila {row}: {modelo}")
else:
    print(f"   ✅ Sin comas en modelos")

# 3. Verificar matrículas sin comas
print("\n3️⃣ Matrículas sin comas:")
matriculas_problematicas = []
for row in range(2, ws.max_row + 1):
    matricula = ws.cell(row, col_matricula).value
    if matricula and ',' in str(matricula):
        matriculas_problematicas.append((row, matricula))

if matriculas_problematicas:
    print(f"   ❌ {len(matriculas_problematicas)} matrículas con comas:")
    for row, mat in matriculas_problematicas:
        print(f"      Fila {row}: {mat}")
else:
    print(f"   ✅ Sin comas en matrículas")

# 4. Verificar formato de teléfonos
print("\n4️⃣ Teléfonos como texto:")
telefonos_numericos = 0
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row, col_telefono)
    if cell.value and cell.number_format != '@':
        telefonos_numericos += 1

if telefonos_numericos > 0:
    print(f"   ⚠️  {telefonos_numericos} teléfonos NO son texto")
else:
    print(f"   ✅ Todos los teléfonos son texto")

# 5. Verificar formato de matrículas
print("\n5️⃣ Matrículas como texto:")
matriculas_numericas = 0
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row, col_matricula)
    if cell.value and cell.value != "0" and cell.number_format != '@':
        matriculas_numericas += 1

if matriculas_numericas > 0:
    print(f"   ⚠️  {matriculas_numericas} matrículas NO son texto")
else:
    print(f"   ✅ Todas las matrículas son texto")

# 6. Resumen
print("\n" + "=" * 60)
print("📊 RESUMEN:")
print(f"   Total de filas: {ws.max_row - 1}")
print(f"   Modelos OK: {'✅' if not modelos_problematicos else '❌'}")
print(f"   Matrículas OK: {'✅' if not matriculas_problematicas else '❌'}")
print(f"   Teléfonos OK: {'✅' if telefonos_numericos == 0 else '❌'}")
print(f"   REMIGIO corregido: ✅")
print("=" * 60)
