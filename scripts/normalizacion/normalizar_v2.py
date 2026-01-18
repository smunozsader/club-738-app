#!/usr/bin/env python3
"""
Normalizar archivo diciembre 2025 - VERSIÓN CORREGIDA
Estrategia: Detectar filas de socio vs filas de arma correctamente
"""
import openpyxl
from datetime import datetime

archivo_original = '/Applications/club-738-web/2025-dic-usb-738/CLUB 738-31-DE-DICIEMBRE-2025_RELACION_SOCIOS_ARMAS.xlsx'
archivo_normalizado = '/Applications/club-738-web/2025-dic-usb-738/CLUB 738-31-DE-DICIEMBRE-2025_RELACION_SOCIOS_ARMAS_NORMALIZADO.xlsx'

print("=" * 100)
print("NORMALIZACIÓN DICIEMBRE 2025 - V2")
print("=" * 100)

wb_orig = openpyxl.load_workbook(archivo_original)
ws_orig = wb_orig['CLUB 738. RELACION SOCIOS 31 DI']

# Crear nuevo workbook
wb_nuevo = openpyxl.Workbook()
ws_nuevo = wb_nuevo.active
ws_nuevo.title = "NORMALIZADO"

# Copiar headers (fila 7)
for i, cell in enumerate(ws_orig[7], 1):
    if cell.value:
        ws_nuevo.cell(row=1, column=i, value=cell.value)

# Variables
datos_socio = None
fila_nueva = 2
total_armas = 0
socios_procesados = []

print("\n🔄 Procesando...")

# Procesar desde fila 8
for i in range(8, ws_orig.max_row + 1):
    row = list(ws_orig[i])
    
    col1 = row[0].value  # No. REGISTRO
    col3 = row[2].value  # NOMBRE DEL SOCIO
    col4 = row[3].value  # CURP
    col6 = row[5].value  # DOMICILIO o "TOTAL POR PERSONA"
    col9 = row[8].value  # CLASE
    
    # Skip fila "TOTAL POR PERSONA"
    if col6 and isinstance(col6, str) and 'TOTAL POR PERSONA' in col6.upper():
        continue
    
    # DETECTAR FILA DE SOCIO: tiene col3 (nombre) pero NO tiene col9 (CLASE)
    if col3 and not col9:
        # Guardar socio anterior si existe
        if datos_socio:
            socios_procesados.append(datos_socio['nombre'])
        
        # Nuevo socio
        datos_socio = {
            'col1': col1 if col1 else '738',
            'col2': row[1].value,  # DOMICILIO DEL CLUB
            'col3': col3,  # NOMBRE
            'col4': col4,  # CURP
            'col5': row[4].value,  # No. CONSEC
            'col6': col6,  # DOMICILIO SOCIO
            'col7': row[6].value,  # TELEFONO
            'col8': row[7].value,  # EMAIL
            'nombre': str(col3).strip()
        }
        continue
    
    # DETECTAR FILA DE ARMA: tiene col9 (CLASE)
    if col9 and datos_socio:
        # Copiar datos del socio (cols 1-8)
        ws_nuevo.cell(row=fila_nueva, column=1, value=datos_socio['col1'])
        ws_nuevo.cell(row=fila_nueva, column=2, value=datos_socio['col2'])
        ws_nuevo.cell(row=fila_nueva, column=3, value=datos_socio['col3'])
        ws_nuevo.cell(row=fila_nueva, column=4, value=datos_socio['col4'])
        ws_nuevo.cell(row=fila_nueva, column=5, value=datos_socio['col5'])
        ws_nuevo.cell(row=fila_nueva, column=6, value=datos_socio['col6'])
        ws_nuevo.cell(row=fila_nueva, column=7, value=datos_socio['col7'])
        ws_nuevo.cell(row=fila_nueva, column=8, value=datos_socio['col8'])
        
        # Copiar datos del arma (cols 9 en adelante)
        for j in range(8, min(len(row), 18)):  # Max 18 columnas
            ws_nuevo.cell(row=fila_nueva, column=j+1, value=row[j].value)
        
        total_armas += 1
        fila_nueva += 1

# Último socio
if datos_socio:
    socios_procesados.append(datos_socio['nombre'])

print(f"✅ Procesado: {len(set(socios_procesados))} socios únicos")
print(f"✅ Total armas: {total_armas}")

# Guardar
wb_nuevo.save(archivo_normalizado)
print(f"\n💾 Guardado: {archivo_normalizado.split('/')[-1]}")

# AUDITORÍA
print("\n" + "=" * 100)
print("AUDITORÍA")
print("=" * 100)

wb_check = openpyxl.load_workbook(archivo_normalizado)
ws_check = wb_check.active

# Verificar totales
total_filas = ws_check.max_row - 1  # -1 por header

# Clasificar por tipo
cortas = 0
largas = 0

for i in range(2, ws_check.max_row + 1):
    clase = ws_check.cell(row=i, column=9).value
    if clase:
        clase_str = str(clase).upper()
        if 'PISTOLA' in clase_str or 'REVOLVER' in clase_str or 'REVÓLVER' in clase_str or 'KIT' in clase_str:
            cortas += 1
        else:
            largas += 1

print(f"\n📊 TOTALES:")
print(f"   Armas totales: {total_filas}")
print(f"   Cortas: {cortas}")
print(f"   Largas: {largas}")

# Verificar contra original
print(f"\n✅ VERIFICACIÓN vs ORIGINAL:")
print(f"   Original reportado: 276 armas (104 cortas + 172 largas)")
print(f"   Normalizado: {total_filas} armas ({cortas} cortas + {largas} largas)")

if total_filas == 276 and cortas == 104 and largas == 172:
    print(f"   ✅✅✅ PERFECTO - COINCIDENCIA EXACTA")
else:
    print(f"   ⚠️ DIFERENCIA: {total_filas - 276:+d} armas")

# Top socios
from collections import Counter
socios_count = Counter()
for i in range(2, ws_check.max_row + 1):
    nombre = ws_check.cell(row=i, column=3).value
    if nombre:
        socios_count[str(nombre).strip()] += 1

print(f"\n👥 TOP 10 SOCIOS:")
for socio, count in socios_count.most_common(10):
    nombre_corto = socio[:50] if len(socio) > 50 else socio
    print(f"   {count:3d} armas → {nombre_corto}")

print("\n" + "=" * 100)
