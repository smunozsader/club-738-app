#!/usr/bin/env python3
"""
Normalizar diciembre 2025 - VERSIÓN FINAL CORRECTA
FORMATO REAL:
- Fila de socio: tiene NOMBRE en col3, puede o no tener CLASE en col9
  * Si tiene CLASE → es la primera arma del socio
  * Si NO tiene CLASE → solo es header del socio
- Filas siguientes: solo CLASE (armas adicionales)
- Fila "TOTAL POR PERSONA": fin del socio
"""
import openpyxl

archivo_original = '/Applications/club-738-web/2025-dic-usb-738/CLUB 738-31-DE-DICIEMBRE-2025_RELACION_SOCIOS_ARMAS.xlsx'
archivo_normalizado = '/Applications/club-738-web/2025-dic-usb-738/CLUB 738-31-DE-DICIEMBRE-2025_RELACION_SOCIOS_ARMAS_NORMALIZADO.xlsx'

print("=" * 100)
print("NORMALIZACIÓN DICIEMBRE 2025 - VERSIÓN FINAL")
print("=" * 100)

wb_orig = openpyxl.load_workbook(archivo_original)
ws_orig = wb_orig['CLUB 738. RELACION SOCIOS 31 DI']

wb_nuevo = openpyxl.Workbook()
ws_nuevo = wb_nuevo.active
ws_nuevo.title = "NORMALIZADO"

# Copiar headers (fila 7)
for i, cell in enumerate(ws_orig[7], 1):
    if cell.value:
        ws_nuevo.cell(row=1, column=i, value=cell.value)

datos_socio = None
fila_nueva = 2
total_armas = 0

print("\n🔄 Procesando...")

for i in range(8, ws_orig.max_row + 1):
    row = list(ws_orig[i])
    
    col3 = row[2].value  # NOMBRE
    col6 = row[5].value  # DOMICILIO o "TOTAL POR PERSONA"
    col9 = row[8].value  # CLASE
    
    # Skip "TOTAL POR PERSONA"
    if col6 and isinstance(col6, str) and 'TOTAL' in col6.upper():
        continue
    
    # CASO 1: Fila con NOMBRE (inicio de nuevo socio)
    if col3:
        # Guardar datos del socio
        datos_socio = {
            'cols': [cell.value for cell in row[:8]]  # Copiar cols 1-8
        }
        
        # Si esta fila TAMBIÉN tiene CLASE → es la primera arma del socio
        if col9:
            # Copiar datos socio + datos arma
            for j in range(18):  # 18 columnas
                ws_nuevo.cell(row=fila_nueva, column=j+1, value=row[j].value if j < len(row) else None)
            total_armas += 1
            fila_nueva += 1
        
        continue
    
    # CASO 2: Fila SIN nombre pero CON CLASE → arma adicional del socio actual
    if col9 and datos_socio:
        # Copiar datos del socio (cols 1-8)
        for j in range(8):
            ws_nuevo.cell(row=fila_nueva, column=j+1, value=datos_socio['cols'][j])
        
        # Copiar datos del arma (cols 9 en adelante)
        for j in range(8, min(len(row), 18)):
            ws_nuevo.cell(row=fila_nueva, column=j+1, value=row[j].value)
        
        total_armas += 1
        fila_nueva += 1

print(f"✅ Total armas procesadas: {total_armas}")

wb_nuevo.save(archivo_normalizado)
print(f"💾 Guardado: {archivo_normalizado.split('/')[-1]}")

# AUDITORÍA
print("\n" + "=" * 100)
print("AUDITORÍA FINAL")
print("=" * 100)

wb_check = openpyxl.load_workbook(archivo_normalizado)
ws_check = wb_check.active

total_filas = ws_check.max_row - 1

# Clasificar
cortas = 0
largas = 0
for i in range(2, ws_check.max_row + 1):
    clase = ws_check.cell(row=i, column=9).value
    if clase:
        clase_str = str(clase).upper()
        if 'PISTOLA' in clase_str or 'REVOLVER' in clase_str or 'REVÓLVER' in clase_str or 'KIT' in clase_str:
            cortas += 1
        else:
            largas += 1

print(f"\n📊 TOTALES NORMALIZADOS:")
print(f"   Total armas: {total_filas}")
print(f"   Cortas: {cortas}")
print(f"   Largas: {largas}")

print(f"\n✅ VERIFICACIÓN:")
print(f"   Original: 276 armas (104 cortas + 172 largas)")
print(f"   Normalizado: {total_filas} armas ({cortas} cortas + {largas} largas)")

if total_filas == 276 and cortas == 104 and largas == 172:
    print(f"\n   ✅✅✅ PERFECTO - COINCIDENCIA EXACTA ✅✅✅")
else:
    print(f"\n   ⚠️ Diferencia: {total_filas - 276:+d} armas")

# Verificar que todas las filas tienen datos completos
filas_completas = 0
for i in range(2, ws_check.max_row + 1):
    nombre = ws_check.cell(row=i, column=3).value
    curp = ws_check.cell(row=i, column=4).value
    clase = ws_check.cell(row=i, column=9).value
    matricula = ws_check.cell(row=i, column=13).value
    
    if nombre and clase:  # Mínimo requerido
        filas_completas += 1

print(f"\n✅ Filas con datos completos: {filas_completas}/{total_filas}")

# Top socios
from collections import Counter
socios_count = Counter()
for i in range(2, ws_check.max_row + 1):
    nombre = ws_check.cell(row=i, column=3).value
    if nombre:
        socios_count[str(nombre).strip()] += 1

print(f"\n👥 TOP 10 SOCIOS CON MÁS ARMAS:")
for socio, count in socios_count.most_common(10):
    nombre_limpio = socio.split('. ')[-1] if '. ' in socio else socio
    print(f"   {count:2d} armas → {nombre_limpio}")

print("\n" + "=" * 100)
