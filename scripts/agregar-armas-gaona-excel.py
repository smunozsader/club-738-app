#!/usr/bin/env python3
"""
Agregar 3 armas nuevas al arsenal de Enrique Gaona Castañeda (credencial 195)
Fecha: 24 Feb 2026
"""

import openpyxl
from datetime import datetime
import shutil

EXCEL_PATH = "/Applications/club-738-web/data/referencias/socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx"

# Datos del socio
SOCIO = {
    "credencial": 195,
    "nombre": "ENRIQUE GAONA CASTAÑEDA",
    "email": "quiquis77@hotmail.com"
}

# Armas nuevas a agregar (aprobadas hoy)
ARMAS_NUEVAS = [
    {"clase": "RIFLE", "calibre": ".22 LR", "marca": "CESKA ZBROJOVKA", "modelo": "CZ 512", "matricula": "J100907", "folio": "A3900899"},
    {"clase": "PISTOLA", "calibre": ".380", "marca": "BERETTA", "modelo": "80 X", "matricula": "Y014871X", "folio": "A3900898"},
    {"clase": "PISTOLA", "calibre": ".22 LR", "marca": "SIG SAUER", "modelo": "P322", "matricula": "73A192310", "folio": "A3916770"}
]

print("=" * 100)
print(f"🔫 AGREGAR {len(ARMAS_NUEVAS)} ARMAS A ENRIQUE GAONA (Credencial {SOCIO['credencial']})")
print("=" * 100)

# 0. Crear backup
backup_path = EXCEL_PATH.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
shutil.copy2(EXCEL_PATH, backup_path)
print(f"\n0️⃣ BACKUP CREADO: {backup_path}")

# 1. Cargar Excel
print("\n1️⃣ CARGANDO EXCEL...")
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

# Encontrar última fila ocupada
ultima_fila = ws.max_row

print(f"   ✅ Excel cargado: {ultima_fila} filas")

# Buscar filas existentes de Enrique Gaona para copiar datos completos
print("\n2️⃣ BUSCANDO DATOS EXISTENTES DE ENRIQUE GAONA...")
fila_gaona = None
for row in range(2, ultima_fila + 1):
    cred = ws[f'C{row}'].value
    if cred == 195:
        fila_gaona = row
        print(f"   ✅ Encontrado en fila {row}")
        break

if not fila_gaona:
    print("   ⚠️ No encontrado por credencial, buscando por email...")
    for row in range(2, ultima_fila + 1):
        email = ws[f'G{row}'].value
        if email and 'quiquis77' in str(email).lower():
            fila_gaona = row
            print(f"   ✅ Encontrado en fila {row} por email")
            break

# Copiar datos de otras columnas de la fila existente
datos_socio = {}
if fila_gaona:
    # Copiar todas las columnas relevantes del socio
    columnas_socio = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
    for col in columnas_socio:
        datos_socio[col] = ws[f'{col}{fila_gaona}'].value
    print(f"   ✅ Datos del socio copiados de fila {fila_gaona}")

# 3. Agregar nuevas filas
print(f"\n3️⃣ AGREGANDO {len(ARMAS_NUEVAS)} ARMAS...")

nueva_fila = ultima_fila + 1
for arma in ARMAS_NUEVAS:
    # Copiar datos del socio si existen
    if datos_socio:
        for col, valor in datos_socio.items():
            ws[f'{col}{nueva_fila}'] = valor
    else:
        # Si no encontramos datos, usar los básicos
        ws[f'C{nueva_fila}'] = SOCIO['credencial']
        ws[f'D{nueva_fila}'] = SOCIO['nombre']
        ws[f'G{nueva_fila}'] = SOCIO['email']
    
    # Datos del arma
    ws[f'N{nueva_fila}'] = arma['clase']
    ws[f'O{nueva_fila}'] = arma['calibre']
    ws[f'P{nueva_fila}'] = arma['marca']
    ws[f'Q{nueva_fila}'] = arma['modelo']
    ws[f'R{nueva_fila}'] = arma['matricula']
    ws[f'S{nueva_fila}'] = arma['folio']
    
    print(f"   ✅ Fila {nueva_fila}: {arma['clase']} {arma['marca']} {arma['modelo']} ({arma['matricula']})")
    nueva_fila += 1

# 4. Guardar Excel
print("\n4️⃣ GUARDANDO EXCEL...")
wb.save(EXCEL_PATH)
print(f"   ✅ Excel guardado: {EXCEL_PATH}")

# 5. Verificar resultado
print("\n5️⃣ VERIFICACIÓN...")
wb2 = openpyxl.load_workbook(EXCEL_PATH)
ws2 = wb2.active
armas_gaona = 0
for row in range(2, ws2.max_row + 1):
    cred = ws2[f'C{row}'].value
    if cred == 195:
        armas_gaona += 1
        print(f"   • Fila {row}: {ws2[f'N{row}'].value} {ws2[f'P{row}'].value} {ws2[f'Q{row}'].value}")

print(f"\n   📊 TOTAL ARMAS DE ENRIQUE GAONA: {armas_gaona}")

print("\n" + "=" * 100)
print("✅ FUENTE DE VERDAD ACTUALIZADA")
print("=" * 100)
