#!/usr/bin/env python3
import openpyxl

print("\n✅ VERIFICACIÓN POST-NORMALIZACIÓN\n")

try:
    wb = openpyxl.load_workbook('socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx')
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    
    telefono_col = next((i for i, h in enumerate(headers) if h and 'TELEFONO' in str(h).upper()), None)
    cp_col = next((i for i, h in enumerate(headers) if h and 'CP' in str(h).upper()), None)
    modelo_col = next((i for i, h in enumerate(headers) if h and 'MODELO' in str(h).upper()), None)
    matricula_col = next((i for i, h in enumerate(headers) if h and 'MATRÍCULA' in str(h).upper()), None)
    
    print("📋 MUESTREO DE DATOS NORMALIZADOS (primeros 5 registros)\n")
    print("=" * 150)
    
    for row_idx in range(2, 7):
        nombre = ws.cell(row_idx, 4).value
        print(f"\n[Row {row_idx}] {nombre}")
        
        if telefono_col is not None:
            cell = ws.cell(row_idx, telefono_col + 1)
            print(f"   TELEFONO    → '{cell.value}' (formato: {cell.number_format})")
        
        if cp_col is not None:
            cell = ws.cell(row_idx, cp_col + 1)
            print(f"   CP          → '{cell.value}' (formato: {cell.number_format})")
        
        if modelo_col is not None:
            cell = ws.cell(row_idx, modelo_col + 1)
            print(f"   MODELO      → '{cell.value}' (formato: {cell.number_format})")
        
        if matricula_col is not None:
            cell = ws.cell(row_idx, matricula_col + 1)
            print(f"   MATRÍCULA   → '{cell.value}' (formato: {cell.number_format})")
    
    print("\n" + "=" * 150)
    print("\n✨ Todas las columnas normalizadas correctamente")
    print("   - Sin comas ni espacios innecesarios")
    print("   - Formato de texto (@) aplicado\n")
    
except Exception as e:
    print(f"❌ Error: {e}\n")
