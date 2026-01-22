#!/usr/bin/env python3
"""
Auditoría de armas para Santiago Alejandro Quintal Paredes
Compara Excel maestro vs Firestore
"""

import openpyxl
import json
import subprocess
import sys
from pathlib import Path

# Configuración
EXCEL_PATH = '/Applications/club-738-web/socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx'
EMAIL = 'squintal158@gmail.com'
SOCIO_NAME = 'SANTIAGO ALEJANDRO QUINTAL PAREDES'

def leer_armas_excel():
    """Lee el archivo Excel y extrae armas para Santiago"""
    print(f"\n{'='*70}")
    print(f"LEYENDO EXCEL: FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx")
    print(f"{'='*70}\n")
    
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        
        print(f"📊 Total de filas en Excel: {ws.max_row}")
        
        armas_excel = []
        headers = None
        
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            # Primera fila = encabezados
            if idx == 1:
                headers = row
                print(f"\n📋 Encabezados encontrados ({len(headers)} columnas):\n")
                for i, h in enumerate(headers):
                    if h:
                        print(f"   [{i}] {h}")
                continue
            
            # Buscar filas con el nombre del socio
            if row and len(row) > 0 and row[0]:
                nombre_en_fila = str(row[0]).strip().upper()
                
                if SOCIO_NAME in nombre_en_fila:
                    print(f"\n✅ ENCONTRADO (fila {idx}): {nombre_en_fila}")
                    
                    # Mapeo de índices (ajustados según estructura real del Excel)
                    arma_dict = {}
                    
                    if len(row) > 1:
                        arma_dict['nombre'] = str(row[0]).strip() if row[0] else ''
                    if len(row) > 2:
                        arma_dict['clase'] = str(row[2]).strip() if row[2] else ''
                    if len(row) > 3:
                        arma_dict['calibre'] = str(row[3]).strip() if row[3] else ''
                    if len(row) > 4:
                        arma_dict['marca'] = str(row[4]).strip() if row[4] else ''
                    if len(row) > 5:
                        arma_dict['modelo'] = str(row[5]).strip() if row[5] else ''
                    if len(row) > 6:
                        arma_dict['matricula'] = str(row[6]).strip() if row[6] else ''
                    if len(row) > 7:
                        arma_dict['folio'] = str(row[7]).strip() if row[7] else ''
                    
                    # Solo agregar si tiene matrícula y folio
                    if arma_dict.get('matricula') and arma_dict.get('folio'):
                        armas_excel.append(arma_dict)
                        print(f"   📌 Arma agregada: {arma_dict['clase']} {arma_dict['marca']} {arma_dict['modelo']}")
        
        print(f"\n✅ Total de armas encontradas en Excel para {SOCIO_NAME}: {len(armas_excel)}\n")
        return armas_excel
        
    except Exception as e:
        print(f"❌ Error al leer Excel: {e}")
        return []


def obtener_armas_firestore():
    """Obtiene armas de Firestore usando el script existente"""
    print(f"\n{'='*70}")
    print(f"LEYENDO FIRESTORE")
    print(f"{'='*70}\n")
    
    # Crear un script temporal de Node.js para leer Firestore
    script_content = f"""
import admin from 'firebase-admin';
import fs from 'fs';
import path from 'path';
import {{ fileURLToPath }} from 'url';

const __dirname = path.dirname(fileURLToPath(import.meta.url));

try {{
  const serviceAccountPath = path.join(__dirname, 'scripts', 'serviceAccountKey.json');
  if (!fs.existsSync(serviceAccountPath)) {{
    console.log(JSON.stringify({{ error: 'serviceAccountKey.json no encontrado' }}, null, 2));
    process.exit(1);
  }}

  const serviceAccount = JSON.parse(fs.readFileSync(serviceAccountPath, 'utf8'));

  admin.initializeApp({{
    credential: admin.credential.cert(serviceAccount),
    projectId: serviceAccount.project_id
  }});

  const db = admin.firestore();
  const email = '{EMAIL}';
  
  const armasSnapshot = await db.collection(`socios/${{email}}/armas`).get();
  
  const armas = [];
  armasSnapshot.forEach((doc) => {{
    armas.push({{
      id: doc.id,
      ...doc.data()
    }});
  }});
  
  console.log(JSON.stringify(armas, null, 2));
}} catch (e) {{
  console.error(JSON.stringify({{ error: e.message }}, null, 2));
  process.exit(1);
}}
"""
    
    temp_script = '/tmp/read_firestore_santiago.mjs'
    
    try:
        with open(temp_script, 'w') as f:
            f.write(script_content)
        
        # Ejecutar desde el directorio del proyecto
        result = subprocess.run(
            ['node', temp_script],
            cwd='/Applications/club-738-web',
            capture_output=True,
            text=True,
            timeout=10
        )
        
        if result.returncode != 0:
            print(f"⚠️  Error ejecutando Node: {result.stderr}")
            return []
        
        armas = json.loads(result.stdout)
        
        if isinstance(armas, dict) and 'error' in armas:
            print(f"❌ Error de Firestore: {armas['error']}")
            return []
        
        print(f"✅ Total de armas en Firestore: {len(armas)}\n")
        
        for arma in armas:
            print(f"   🔫 {arma.get('clase', 'N/A')} - {arma.get('marca', 'N/A')} {arma.get('modelo', 'N/A')}")
            print(f"      Calibre: {arma.get('calibre', 'N/A')}, Matrícula: {arma.get('matricula', 'N/A')}, Folio: {arma.get('folio', 'N/A')}\n")
        
        return armas
        
    except subprocess.TimeoutExpired:
        print("❌ Timeout al leer Firestore")
        return []
    except json.JSONDecodeError:
        print(f"❌ Error al parsear JSON de Firestore: {result.stdout}")
        return []
    except Exception as e:
        print(f"❌ Error general: {e}")
        return []
    finally:
        if Path(temp_script).exists():
            Path(temp_script).unlink()


def normalizar_arma(arma):
    """Normaliza arma para comparación"""
    # Convertir a string clave normalizada
    clase = str(arma.get('clase', '')).upper().strip()
    marca = str(arma.get('marca', '')).upper().strip()
    modelo = str(arma.get('modelo', '')).upper().strip()
    matricula = str(arma.get('matricula', '')).upper().strip().replace(' ', '')
    folio = str(arma.get('folio', '')).upper().strip()
    
    return f"{clase}|{marca}|{modelo}|{matricula}|{folio}"


def comparar_armas(armas_excel, armas_firestore):
    """Compara armas Excel vs Firestore"""
    print(f"\n{'='*70}")
    print(f"COMPARACIÓN EXCEL vs FIRESTORE")
    print(f"{'='*70}\n")
    
    # Normalizar claves
    excel_keys = set()
    firestore_keys = set()
    
    excel_map = {}
    firestore_map = {}
    
    for arma in armas_excel:
        key = normalizar_arma(arma)
        excel_keys.add(key)
        excel_map[key] = arma
    
    for arma in armas_firestore:
        key = normalizar_arma(arma)
        firestore_keys.add(key)
        firestore_map[key] = arma
    
    # Categorías
    en_excel_y_firestore = excel_keys & firestore_keys
    solo_en_excel = excel_keys - firestore_keys
    solo_en_firestore = firestore_keys - excel_keys
    
    # Reportar
    print(f"✅ EN AMBOS (Excel ∩ Firestore): {len(en_excel_y_firestore)}")
    for key in sorted(en_excel_y_firestore):
        arma = excel_map[key]
        print(f"   ✓ {arma.get('clase')} {arma.get('marca')} {arma.get('modelo')}")
        print(f"     Cal: {arma.get('calibre')}, Mat: {arma.get('matricula')}, Folio: {arma.get('folio')}")
    
    print(f"\n❌ SOLO EN EXCEL (no en Firestore): {len(solo_en_excel)}")
    for key in sorted(solo_en_excel):
        arma = excel_map[key]
        print(f"   ✗ {arma.get('clase')} {arma.get('marca')} {arma.get('modelo')}")
        print(f"     Cal: {arma.get('calibre')}, Mat: {arma.get('matricula')}, Folio: {arma.get('folio')}")
    
    print(f"\n⚠️  SOLO EN FIRESTORE (no en Excel): {len(solo_en_firestore)}")
    for key in sorted(solo_en_firestore):
        arma = firestore_map[key]
        print(f"   ⚠ {arma.get('clase')} {arma.get('marca')} {arma.get('modelo')}")
        print(f"     Cal: {arma.get('calibre')}, Mat: {arma.get('matricula')}, Folio: {arma.get('folio')}")
    
    print(f"\n{'='*70}")
    print(f"RESUMEN")
    print(f"{'='*70}")
    print(f"Armas en Excel: {len(armas_excel)}")
    print(f"Armas en Firestore: {len(armas_firestore)}")
    print(f"Coincidencias: {len(en_excel_y_firestore)}")
    print(f"Faltantes (Excel → Firestore): {len(solo_en_excel)}")
    print(f"Extras (Firestore → Excel): {len(solo_en_firestore)}")
    print(f"{'='*70}\n")


if __name__ == '__main__':
    # Leer Excel
    armas_excel = leer_armas_excel()
    
    # Leer Firestore
    armas_firestore = obtener_armas_firestore()
    
    # Comparar
    if armas_excel or armas_firestore:
        comparar_armas(armas_excel, armas_firestore)
    else:
        print("\n⚠️  No se encontraron datos para comparar")
