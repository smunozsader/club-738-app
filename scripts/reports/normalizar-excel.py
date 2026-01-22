#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import numbers as num_formats
import re

print("\n🔧 NORMALIZACIÓN DE COLUMNAS NUMÉRICAS\n")

try:
    # Cargar archivo
    print("📂 Cargando FUENTE_DE_VERDAD.xlsx...")
    wb = openpyxl.load_workbook('socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx')
    ws = wb.active
    
    # Identificar columnas
    headers = [cell.value for cell in ws[1]]
    
    telefono_col = next((i for i, h in enumerate(headers) if h and 'TELEFONO' in str(h).upper()), None)
    cp_col = next((i for i, h in enumerate(headers) if h and 'CP' in str(h).upper()), None)
    modelo_col = next((i for i, h in enumerate(headers) if h and 'MODELO' in str(h).upper()), None)
    matricula_col = next((i for i, h in enumerate(headers) if h and 'MATRÍCULA' in str(h).upper()), None)
    
    cols_to_fix = {
        'TELEFONO': (telefono_col, 'TELEFONO'),
        'CP': (cp_col, 'CP'),
        'MODELO': (modelo_col, 'MODELO'),
        'MATRICULA': (matricula_col, 'MATRICULA')
    }
    
    print(f"✅ Columnas identificadas:")
    for name, (idx, label) in cols_to_fix.items():
        if idx is not None:
            print(f"   {name:15s} → Columna {idx + 1} ({headers[idx]})")
    
    print("\n🔄 Normalizando...\n")
    
    # Procesar cada fila
    cambios_total = 0
    cambios_por_col = {
        'TELEFONO': 0,
        'CP': 0,
        'MODELO': 0,
        'MATRICULA': 0
    }
    
    for row_idx in range(2, ws.max_row + 1):
        # TELEFONO
        if telefono_col is not None:
            cell = ws.cell(row_idx, telefono_col + 1)
            if cell.value:
                valor_str = str(cell.value).strip()
                # Remover espacios y comas
                valor_limpio = valor_str.replace(' ', '').replace(',', '')
                if valor_str != valor_limpio:
                    cell.value = valor_limpio
                    cell.number_format = '@'  # Formato texto
                    cambios_por_col['TELEFONO'] += 1
                    cambios_total += 1
                else:
                    cell.number_format = '@'  # Asegurar formato texto
            
        # CP
        if cp_col is not None:
            cell = ws.cell(row_idx, cp_col + 1)
            if cell.value:
                valor_str = str(cell.value).strip()
                # Remover espacios y comas
                valor_limpio = valor_str.replace(' ', '').replace(',', '')
                if valor_str != valor_limpio:
                    cell.value = valor_limpio
                    cell.number_format = '@'  # Formato texto
                    cambios_por_col['CP'] += 1
                    cambios_total += 1
                else:
                    cell.number_format = '@'  # Asegurar formato texto
        
        # MODELO
        if modelo_col is not None:
            cell = ws.cell(row_idx, modelo_col + 1)
            if cell.value:
                valor_str = str(cell.value).strip()
                # Remover comas innecesarias
                valor_limpio = valor_str.replace(',', '')
                if valor_str != valor_limpio:
                    cell.value = valor_limpio
                    cell.number_format = '@'  # Formato texto
                    cambios_por_col['MODELO'] += 1
                    cambios_total += 1
                else:
                    cell.number_format = '@'  # Asegurar formato texto
        
        # MATRICULA
        if matricula_col is not None:
            cell = ws.cell(row_idx, matricula_col + 1)
            if cell.value:
                valor_str = str(cell.value).strip()
                # Remover comas innecesarias
                valor_limpio = valor_str.replace(',', '')
                if valor_str != valor_limpio:
                    cell.value = valor_limpio
                    cell.number_format = '@'  # Formato texto
                    cambios_por_col['MATRICULA'] += 1
                    cambios_total += 1
                else:
                    cell.number_format = '@'  # Asegurar formato texto
    
    # Guardar
    print(f"💾 Guardando cambios...\n")
    wb.save('socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx')
    
    print(f"✅ NORMALIZACIÓN COMPLETADA\n")
    print(f"📊 Cambios por columna:")
    for col_name, cambios in cambios_por_col.items():
        print(f"   {col_name:15s}: {cambios} filas normalizadas")
    print(f"\n📈 Total de cambios: {cambios_total}")
    print(f"✨ Todas las columnas ahora con formato TEXTO (@)\n")
    
except Exception as e:
    print(f"❌ Error: {e}\n")
    import traceback
    traceback.print_exc()
