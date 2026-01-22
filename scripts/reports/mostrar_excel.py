#!/usr/bin/env python3
"""
Mostrar contenido del Excel FUENTE_DE_VERDAD en la terminal
"""

import openpyxl
import pandas as pd
from pathlib import Path

EXCEL_PATH = "/Applications/club-738-web/socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx"

print("=" * 120)
print("📊 FUENTE DE VERDAD - CLUB 738")
print("=" * 120)

# 1. Cargar con openpyxl para ver estructura
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

print(f"\n📋 Hoja activa: {ws.title}")
print(f"📌 Filas totales: {ws.max_row}")
print(f"📌 Columnas totales: {ws.max_column}")

# 2. Mostrar encabezados
print("\n" + "=" * 120)
print("ENCABEZADOS (Fila 1):")
print("=" * 120)
headers = [cell.value for cell in ws[1]]
for col_idx, header in enumerate(headers, 1):
    print(f"  Col {col_idx}: {header}")

# 3. Cargar con pandas para mejor visualización
df = pd.read_excel(EXCEL_PATH)

print("\n" + "=" * 120)
print("PRIMERAS 5 FILAS:")
print("=" * 120)
print(df.head(5).to_string())

# 4. Buscar a Ricardo Soberanis
print("\n" + "=" * 120)
print("BÚSQUEDA: RICARDO SOBERANIS (verificar corrección de calibre):")
print("=" * 120)

ricardo_rows = df[df['NOMBRE SOCIO'].str.contains('RICARDO.*SOBERANIS', case=False, na=False, regex=True)]
if len(ricardo_rows) > 0:
    print(f"✅ Encontradas {len(ricardo_rows)} fila(s) para Ricardo Soberanis:")
    print(ricardo_rows[['No. CREDENCIAL', 'NOMBRE SOCIO', 'EMAIL', 'CLASE', 'CALIBRE', 'MARCA', 'MODELO', 'MATRÍCULA', 'FOLIO']].to_string())
else:
    print("❌ No encontrado")

# 5. Total de registros
print("\n" + "=" * 120)
print("ESTADÍSTICAS:")
print("=" * 120)
print(f"Total de socios: {len(df)}")
print(f"Total de armas registradas: {len(df)}")

# 6. Últimas 5 filas
print("\n" + "=" * 120)
print("ÚLTIMAS 5 FILAS (incluyendo nuevos registros):")
print("=" * 120)
print(df.tail(5).to_string())

print("\n" + "=" * 120)
