#!/usr/bin/env python3
import openpyxl
import os

excel_path = "socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx"

if not os.path.exists(excel_path):
    print(f"✗ Archivo no existe: {excel_path}")
    exit(1)

wb = openpyxl.load_workbook(excel_path)
ws = wb.active

print(f"📊 Excel: {os.path.basename(excel_path)}")
print(f"   Hoja activa: {ws.title}")
print(f"   Total filas: {ws.max_row}")
print(f"   Total columnas: {ws.max_column}\n")

# Mostrar encabezados
print("📋 ENCABEZADOS:")
headers = []
for col in range(1, ws.max_column + 1):
    header = ws.cell(1, col).value
    headers.append(header)
    print(f"   Col {col}: {header}")

print("\n📋 FILAS DE RICARDO ANTONIO SOBERANIS GAMBOA:")
ricardo_rows = []
for row in range(2, ws.max_row + 1):
    socio_col = ws.cell(row, 2).value  # Asumiendo que columna 2 es nombre
    if socio_col and "RICARDO" in str(socio_col).upper():
        print(f"\n   Fila {row}:")
        ricardo_rows.append(row)
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row, col).value
            print(f"      {headers[col-1]}: {cell_value}")

print(f"\n✓ Total filas de Ricardo encontradas: {len(ricardo_rows)}")
print(f"   Filas: {ricardo_rows}")
