#!/usr/bin/env python3
"""
Verificar NUEVA FUENTE DE VERDAD - Auditoría completa
"""
import openpyxl
from collections import Counter

archivo = "data/socios/2026_ENERO_RELACION_SOCIOS_ARMAS_MASTER.xlsx"

print("=" * 80)
print("AUDITORÍA: NUEVA FUENTE DE VERDAD - ENERO 2026")
print("=" * 80)

wb = openpyxl.load_workbook(archivo)
ws = wb.active

# Contar armas
total_armas = ws.max_row - 1  # Menos header

# Contar por clase
armas_cortas = 0
armas_largas = 0
clases = []

for i in range(2, ws.max_row + 1):
    clase = ws.cell(i, 12).value  # Col 12 = CLASE
    if clase:
        clases.append(clase)
        clase_upper = str(clase).upper()
        if any(x in clase_upper for x in ['PISTOLA', 'REVOLVER', 'KIT']):
            armas_cortas += 1
        else:
            armas_largas += 1

print(f"\n📊 TOTAL DE ARMAS: {total_armas}")
print(f"   - Cortas: {armas_cortas}")
print(f"   - Largas: {armas_largas}")

# Contar socios
curps = set()
nombres_socios = []
for i in range(2, ws.max_row + 1):
    curp = ws.cell(i, 5).value  # Col 5 = CURP
    nombre = ws.cell(i, 4).value  # Col 4 = NOMBRE
    if curp:
        curps.add(curp)
        if nombre:
            nombres_socios.append(nombre)

print(f"\n👥 TOTAL DE SOCIOS: {len(curps)}")

# Contar armas por socio
armas_por_socio = Counter(nombres_socios)
top_10 = armas_por_socio.most_common(10)

print(f"\n🔝 TOP 10 SOCIOS POR CANTIDAD DE ARMAS:")
for i, (socio, cantidad) in enumerate(top_10, 1):
    print(f"   {i:2d}. {socio}: {cantidad} armas")

# Verificar direcciones estructuradas
print(f"\n🏠 VERIFICACIÓN DE DIRECCIONES:")
con_direccion = 0
sin_direccion = 0
curps_verificados = set()

for i in range(2, ws.max_row + 1):
    curp = ws.cell(i, 5).value
    if curp and curp not in curps_verificados:
        curps_verificados.add(curp)
        calle = ws.cell(i, 7).value
        colonia = ws.cell(i, 8).value
        
        if calle or colonia:
            con_direccion += 1
        else:
            sin_direccion += 1

print(f"   ✅ Socios con dirección estructurada: {con_direccion}")
print(f"   ⚠️  Socios sin dirección estructurada: {sin_direccion}")

# Verificar las 4 armas nuevas
print(f"\n🆕 VERIFICACIÓN DE 4 ARMAS NUEVAS:")
armas_nuevas_buscar = [
    '73-H21YT-001717',  # Iván Cabo - Retay Gordion
    'FP40104',          # Iván Cabo - Shadow 2
    'DP25087',          # Gardoni - Shadow 2
    'C647155'           # Arechiga - P07
]

encontradas = []
for i in range(2, ws.max_row + 1):
    matricula = ws.cell(i, 16).value  # Col 16 = MATRICULA
    if matricula and str(matricula).strip() in armas_nuevas_buscar:
        nombre = ws.cell(i, 4).value
        marca = ws.cell(i, 14).value
        modelo = ws.cell(i, 15).value
        encontradas.append(matricula)
        print(f"   ✅ {matricula}: {nombre} - {marca} {modelo}")

if len(encontradas) == 4:
    print(f"\n✅✅✅ PERFECTO - Las 4 armas nuevas están presentes")
else:
    print(f"\n⚠️  FALTA: {4 - len(encontradas)} arma(s)")
    faltantes = set(armas_nuevas_buscar) - set(encontradas)
    for f in faltantes:
        print(f"   ❌ {f}")

# Verificar que NO hay duplicados
print(f"\n🔍 VERIFICACIÓN DE DUPLICADOS:")
matriculas = []
for i in range(2, ws.max_row + 1):
    mat = ws.cell(i, 16).value
    if mat:
        matriculas.append(str(mat).strip())

duplicados = [m for m in set(matriculas) if matriculas.count(m) > 1]
if duplicados:
    print(f"   ⚠️  {len(duplicados)} matrículas duplicadas encontradas:")
    for dup in duplicados[:5]:
        print(f"      - {dup}")
else:
    print(f"   ✅ NO hay duplicados")

# Comparación con objetivo
print(f"\n🎯 COMPARACIÓN CON OBJETIVO:")
print(f"   Diciembre 2025: 276 armas (104 cortas + 172 largas)")
print(f"   NUEVA FUENTE:   {total_armas} armas ({armas_cortas} cortas + {armas_largas} largas)")
print(f"   Diferencia:     +{total_armas - 276} armas")

if total_armas == 280 and armas_cortas == 107 and armas_largas == 173:
    print(f"\n✅✅✅ PERFECTO - NUEVA FUENTE DE VERDAD VERIFICADA")
else:
    print(f"\n⚠️  Revisar números")

wb.close()
