#!/usr/bin/env python3
"""
Análisis comparativo CORRECTO entre diciembre 2025 y enero 2026
"""
import openpyxl
from collections import defaultdict

archivo_dic = '/Applications/club-738-web/2025-dic-usb-738/CLUB 738-31-DE-DICIEMBRE-2025_RELACION_SOCIOS_ARMAS_NORMALIZADO.xlsx'
archivo_ene = '/Applications/club-738-web/data/socios/Copy of 2026.31.01_RELACION_SOCIOS_ARMAS_SEPARADO_verified.xlsx'

print("=" * 100)
print("ANÁLISIS COMPARATIVO: DICIEMBRE 2025 (USB 32 ZM) vs ENERO 2026 (ACTUAL)")
print("=" * 100)

# ========== DICIEMBRE 2025 ==========
print("\n📅 DICIEMBRE 2025 - Archivo normalizado (entregado a 32 ZM)")
print("-" * 100)

wb_dic = openpyxl.load_workbook(archivo_dic)
ws_dic = wb_dic.active

# Archivo normalizado tiene headers en fila 1, datos desde fila 2
headers_dic = [cell.value for cell in ws_dic[1]]

total_armas_dic = 0
cortas_dic = 0
largas_dic = 0
armas_dic_set = set()
socios_dic = {}

for i, row in enumerate(ws_dic.iter_rows(min_row=8, values_only=True), 8):
    # Detectar fila de "TOTAL POR PERSONA" (skip)
    if row[5] and 'TOTAL POR PERSONA' in str(row[5]).upper():
        continue
    
    # Detectar fila de arma (Col 9 tiene CLASE)
    if row[8]:  # Col 9 (índice 8)
        clase = str(row[8]).strip().upper()
        marca = str(row[10]).strip() if row[10] else ""
        modelo = str(row[11]).strip() if row[11] else ""
        matricula = str(row[12]).strip() if row[12] else ""
        folio = str(row[13]).strip() if row[13] else ""
        nombre_socio = str(row[2]).strip() if row[2] else ""
        curp = str(row[3]).strip() if row[3] else ""
        
        total_armas_dic += 1
        armas_dic_set.add(matricula)
        
        # Clasificar
        if 'PISTOLA' in clase or 'REVÓLVER' in clase or 'REVOLVER' in clase or 'KIT' in clase:
            cortas_dic += 1
        else:
            largas_dic += 1
        
        # Contar por socio
        if curp and curp != 'None':
            if curp not in socios_dic:
                socios_dic[curp] = {
                    'nombre': nombre_socio,
                    'total': 0,
                    'cortas': 0,
                    'largas': 0,
                    'armas': []
                }
            socios_dic[curp]['total'] += 1
            socios_dic[curp]['armas'].append(matricula)
            if 'PISTOLA' in clase or 'REVÓLVER' in clase or 'REVOLVER' in clase or 'KIT' in clase:
                socios_dic[curp]['cortas'] += 1
            else:
                socios_dic[curp]['largas'] += 1

print(f"📊 TOTALES DICIEMBRE 2025:")
print(f"   Total armas: {total_armas_dic}")
print(f"   Armas cortas: {cortas_dic}")
print(f"   Armas largas: {largas_dic}")
print(f"   Total socios: {len(socios_dic)}")

# Verificar contra lo reportado
print(f"\n✅ VERIFICACIÓN vs Reporte a 32 ZM:")
print(f"   Reportado: 276 armas (104 cortas + 172 largas)")
print(f"   Encontrado: {total_armas_dic} armas ({cortas_dic} cortas + {largas_dic} largas)")
if total_armas_dic == 276 and cortas_dic == 104 and largas_dic == 172:
    print(f"   ✅ COINCIDE EXACTAMENTE")
else:
    print(f"   ⚠️ DIFERENCIA: {total_armas_dic - 276:+d} armas")

# ========== ENERO 2026 ==========
print("\n" + "=" * 100)
print("📅 ENERO 2026 - Archivo actual (post-sincronización)")
print("-" * 100)

wb_ene = openpyxl.load_workbook(archivo_ene)
ws_ene = wb_ene.active

headers_ene = [cell.value for cell in ws_ene[1]]

col_clase_ene = headers_ene.index('CLASE') + 1
col_matricula_ene = headers_ene.index('MATRÍCULA') + 1
col_email_ene = headers_ene.index('EMAIL') + 1
col_nombre_ene = headers_ene.index('NOMBRE DEL SOCIO') + 1
col_curp_ene = headers_ene.index('CURP') + 1

total_armas_ene = 0
cortas_ene = 0
largas_ene = 0
armas_ene_set = set()
socios_ene = {}

for row in ws_ene.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    
    clase = str(row[col_clase_ene - 1]).strip().upper() if row[col_clase_ene - 1] else ""
    matricula = str(row[col_matricula_ene - 1]).strip() if row[col_matricula_ene - 1] else ""
    email = str(row[col_email_ene - 1]).strip().lower() if row[col_email_ene - 1] else ""
    nombre = str(row[col_nombre_ene - 1]).strip() if row[col_nombre_ene - 1] else ""
    curp = str(row[col_curp_ene - 1]).strip() if row[col_curp_ene - 1] else ""
    
    if clase and clase != 'NONE':
        total_armas_ene += 1
        armas_ene_set.add(matricula)
        
        if 'PISTOLA' in clase or 'REVÓLVER' in clase or 'REVOLVER' in clase or 'KIT' in clase:
            cortas_ene += 1
        else:
            largas_ene += 1
        
        # Por CURP
        if curp and curp != 'None':
            if curp not in socios_ene:
                socios_ene[curp] = {
                    'nombre': nombre,
                    'email': email,
                    'total': 0,
                    'cortas': 0,
                    'largas': 0,
                    'armas': []
                }
            socios_ene[curp]['total'] += 1
            socios_ene[curp]['armas'].append(matricula)
            if 'PISTOLA' in clase or 'REVÓLVER' in clase or 'REVOLVER' in clase or 'KIT' in clase:
                socios_ene[curp]['cortas'] += 1
            else:
                socios_ene[curp]['largas'] += 1

print(f"📊 TOTALES ENERO 2026:")
print(f"   Total armas: {total_armas_ene}")
print(f"   Armas cortas: {cortas_ene}")
print(f"   Armas largas: {largas_ene}")
print(f"   Total socios: {len(socios_ene)}")

# ========== DIFERENCIAS ==========
print("\n" + "=" * 100)
print("🔍 ANÁLISIS DE DIFERENCIAS")
print("=" * 100)

diff_total = total_armas_ene - total_armas_dic
diff_cortas = cortas_ene - cortas_dic
diff_largas = largas_ene - largas_dic

print(f"\n📈 CAMBIO EN TOTALES:")
print(f"   Total armas:  {total_armas_dic} → {total_armas_ene} ({diff_total:+d})")
print(f"   Cortas:       {cortas_dic} → {cortas_ene} ({diff_cortas:+d})")
print(f"   Largas:       {largas_dic} → {largas_ene} ({diff_largas:+d})")

# Armas nuevas y eliminadas
nuevas = armas_ene_set - armas_dic_set
eliminadas = armas_dic_set - armas_ene_set

print(f"\n🆕 ARMAS NUEVAS (agregadas entre dic-ene): {len(nuevas)}")
if len(nuevas) <= 30:
    for mat in sorted(nuevas):
        if mat and mat != 'None':
            # Buscar a quién pertenece
            for curp, data in socios_ene.items():
                if mat in data['armas']:
                    print(f"   • {mat:20s} → {data['nombre']}")
                    break
else:
    print(f"   (Demasiadas para listar individualmente)")

print(f"\n❌ ARMAS ELIMINADAS (estaban en dic, no en ene): {len(eliminadas)}")
if len(eliminadas) <= 30:
    for mat in sorted(eliminadas):
        if mat and mat != 'None':
            for curp, data in socios_dic.items():
                if mat in data['armas']:
                    print(f"   • {mat:20s} → {data['nombre']}")
                    break
else:
    print(f"   (Demasiadas para listar)")

# Socios con cambios
print("\n" + "=" * 100)
print("👥 SOCIOS CON CAMBIOS EN ARSENAL (por CURP)")
print("=" * 100)

todos_curps = set(socios_dic.keys()) | set(socios_ene.keys())
cambios = []

for curp in todos_curps:
    dic_data = socios_dic.get(curp, {'nombre': '', 'total': 0, 'cortas': 0, 'largas': 0})
    ene_data = socios_ene.get(curp, {'nombre': '', 'total': 0, 'cortas': 0, 'largas': 0, 'email': ''})
    
    if dic_data['total'] != ene_data['total']:
        cambios.append({
            'curp': curp,
            'nombre': ene_data.get('nombre') or dic_data['nombre'],
            'email': ene_data.get('email', ''),
            'dic_total': dic_data['total'],
            'ene_total': ene_data['total'],
            'diff': ene_data['total'] - dic_data['total']
        })

if cambios:
    # Ordenar por magnitud del cambio
    cambios_ordenados = sorted(cambios, key=lambda x: abs(x['diff']), reverse=True)
    
    print(f"\nTotal de socios con cambios: {len(cambios)}\n")
    
    for c in cambios_ordenados:
        print(f"{c['nombre']}")
        print(f"   CURP: {c['curp']}")
        if c['email']:
            print(f"   Email: {c['email']}")
        print(f"   Diciembre: {c['dic_total']} armas")
        print(f"   Enero: {c['ene_total']} armas")
        print(f"   Cambio: {c['diff']:+d}")
        print()
else:
    print("✅ No hay cambios en arsenales por socio")

print("=" * 100)
