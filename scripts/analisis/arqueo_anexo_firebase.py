#!/usr/bin/env python3
"""
ARQUEO/AUDITORÍA: Anexo A vs Firebase
- Anexo A: Fuente oficial con 76 socios
- Firebase: Identifica socios por EMAIL
"""
import openpyxl
from collections import defaultdict

anexo_a = "2025-dic-usb-738/CLUB 738-31-DE-DICIEMBRE-2025_ANEXOS A, B Y C.xlsx"
nueva_fuente = "data/socios/2026_ENERO_RELACION_SOCIOS_ARMAS_MASTER.xlsx"

print("=" * 80)
print("ARQUEO: ANEXO A (76 SOCIOS) vs FIREBASE")
print("=" * 80)

# 1. LEER ANEXO A - FUENTE OFICIAL
print("\n📋 PASO 1: Leyendo Anexo A (fuente oficial)...")
wb_anexo = openpyxl.load_workbook(anexo_a)
ws_anexo = wb_anexo["Anexo A"]

# Headers en fila 6:
# Col 1: No. DE REGISTRO
# Col 2: NOMBRE SOCIO  
# Col 3: CURP
# Col 4: TELEFONO
# Col 5: EMAIL
# Col 6: No. DE SOCIO (credencial)
# Col 7: ARMAS CORTAS
# Col 8: ARMAS LARGAS
# Col 9: FECHA ALTA

socios_anexo = {}  # key = email
for i in range(7, ws_anexo.max_row + 1):
    email = ws_anexo.cell(i, 5).value
    nombre = ws_anexo.cell(i, 2).value
    
    if email and nombre and "TOTAL" not in str(nombre).upper():
        email_clean = str(email).strip().lower()
        socios_anexo[email_clean] = {
            'fila': i,
            'registro': ws_anexo.cell(i, 1).value,
            'nombre': ws_anexo.cell(i, 2).value,
            'curp': ws_anexo.cell(i, 3).value,
            'telefono': ws_anexo.cell(i, 4).value,
            'email': email_clean,
            'num_socio': ws_anexo.cell(i, 6).value,
            'armas_cortas': ws_anexo.cell(i, 7).value or 0,
            'armas_largas': ws_anexo.cell(i, 8).value or 0,
            'fecha_alta': ws_anexo.cell(i, 9).value
        }

print(f"   ✅ {len(socios_anexo)} socios en Anexo A")

# 2. LEER NUEVA FUENTE (representa lo que está en Firebase)
print("\n📋 PASO 2: Leyendo Nueva Fuente de Verdad (Firebase simulado)...")
wb_nueva = openpyxl.load_workbook(nueva_fuente)
ws_nueva = wb_nueva.active

# Contar armas por email
armas_por_email = defaultdict(lambda: {'cortas': 0, 'largas': 0, 'curp': None, 'nombre': None})

for i in range(2, ws_nueva.max_row + 1):
    curp = ws_nueva.cell(i, 5).value
    nombre = ws_nueva.cell(i, 4).value
    clase = ws_nueva.cell(i, 12).value
    
    if curp:
        # Buscar email de este CURP en Anexo A
        email_encontrado = None
        for email, datos in socios_anexo.items():
            if datos['curp'] and str(datos['curp']).strip() == str(curp).strip():
                email_encontrado = email
                break
        
        if email_encontrado:
            armas_por_email[email_encontrado]['curp'] = curp
            armas_por_email[email_encontrado]['nombre'] = nombre
            
            if clase and clase != "0":
                clase_upper = str(clase).upper()
                if any(x in clase_upper for x in ['PISTOLA', 'REVOLVER', 'KIT']):
                    armas_por_email[email_encontrado]['cortas'] += 1
                else:
                    armas_por_email[email_encontrado]['largas'] += 1

print(f"   ✅ {len(armas_por_email)} socios con datos en Nueva Fuente")

# 3. ARQUEO - COMPARAR
print("\n" + "=" * 80)
print("📊 RESULTADO DEL ARQUEO")
print("=" * 80)

# 3.1 Socios en Anexo A pero NO en Firebase
faltantes_firebase = []
for email, datos in socios_anexo.items():
    if email not in armas_por_email:
        faltantes_firebase.append(datos)

if faltantes_firebase:
    print(f"\n❌ SOCIOS EN ANEXO A PERO NO EN FIREBASE: {len(faltantes_firebase)}")
    print("-" * 80)
    for socio in sorted(faltantes_firebase, key=lambda x: x['num_socio']):
        print(f"   {socio['num_socio']:3s}. {socio['nombre']}")
        print(f"       Email: {socio['email']}")
        print(f"       CURP: {socio['curp']}")
        print(f"       Armas esperadas: {socio['armas_cortas']} cortas + {socio['armas_largas']} largas")
        print()

# 3.2 Socios en Firebase pero NO en Anexo A
extras_firebase = []
for email in armas_por_email:
    if email not in socios_anexo:
        extras_firebase.append(email)

if extras_firebase:
    print(f"\n⚠️  SOCIOS EN FIREBASE PERO NO EN ANEXO A: {len(extras_firebase)}")
    print("-" * 80)
    for email in extras_firebase:
        datos = armas_por_email[email]
        print(f"   {datos['nombre']}")
        print(f"   Email: {email}")
        print(f"   CURP: {datos['curp']}")
        print()

# 3.3 Discrepancias en cantidad de armas
discrepancias = []
for email, datos_anexo in socios_anexo.items():
    if email in armas_por_email:
        datos_firebase = armas_por_email[email]
        
        armas_anexo_total = int(datos_anexo['armas_cortas']) + int(datos_anexo['armas_largas'])
        armas_firebase_total = datos_firebase['cortas'] + datos_firebase['largas']
        
        if armas_anexo_total != armas_firebase_total:
            discrepancias.append({
                'email': email,
                'nombre': datos_anexo['nombre'],
                'num_socio': datos_anexo['num_socio'],
                'anexo_cortas': datos_anexo['armas_cortas'],
                'anexo_largas': datos_anexo['armas_largas'],
                'anexo_total': armas_anexo_total,
                'firebase_cortas': datos_firebase['cortas'],
                'firebase_largas': datos_firebase['largas'],
                'firebase_total': armas_firebase_total,
                'diferencia': armas_firebase_total - armas_anexo_total
            })

if discrepancias:
    print(f"\n⚠️  DISCREPANCIAS EN CANTIDAD DE ARMAS: {len(discrepancias)}")
    print("-" * 80)
    for d in sorted(discrepancias, key=lambda x: abs(x['diferencia']), reverse=True):
        print(f"   {d['num_socio']:3s}. {d['nombre']}")
        print(f"       Email: {d['email']}")
        print(f"       Anexo A:  {d['anexo_cortas']} cortas + {d['anexo_largas']} largas = {d['anexo_total']} TOTAL")
        print(f"       Firebase: {d['firebase_cortas']} cortas + {d['firebase_largas']} largas = {d['firebase_total']} TOTAL")
        print(f"       Diferencia: {d['diferencia']:+d}")
        print()

# 4. RESUMEN FINAL
print("\n" + "=" * 80)
print("📊 RESUMEN FINAL DEL ARQUEO")
print("=" * 80)
print(f"Socios en Anexo A (oficial):           {len(socios_anexo)}")
print(f"Socios en Firebase:                    {len(armas_por_email)}")
print(f"Socios faltantes en Firebase:          {len(faltantes_firebase)}")
print(f"Socios extras en Firebase:             {len(extras_firebase)}")
print(f"Socios con discrepancias de armas:     {len(discrepancias)}")

if len(faltantes_firebase) == 0 and len(extras_firebase) == 0 and len(discrepancias) == 0:
    print("\n✅✅✅ ARQUEO PERFECTO - TODO COINCIDE")
else:
    print(f"\n⚠️  SE ENCONTRARON {len(faltantes_firebase) + len(extras_firebase) + len(discrepancias)} INCONSISTENCIAS")

# Guardar reporte
with open('ARQUEO_ANEXO_A_VS_FIREBASE.txt', 'w', encoding='utf-8') as f:
    f.write("ARQUEO: ANEXO A vs FIREBASE\n")
    f.write("=" * 80 + "\n\n")
    
    if faltantes_firebase:
        f.write(f"SOCIOS FALTANTES EN FIREBASE: {len(faltantes_firebase)}\n")
        f.write("-" * 80 + "\n")
        for socio in sorted(faltantes_firebase, key=lambda x: x['num_socio']):
            f.write(f"{socio['num_socio']}. {socio['nombre']}\n")
            f.write(f"   Email: {socio['email']}\n")
            f.write(f"   CURP: {socio['curp']}\n")
            f.write(f"   Armas: {socio['armas_cortas']} cortas + {socio['armas_largas']} largas\n\n")
    
    if discrepancias:
        f.write(f"\nDISCREPANCIAS DE ARMAS: {len(discrepancias)}\n")
        f.write("-" * 80 + "\n")
        for d in sorted(discrepancias, key=lambda x: abs(x['diferencia']), reverse=True):
            f.write(f"{d['num_socio']}. {d['nombre']}\n")
            f.write(f"   Anexo A: {d['anexo_total']} armas | Firebase: {d['firebase_total']} armas | Dif: {d['diferencia']:+d}\n\n")

print(f"\n💾 Reporte guardado en: ARQUEO_ANEXO_A_VS_FIREBASE.txt")

wb_anexo.close()
wb_nueva.close()
