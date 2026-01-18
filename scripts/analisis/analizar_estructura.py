#!/usr/bin/env python3
import openpyxl

archivo = '/Applications/club-738-web/2025-dic-usb-738/CLUB 738-31-DE-DICIEMBRE-2025_RELACION_SOCIOS_ARMAS.xlsx'
wb = openpyxl.load_workbook(archivo)
ws = wb['CLUB 738. RELACION SOCIOS 31 DI']

print('ANÁLISIS DE ESTRUCTURA DE FILAS:')
print('=' * 100)

# Contar filas con CLASE
count_armas = 0
count_socios_con_nombre = 0

for i in range(8, ws.max_row + 1):
    row = list(ws[i])
    col3 = row[2].value
    col6 = row[5].value
    col9 = row[8].value
    
    # Skip "TOTAL POR PERSONA"
    if col6 and 'TOTAL' in str(col6).upper():
        continue
    
    # Contar armas (tienen CLASE en col9)
    if col9:
        count_armas += 1
    
    # Contar filas con nombre (col3 tiene valor Y NO tiene col9)
    if col3 and not col9:
        count_socios_con_nombre += 1

print(f'Total filas con CLASE (armas): {count_armas}')
print(f'Total filas con NOMBRE pero sin CLASE (socios): {count_socios_con_nombre}')

# Ahora buscar filas problemáticas: tienen CLASE pero TAMBIÉN tienen NOMBRE en col3
print('\n⚠️ FILAS PROBLEMÁTICAS (tienen CLASE Y NOMBRE):')
problemas = 0
for i in range(8, ws.max_row + 1):
    row = list(ws[i])
    col3 = row[2].value
    col9 = row[8].value
    
    if col3 and col9:
        problemas += 1
        if problemas <= 20:
            col1 = str(row[0].value)[:10] if row[0].value else 'None'
            col3_str = str(col3)[:40]
            col9_str = str(col9)[:20]
            print(f'Fila {i}: Col1={col1} | Nombre={col3_str} | Clase={col9_str}')

print(f'\nTotal filas problemáticas: {problemas}')

print('\n' + '=' * 100)
print('CONCLUSIÓN:')
if problemas > 0:
    print(f'❌ HAY {problemas} FILAS donde el NOMBRE y la CLASE están en la MISMA FILA')
    print('   Esto significa que NO todas las armas están en filas separadas del nombre')
else:
    print('✅ TODAS las armas están en filas separadas del nombre del socio')
