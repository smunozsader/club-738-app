#!/usr/bin/env python3
"""
Identificar socio sin dirección estructurada
"""
import openpyxl

archivo = "data/socios/2026_ENERO_RELACION_SOCIOS_ARMAS_MASTER.xlsx"

print("=" * 80)
print("BUSCANDO SOCIO SIN DIRECCIÓN ESTRUCTURADA")
print("=" * 80)

wb = openpyxl.load_workbook(archivo)
ws = wb.active

# Revisar cada socio (por CURP único)
socios_revisados = set()

for i in range(2, ws.max_row + 1):
    curp = ws.cell(i, 5).value
    nombre = ws.cell(i, 4).value
    calle = ws.cell(i, 7).value
    colonia = ws.cell(i, 8).value
    ciudad = ws.cell(i, 9).value
    
    if curp and curp not in socios_revisados:
        socios_revisados.add(curp)
        
        if not calle and not colonia:
            print(f"\n❌ SOCIO SIN DIRECCIÓN:")
            print(f"   Nombre: {nombre}")
            print(f"   CURP: {curp}")
            print(f"   Fila: {i}")
            print(f"   Calle: {calle}")
            print(f"   Colonia: {colonia}")
            print(f"   Ciudad: {ciudad}")

print(f"\n📊 Total socios revisados: {len(socios_revisados)}")

wb.close()
