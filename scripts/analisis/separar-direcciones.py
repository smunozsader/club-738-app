#!/usr/bin/env python3

"""
Script: Separar direcciones de socios en columnas normalizadas

Archivo: 2025.31.12_RELACION_SOCIOS_ARMAS copia con direccion.xlsx
Sheet: Sheet1

Entrada: Columna A con direcciones completas
Salida: Columnas A-F con datos separados (CALLE, COLONIA, CIUDAD, MUNICIPIO, ESTADO, CP)

Formato típico de direcciones:
- CALLE 44 No. 438 x 21 y 23, FRACC. LOS PINOS, MÉRIDA, YUC., CP 97138
- Componentes separados por comas

Uso:
   python3 scripts/separar-direcciones.py
"""

import openpyxl
import re
from pathlib import Path

# ============================================================================
# FUNCIONES DE PARSING
# ============================================================================

def limpiar_texto(texto):
    """Limpia espacios extra y normaliza."""
    if not texto:
        return ""
    return re.sub(r'\s+', ' ', texto.strip())

def extraer_codigo_postal(texto):
    """Extrae código postal (5 dígitos)."""
    match = re.search(r'CP\s*(\d{5})', texto, re.IGNORECASE)
    if match:
        return match.group(1)
    
    # Buscar 5 dígitos al final
    match = re.search(r'(\d{5})\s*$', texto)
    if match:
        return match.group(1)
    
    return ""

def separar_direccion(direccion_completa):
    """
    Separa una dirección completa en componentes.
    
    Retorna:
        dict con: calle, colonia, ciudad, municipio, estado, cp
    """
    # Inicializar todas las variables
    calle = ''
    colonia = ''
    ciudad = ''
    municipio = ''
    estado = ''
    cp = ''
    
    if not direccion_completa:
        return {
            'calle': calle,
            'colonia': colonia,
            'ciudad': ciudad,
            'municipio': municipio,
            'estado': estado,
            'cp': cp
        }
    
    direccion = str(direccion_completa).strip()
    
    # Extraer código postal
    cp = extraer_codigo_postal(direccion)
    
    # Remover código postal del texto
    direccion_sin_cp = re.sub(r',?\s*C\.?P\.?\s*\d{5}', '', direccion, flags=re.IGNORECASE)
    direccion_sin_cp = re.sub(r'\d{5}\s*$', '', direccion_sin_cp)
    
    # Dividir por comas
    partes = [limpiar_texto(p) for p in direccion_sin_cp.split(',')]
    partes = [p for p in partes if p]  # Remover vacíos
    
    if len(partes) == 0:
        pass
    elif len(partes) == 1:
        # Solo calle
        calle = partes[0]
    elif len(partes) == 2:
        # CALLE, CIUDAD
        calle = partes[0]
        # Intentar separar CIUDAD, ESTADO
        if 'YUC' in partes[1].upper() or 'CAMP' in partes[1].upper():
            ciudad_estado = partes[1].split()
            if len(ciudad_estado) >= 2:
                ciudad = ' '.join(ciudad_estado[:-1])
                estado = ciudad_estado[-1].rstrip('.,')
            else:
                ciudad = partes[1]
        else:
            ciudad = partes[1]
    elif len(partes) == 3:
        # CALLE, COLONIA, CIUDAD-ESTADO
        # O podría ser: CALLE, CIUDAD, MUNICIPIO-ESTADO
        calle = partes[0]
        
        # Verificar si parte[2] tiene 2 palabras separadas por espacio (CIUDAD ESTADO)
        ultima = partes[2]
        palabras_ultima = ultima.split()
        
        # Si la última parte tiene estado (YUC, CAMP, etc.)
        if any(est in ultima.upper() for est in ['YUC', 'CAMP', 'Q.ROO', 'CDMX']):
            # Puede ser: COLONIA, CIUDAD, ESTADO o CIUDAD, MUNICIPIO, ESTADO
            # Si partes[1] es una localidad pequeña y última tiene municipio
            # Ejemplo: BECANCHEN, TEKAX, YUC
            if len(palabras_ultima) >= 2:
                posible_municipio = ' '.join(palabras_ultima[:-1])
                estado = palabras_ultima[-1].rstrip('.,')
                
                # Si posible_municipio es conocido, entonces partes[1] es ciudad
                if posible_municipio.upper() in ['TEKAX', 'VALLADOLID', 'TIZIMÍN', 'MÉRIDA']:
                    ciudad = partes[1]  # BECANCHEN
                    municipio = posible_municipio  # TEKAX
                else:
                    # Es COLONIA, CIUDAD, ESTADO
                    colonia = partes[1]
                    ciudad = posible_municipio
            else:
                ciudad = partes[1]
        else:
            # CALLE, COLONIA, CIUDAD
            colonia = partes[1]
            ciudad = ultima
            
    elif len(partes) >= 4:
        # CALLE, COLONIA, CIUDAD, ESTADO o más
        # O CALLE, CIUDAD, MUNICIPIO, ESTADO (sin colonia)
        calle = partes[0]
        
        # Normalizar última parte (estado)
        ultima_parte = partes[-1].upper().strip().rstrip('.,')
        
        if 'YUC' in ultima_parte:
            estado = 'YUCATÁN'
        elif 'CAMP' in ultima_parte:
            estado = 'CAMPECHE'
        elif 'Q.ROO' in ultima_parte or 'QUINTANA ROO' in ultima_parte:
            estado = 'QUINTANA ROO'
        else:
            estado = ultima_parte
        
        if len(partes) == 4:
            # Determinar si es:
            # Caso A: CALLE, COLONIA, CIUDAD, ESTADO
            # Caso B: CALLE, CIUDAD, MUNICIPIO, ESTADO
            
            # Detectar colonia markers
            parte1_upper = partes[1].upper().strip()
            es_colonia = any(marcador in parte1_upper for marcador in 
                             ['FRACC.', 'FRACC', 'FRACCIONAMIENTO', 'COL.', 'COL', 'COLONIA',
                              'RESIDENCIAL', 'PRIVADA', 'PRIVADA.', 'PRIV.', 'COTO'])
            
            if es_colonia:
                # Caso A: CALLE, COLONIA, CIUDAD, ESTADO
                colonia = partes[1]
                ciudad = partes[2]
                # Municipio = ciudad si es cabecera homónima
                ciudad_upper = partes[2].upper().strip()
                cabeceras_homonimas = ['MÉRIDA', 'VALLADOLID', 'TEKAX', 'TIZIMÍN', 'PROGRESO',
                                       'MOTUL', 'TICUL', 'UMÁN', 'CAMPECHE', 'CHAMPOTÓN',
                                       'HECELCHAKÁN', 'CALKINÍ', 'HOPELCHÉN', 'OXKUTZCAB',
                                       'CONKAL', 'KANASÍN', 'IZAMAL']
                if ciudad_upper in cabeceras_homonimas:
                    municipio = ciudad
            else:
                # Verificar si partes[2] es municipio conocido
                parte2_upper = partes[2].upper().strip()
                municipios_conocidos = ['TEKAX', 'VALLADOLID', 'TIZIMÍN', 'MÉRIDA', 'PROGRESO', 
                                         'MOTUL', 'TICUL', 'UMÁN', 'OXKUTZCAB', 'CONKAL',
                                         'KANASÍN', 'CAMPECHE', 'CARMEN', 'CALKINÍ', 'CHAMPOTÓN',
                                         'HOPELCHÉN', 'CANDELARIA', 'HECELCHAKÁN', 'CAUCEL']
                
                if parte2_upper in municipios_conocidos:
                    # Caso B: CALLE, CIUDAD, MUNICIPIO, ESTADO
                    ciudad = partes[1]
                    municipio = partes[2]
                else:
                    # Por defecto: CALLE, COLONIA, CIUDAD, ESTADO
                    colonia = partes[1]
                    ciudad = partes[2]
        
        elif len(partes) >= 5:
            # CALLE, COLONIA, CIUDAD, MUNICIPIO, ESTADO
            colonia = partes[1]
            ciudad = partes[2]
            municipio = partes[3]
    
    # Normalizar estado
    estado_map = {
        'YUC': 'YUCATÁN',
        'YUC.': 'YUCATÁN',
        'YUCATÁN': 'YUCATÁN',
        'CAMP': 'CAMPECHE',
        'CAMP.': 'CAMPECHE',
        'CAMPECHE': 'CAMPECHE',
        'Q.ROO': 'QUINTANA ROO',
        'Q. ROO': 'QUINTANA ROO',
        'QUINTANA ROO': 'QUINTANA ROO',
        'CDMX': 'CIUDAD DE MÉXICO',
        'CIUDAD DE MÉXICO': 'CIUDAD DE MÉXICO'
    }
    
    estado_upper = estado.upper().strip()
    estado = estado_map.get(estado_upper, estado)
    
    # Normalizar ciudad
    ciudad_normalizada = ciudad.upper().strip()
    
    # SOLO duplicar ciudad en municipio si:
    # 1. Municipio está vacío
    # 2. Ciudad es una cabecera municipal conocida (no una localidad)
    
    # Cabeceras municipales homónimas de Yucatán (ciudad = municipio)
    cabeceras_yuc = {
        'MÉRIDA': 'MÉRIDA',
        'MERIDA': 'MÉRIDA',
        'VALLADOLID': 'VALLADOLID',
        'TIZIMÍN': 'TIZIMÍN',
        'TIZIMIN': 'TIZIMÍN',
        'PROGRESO': 'PROGRESO',
        'MOTUL': 'MOTUL',
        'TICUL': 'TICUL',
        'UMÁN': 'UMÁN',
        'UMAN': 'UMÁN',
        'OXKUTZCAB': 'OXKUTZCAB',
        'TEKAX': 'TEKAX',  # Solo si dice TEKAX, no BECANCHEN
        'IZAMAL': 'IZAMAL',
        'KANASÍN': 'KANASÍN',
        'KANASIN': 'KANASÍN'
    }
    
    # Cabeceras de Campeche
    cabeceras_camp = {
        'CAMPECHE': 'CAMPECHE',
        'SAN FRANCISCO DE CAMPECHE': 'CAMPECHE',
        'CD. DEL CARMEN': 'CARMEN',
        'CIUDAD DEL CARMEN': 'CARMEN',
        'CARMEN': 'CARMEN',
        'CALKINÍ': 'CALKINÍ',
        'CALKINI': 'CALKINÍ',
        'CHAMPOTÓN': 'CHAMPOTÓN',
        'CHAMPOTON': 'CHAMPOTÓN',
        'HOPELCHÉN': 'HOPELCHÉN',
        'HOPELCHEN': 'HOPELCHÉN'
    }
    
    # Cabeceras de Quintana Roo
    cabeceras_qroo = {
        'CHETUMAL': 'OTHÓN P. BLANCO',
        'CANCÚN': 'BENITO JUÁREZ',
        'CANCUN': 'BENITO JUÁREZ',
        'PLAYA DEL CARMEN': 'SOLIDARIDAD',
        'COZUMEL': 'COZUMEL'
    }
    
    # Alcaldías de CDMX (duplicar ciudad en municipio)
    alcaldias_cdmx = [
        'MIGUEL HIDALGO', 'BENITO JUÁREZ', 'CUAUHTÉMOC', 'COYOACÁN',
        'ÁLVARO OBREGÓN', 'TLALPAN', 'IZTAPALAPA', 'GUSTAVO A. MADERO',
        'VENUSTIANO CARRANZA', 'AZCAPOTZALCO', 'IZTACALCO', 'MAGDALENA CONTRERAS',
        'CUAJIMALPA', 'TLÁHUAC', 'XOCHIMILCO', 'MILPA ALTA'
    ]
    
    # Solo asignar municipio si está vacío Y la ciudad es cabecera
    if not municipio or municipio.strip() == '':
        # Normalizar ciudad
        if ciudad_normalizada in cabeceras_yuc:
            ciudad = cabeceras_yuc[ciudad_normalizada]
            municipio = ciudad
        elif ciudad_normalizada in cabeceras_camp:
            ciudad_norm = ciudad_normalizada
            # Para Carmen, mantener "CD. DEL CARMEN" como ciudad
            if 'CARMEN' in ciudad_norm:
                ciudad = 'CD. DEL CARMEN'
            municipio = cabeceras_camp[ciudad_norm]
        elif ciudad_normalizada in cabeceras_qroo:
            municipio = cabeceras_qroo[ciudad_normalizada]
        elif ciudad_normalizada in alcaldias_cdmx:
            # Para CDMX, duplicar alcaldía en municipio
            municipio = ciudad
    
    return {
        'calle': calle,
        'colonia': colonia,
        'ciudad': ciudad,
        'municipio': municipio,
        'estado': estado,
        'cp': cp
    }

# ============================================================================
# PROCESAMIENTO DEL EXCEL
# ============================================================================

def procesar_excel():
    """Procesa el Excel y separa las direcciones."""
    
    archivo_path = '/Applications/club-738-web/data/socios/2025.31.12_RELACION_SOCIOS_ARMAS copia con direccion.xlsx'
    
    print('📖 Cargando archivo Excel...')
    wb = openpyxl.load_workbook(archivo_path)
    ws = wb['Sheet1']
    
    print(f'✅ Archivo cargado: {ws.max_row} filas\n')
    
    # Procesar cada fila
    print('🔄 Procesando direcciones...\n')
    procesadas = 0
    errores = 0
    
    for row_num in range(2, ws.max_row + 1):
        direccion_completa = ws.cell(row_num, 1).value  # Columna A
        
        if not direccion_completa:
            continue
        
        try:
            partes = separar_direccion(direccion_completa)
            
            # Escribir en columnas
            ws.cell(row_num, 1).value = partes['calle']         # A: CALLE
            ws.cell(row_num, 2).value = partes['colonia']       # B: COLONIA
            ws.cell(row_num, 3).value = partes['ciudad']        # C: CIUDAD
            ws.cell(row_num, 4).value = partes['municipio']     # D: MUNICIPIO
            ws.cell(row_num, 5).value = partes['estado']        # E: ESTADO
            ws.cell(row_num, 6).value = partes['cp']            # F: CODIGO POSTAL
            
            procesadas += 1
            
            # Mostrar progreso cada 50 filas
            if procesadas % 50 == 0:
                print(f'   ✓ {procesadas} direcciones procesadas...')
                
        except Exception as e:
            print(f'   ⚠️  Error en fila {row_num}: {e}')
            errores += 1
    
    print(f'\n✅ Procesamiento completo:')
    print(f'   - Direcciones procesadas: {procesadas}')
    print(f'   - Errores: {errores}')
    
    # Guardar archivo
    print(f'\n💾 Guardando archivo...')
    wb.save(archivo_path)
    print(f'✅ Archivo guardado: {archivo_path}\n')
    
    # Mostrar muestra de resultados
    print('📋 Muestra de resultados (primeras 5 filas):')
    print('-' * 130)
    print(f"{'CALLE':<35} | {'COLONIA':<20} | {'CIUDAD':<12} | {'MUNICIPIO':<12} | {'ESTADO':<10} | {'CP':<6}")
    print('-' * 130)
    
    for row_num in range(2, min(7, ws.max_row + 1)):
        calle = ws.cell(row_num, 1).value or ''
        colonia = ws.cell(row_num, 2).value or ''
        ciudad = ws.cell(row_num, 3).value or ''
        municipio = ws.cell(row_num, 4).value or ''
        estado = ws.cell(row_num, 5).value or ''
        cp = ws.cell(row_num, 6).value or ''
        
        print(f"{calle[:34]:<35} | {colonia[:19]:<20} | {ciudad[:11]:<12} | {municipio[:11]:<12} | {estado[:9]:<10} | {cp:<6}")
    
    print('-' * 130)

# ============================================================================
# MAIN
# ============================================================================

def main():
    try:
        print('🚀 SEPARACIÓN DE DIRECCIONES')
        print('=' * 50 + '\n')
        
        procesar_excel()
        
        print('\n✅ ¡PROCESO COMPLETADO!\n')
        print('📝 Pasos siguientes:')
        print('   1. Revisa Sheet1 en el archivo Excel')
        print('   2. Verifica que las direcciones estén correctamente separadas')
        print('   3. Si todo es correcto, copia y pega a CLUB 738. RELACION SOCIOS 31 DI')
        
    except Exception as err:
        print(f'\n❌ ERROR: {err}')
        import traceback
        traceback.print_exc()
        return 1
    
    return 0

if __name__ == '__main__':
    exit(main())
