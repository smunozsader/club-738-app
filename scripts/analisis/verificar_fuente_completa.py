#!/usr/bin/env python3
"""
Verificar que la nueva fuente de verdad quedó correcta
"""
import openpyxl

archivo = "data/socios/2026_ENERO_FUENTE_VERDAD_COMPLETA_76_SOCIOS.xlsx"

print("=" * 80)
print("VERIFICACIÓN: FUENTE DE VERDAD COMPLETA - 76 SOCIOS")
print("=" * 80)

wb = openpyxl.load_workbook(archivo)
ws = wb.active

# Verificar corrección de Agustín
print("\n✅ 1. VERIFICANDO CORRECCIÓN DE AGUSTÍN MORENO:")
print("-" * 80)

for i in range(2, ws.max_row + 1):
    nombre = ws.cell(i, 4).value
    if nombre and "AGUSTIN MORENO" in str(nombre).upper():
        print(f"\nFila {i}:")
        print(f"  Nombre: {nombre}")
        print(f"  CURP: {ws.cell(i, 5).value}")
        print(f"  Teléfono: {ws.cell(i, 6).value}")
        print(f"  Email: {ws.cell(i, 7).value}")
        break

# Estadísticas generales
print("\n\n📊 2. ESTADÍSTICAS GENERALES:")
print("-" * 80)

curps_unicos = set()
emails_unicos = set()
armas_count = 0
socios_sin_armas = 0

for i in range(2, ws.max_row + 1):
    curp = ws.cell(i, 5).value
    email = ws.cell(i, 7).value
    clase = ws.cell(i, 14).value
    
    if curp:
        curps_unicos.add(curp)
    if email:
        emails_unicos.add(email.lower().strip())
    
    if clase and clase != "0":
        armas_count += 1
    elif clase == "0":
        socios_sin_armas += 1

print(f"Total filas (sin header): {ws.max_row - 1}")
print(f"Socios únicos (CURP): {len(curps_unicos)}")
print(f"Emails únicos: {len(emails_unicos)}")
print(f"Armas registradas: {armas_count}")
print(f"Socios sin armas (clase=0): {socios_sin_armas}")

# Verificar emails duplicados
print("\n\n📧 3. VERIFICANDO EMAILS DUPLICADOS:")
print("-" * 80)

emails_dict = {}
for i in range(2, ws.max_row + 1):
    email = ws.cell(i, 7).value
    nombre = ws.cell(i, 4).value
    if email:
        email_clean = email.lower().strip()
        if email_clean not in emails_dict:
            emails_dict[email_clean] = []
        if nombre not in emails_dict[email_clean]:
            emails_dict[email_clean].append(nombre)

duplicados = {k: v for k, v in emails_dict.items() if len(v) > 1}
if duplicados:
    print(f"⚠️  {len(duplicados)} emails duplicados encontrados:")
    for email, nombres in duplicados.items():
        print(f"\n  {email}:")
        for nombre in nombres:
            print(f"    - {nombre}")
else:
    print("✅ NO hay emails duplicados")

# Verificar columnas completas
print("\n\n📋 4. VERIFICANDO COLUMNAS CLAVE:")
print("-" * 80)

headers = [ws.cell(1, j).value for j in range(1, ws.max_column + 1)]
print(f"Total columnas: {len(headers)}")
print(f"\nHeaders:")
for i, h in enumerate(headers, 1):
    print(f"  {i:2d}. {h}")

wb.close()

print("\n" + "=" * 80)
print("✅ VERIFICACIÓN COMPLETADA")
print("=" * 80)
