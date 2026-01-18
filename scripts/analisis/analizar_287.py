#!/usr/bin/env python3
"""
Análisis: ¿De dónde salió la cifra de 287 armas?
"""
import openpyxl

# Archivo actual "fuente de verdad"
archivo_actual = "data/socios/Copy of 2026.31.01_RELACION_SOCIOS_ARMAS_SEPARADO_verified.xlsx"

wb = openpyxl.load_workbook(archivo_actual)
ws = wb.active

print("=" * 80)
print("ANÁLISIS: ¿De dónde salió 287 armas?")
print("=" * 80)

# Encontrar headers - intentar varias columnas
header_row = None
for i in range(1, 20):
    # Probar col 9 (CLASE)
    cell_val = ws.cell(i, 9).value
    if cell_val and 'CLASE' in str(cell_val).upper():
        header_row = i
        break
    # Probar col 1
    cell_val = ws.cell(i, 1).value
    if cell_val and 'CREDENCIAL' in str(cell_val).upper():
        header_row = i
        break

if header_row is None:
    # Si no se encuentra, asumir fila 1
    print("⚠️  No se encontraron headers, asumiendo fila 1")
    header_row = 1
    # Mostrar primeras filas
    for i in range(1, 5):
        for j in range(1, 10):
            val = ws.cell(i, j).value
            if val:
                print(f"  Fila {i}, Col {j}: {val}")

print(f"\n📋 Headers encontrados en fila: {header_row}")
print(f"📊 Total de filas en Excel: {ws.max_row}")

# Contar filas con datos de armas (tienen CLASE)
total_filas = 0
filas_con_clase = 0
filas_sin_clase = 0
clases_vacias = 0

for i in range(header_row + 1, ws.max_row + 1):
    row = [ws.cell(i, j).value for j in range(1, ws.max_column + 1)]
    
    # Verificar si tiene algún dato
    if any(cell for cell in row):
        total_filas += 1
        
        # Col 9 = CLASE (índice 8)
        clase = row[8] if len(row) > 8 else None
        
        if clase and str(clase).strip():
            filas_con_clase += 1
        else:
            filas_sin_clase += 1
            # Mostrar qué tiene en otras columnas
            nombre = row[2] if len(row) > 2 else None
            if nombre:
                print(f"  Fila {i}: Tiene NOMBRE pero no CLASE → {nombre}")

print(f"\n📈 CONTEO DE FILAS:")
print(f"   Total de filas con datos: {total_filas}")
print(f"   Filas con CLASE (armas): {filas_con_clase}")
print(f"   Filas sin CLASE (socios sin armas): {filas_sin_clase}")

print(f"\n💡 CONCLUSIÓN:")
print(f"   El archivo tiene {ws.max_row} filas totales")
print(f"   De las cuales {filas_con_clase} tienen armas registradas")
print(f"   Y {filas_sin_clase} son socios sin armas o filas de separación")

print(f"\n🎯 El número 287 probablemente vino de:")
print(f"   - Conteo inicial antes de limpieza de duplicados")
print(f"   - O conteo de todas las filas (no solo las que tienen CLASE)")

wb.close()
