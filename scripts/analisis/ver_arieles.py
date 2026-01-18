#!/usr/bin/env python3
import openpyxl

archivo = 'data/socios/2026_ENERO_FUENTE_VERDAD_COMPLETA_76_SOCIOS.xlsx'
wb = openpyxl.load_workbook(archivo)
ws = wb.active

print('TELÉFONOS DE LOS DOS ARIELES:')
print('=' * 80)

arieles = []
for i in range(2, ws.max_row + 1):
    nombre = ws.cell(i, 4).value
    curp = ws.cell(i, 5).value
    
    if nombre and 'ARIEL' in str(nombre).upper() and curp:
        if curp not in [a['curp'] for a in arieles]:
            arieles.append({
                'nombre': nombre,
                'curp': curp,
                'credencial': ws.cell(i, 3).value,
                'telefono': ws.cell(i, 6).value,
                'email': ws.cell(i, 7).value
            })

for i, a in enumerate(arieles, 1):
    print(f'\n{i}. {a["nombre"]}')
    print(f'   No. Credencial: {a["credencial"]}')
    print(f'   CURP: {a["curp"]}')
    print(f'   Teléfono: {a["telefono"]}')
    print(f'   Email: {a["email"]}')

print(f'\n\nTotal Arieles: {len(arieles)}')

if len(arieles) == 2:
    if arieles[0]['telefono'] == arieles[1]['telefono']:
        print('\n⚠️  ¡TELÉFONO DUPLICADO!')
        print(f'   Ambos tienen: {arieles[0]["telefono"]}')
    else:
        print('\n✅ Teléfonos diferentes')

wb.close()
