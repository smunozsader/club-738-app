#!/usr/bin/env python3
"""
Análisis comparativo FINAL: Diciembre 2025 (normalizado) vs Enero 2026
"""
import openpyxl

archivo_dic = '/Applications/club-738-web/2025-dic-usb-738/CLUB 738-31-DE-DICIEMBRE-2025_RELACION_SOCIOS_ARMAS_NORMALIZADO.xlsx'
archivo_ene = '/Applications/club-738-web/data/socios/Copy of 2026.31.01_RELACION_SOCIOS_ARMAS_SEPARADO_verified.xlsx'

print("=" * 100)
print("ANÁLISIS COMPARATIVO FINAL: DICIEMBRE 2025 vs ENERO 2026")
print("=" * 100)

# DICIEMBRE 2025 (normalizado)
wb_dic = openpyxl.load_workbook(archivo_dic)
ws_dic = wb_dic.active

headers_dic = [cell.value for cell in ws_dic[1]]
col_clase_d = headers_dic.index('CLASE') + 1
col_mat_d = headers_dic.index('MATRÍCULA') + 1
col_curp_d = headers_dic.index('CURP') + 1
# Buscar nombre con o sin salto de línea
try:
    col_nombre_d = headers_dic.index('NOMBRE DEL SOCIO (No. CREDENCIAL)') + 1
except ValueError:
    col_nombre_d = headers_dic.index('NOMBRE DEL SOCIO \n(No. CREDENCIAL)') + 1

total_dic = 0
cortas_dic = 0
largas_dic = 0
armas_dic = set()
socios_dic = {}

for row in ws_dic.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    
    clase = str(row[col_clase_d - 1]).upper() if row[col_clase_d - 1] else ""
    mat = str(row[col_mat_d - 1]).strip() if row[col_mat_d - 1] else ""
    curp = str(row[col_curp_d - 1]).strip() if row[col_curp_d - 1] else ""
    nombre = str(row[col_nombre_d - 1]).strip() if row[col_nombre_d - 1] else ""
    
    if clase and clase != 'NONE':
        total_dic += 1
        armas_dic.add(mat)
        
        if 'PISTOLA' in clase or 'REVOLVER' in clase or 'KIT' in clase:
            cortas_dic += 1
        else:
            largas_dic += 1
        
        if curp and curp != 'None':
            if curp not in socios_dic:
                socios_dic[curp] = {'nombre': nombre, 'total': 0, 'armas': []}
            socios_dic[curp]['total'] += 1
            socios_dic[curp]['armas'].append(mat)

print(f"\n📅 DICIEMBRE 2025 (normalizado):")
print(f"   Total armas: {total_dic}")
print(f"   Cortas: {cortas_dic}")
print(f"   Largas: {largas_dic}")
print(f"   Socios: {len(socios_dic)}")

# ENERO 2026
wb_ene = openpyxl.load_workbook(archivo_ene)
ws_ene = wb_ene.active

headers_ene = [cell.value for cell in ws_ene[1]]
col_clase_e = headers_ene.index('CLASE') + 1
col_mat_e = headers_ene.index('MATRÍCULA') + 1
col_curp_e = headers_ene.index('CURP') + 1
col_nombre_e = headers_ene.index('NOMBRE DEL SOCIO') + 1

total_ene = 0
cortas_ene = 0
largas_ene = 0
armas_ene = set()
socios_ene = {}

for row in ws_ene.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    
    clase = str(row[col_clase_e - 1]).upper() if row[col_clase_e - 1] else ""
    mat = str(row[col_mat_e - 1]).strip() if row[col_mat_e - 1] else ""
    curp = str(row[col_curp_e - 1]).strip() if row[col_curp_e - 1] else ""
    nombre = str(row[col_nombre_e - 1]).strip() if row[col_nombre_e - 1] else ""
    
    if clase and clase != 'NONE':
        total_ene += 1
        armas_ene.add(mat)
        
        if 'PISTOLA' in clase or 'REVOLVER' in clase or 'KIT' in clase:
            cortas_ene += 1
        else:
            largas_ene += 1
        
        if curp and curp != 'None':
            if curp not in socios_ene:
                socios_ene[curp] = {'nombre': nombre, 'total': 0, 'armas': []}
            socios_ene[curp]['total'] += 1
            socios_ene[curp]['armas'].append(mat)

print(f"\n📅 ENERO 2026 (actual):")
print(f"   Total armas: {total_ene}")
print(f"   Cortas: {cortas_ene}")
print(f"   Largas: {largas_ene}")
print(f"   Socios: {len(socios_ene)}")

# DIFERENCIAS
nuevas = armas_ene - armas_dic
eliminadas = armas_dic - armas_ene

print(f"\n" + "=" * 100)
print("🔍 DIFERENCIAS")
print("=" * 100)
print(f"\n📈 CAMBIO: {total_dic} → {total_ene} ({total_ene - total_dic:+d} armas)")
print(f"   Cortas: {cortas_dic} → {cortas_ene} ({cortas_ene - cortas_dic:+d})")
print(f"   Largas: {largas_dic} → {largas_ene} ({largas_ene - largas_dic:+d})")

print(f"\n🆕 ARMAS NUEVAS ({len(nuevas)}):")
for mat in sorted(nuevas):
    if mat and mat != 'None':
        # Buscar socio
        for curp, data in socios_ene.items():
            if mat in data['armas']:
                print(f"   • {mat:25s} → {data['nombre']}")
                break

print(f"\n❌ ARMAS ELIMINADAS ({len(eliminadas)}):")
for mat in sorted(eliminadas):
    if mat and mat != 'None':
        for curp, data in socios_dic.items():
            if mat in data['armas']:
                print(f"   • {mat:25s} → {data['nombre']}")
                break

# Socios con cambios
print(f"\n" + "=" * 100)
print("👥 SOCIOS CON CAMBIOS")
print("=" * 100)

todos_curps = set(socios_dic.keys()) | set(socios_ene.keys())
cambios = []

for curp in todos_curps:
    dic_data = socios_dic.get(curp, {'nombre': '', 'total': 0})
    ene_data = socios_ene.get(curp, {'nombre': '', 'total': 0})
    
    if dic_data['total'] != ene_data['total']:
        cambios.append({
            'nombre': ene_data.get('nombre') or dic_data['nombre'],
            'dic': dic_data['total'],
            'ene': ene_data['total'],
            'diff': ene_data['total'] - dic_data['total']
        })

cambios_ordenados = sorted(cambios, key=lambda x: abs(x['diff']), reverse=True)

print(f"\nTotal socios con cambios: {len(cambios)}\n")

for c in cambios_ordenados[:20]:  # Top 20
    print(f"{c['nombre']}")
    print(f"   Diciembre: {c['dic']} → Enero: {c['ene']} ({c['diff']:+d})")

print("\n" + "=" * 100)
