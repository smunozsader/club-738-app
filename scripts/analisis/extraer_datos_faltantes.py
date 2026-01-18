#!/usr/bin/env python3
"""
Extraer datos completos de los 11 socios faltantes del Anexo A
"""
import openpyxl

anexo_a = "2025-dic-usb-738/CLUB 738-31-DE-DICIEMBRE-2025_ANEXOS A, B Y C.xlsx"

faltantes_curps = [
    " MECR871030HYNNRG05",  # Nota: tiene espacio al inicio
    "AERF781023MDFRMR09",
    "CAPG891110HYNMCD00",
    "CECR890104HYNRBL02",
    "FEPR920403HYNRRC06",
    "GOAE840623HYNMRD00",
    "GOMA940118MVZMNM00",
    "GOXK740906HNERXR09",
    "PUSJ000131HYNCSSA4",
    "RODY940625HYNMSL01",
    "RUES971109HCCDSN07"
]

print("=" * 80)
print("DATOS COMPLETOS DE 11 SOCIOS FALTANTES")
print("=" * 80)

wb = openpyxl.load_workbook(anexo_a)
ws = wb["Anexo A"]

# Headers en fila 6:
# Col 1: No. DE REGISTRO
# Col 2: NOMBRE SOCIO
# Col 3: CURP
# Col 4: TELEFONO
# Col 5: EMAIL
# Col 6: No. DE SOCIO
# Col 7: ARMAS CORTAS
# Col 8: ARMAS LARGAS
# Col 9: FECHA ALTA

socios_encontrados = []

for i in range(7, ws.max_row + 1):
    curp = ws.cell(i, 3).value
    if curp and (curp in faltantes_curps or curp.strip() in [c.strip() for c in faltantes_curps]):
        socio = {
            'fila': i,
            'registro': ws.cell(i, 1).value,
            'nombre': ws.cell(i, 2).value,
            'curp': curp.strip() if curp else None,
            'telefono': ws.cell(i, 4).value,
            'email': ws.cell(i, 5).value,
            'num_socio': ws.cell(i, 6).value,
            'armas_cortas': ws.cell(i, 7).value,
            'armas_largas': ws.cell(i, 8).value,
            'fecha_alta': ws.cell(i, 9).value
        }
        socios_encontrados.append(socio)
        
        print(f"\n{len(socios_encontrados)}. {socio['nombre']}")
        print(f"   CURP: {socio['curp']}")
        print(f"   Email: {socio['email']}")
        print(f"   Teléfono: {socio['telefono']}")
        print(f"   No. Socio: {socio['num_socio']}")
        print(f"   Armas: {socio['armas_cortas']} cortas + {socio['armas_largas']} largas")
        print(f"   Fecha alta: {socio['fecha_alta']}")

print(f"\n{'='*80}")
print(f"📊 TOTAL ENCONTRADOS: {len(socios_encontrados)}/11")

if len(socios_encontrados) < 11:
    print(f"\n⚠️  FALTAN {11 - len(socios_encontrados)} SOCIOS")
    curps_encontrados = [s['curp'] for s in socios_encontrados]
    for curp in faltantes_curps:
        if curp.strip() not in curps_encontrados:
            print(f"   - {curp}")

# Contar armas totales
total_cortas = sum(int(s['armas_cortas']) if s['armas_cortas'] else 0 for s in socios_encontrados)
total_largas = sum(int(s['armas_largas']) if s['armas_largas'] else 0 for s in socios_encontrados)

print(f"\n📊 ARMAS DE LOS 11 SOCIOS:")
print(f"   Cortas: {total_cortas}")
print(f"   Largas: {total_largas}")
print(f"   TOTAL: {total_cortas + total_largas}")

wb.close()
