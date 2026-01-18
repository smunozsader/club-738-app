#!/usr/bin/env python3
"""
Verificar y normalizar todos los campos numéricos como texto
"""

from openpyxl import load_workbook
from openpyxl.styles import Alignment

archivo = 'socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx'

print(f"📂 Abriendo: {archivo}")
wb = load_workbook(archivo)
ws = wb.active

headers = [cell.value for cell in ws[1]]
col_telefono = headers.index("TELEFONO") + 1
col_matricula = headers.index("MATRÍCULA") + 1
col_nombre = headers.index("NOMBRE SOCIO") + 1

print(f"\n🔧 Verificando formato de campos...")

cambios_telefono = 0
cambios_matricula = 0

for row in range(2, ws.max_row + 1):
    # Teléfonos
    telefono = ws.cell(row, col_telefono).value
    if telefono:
        # Asegurar formato texto
        cell = ws.cell(row, col_telefono)
        if cell.number_format != '@':
            cell.number_format = '@'
            cambios_telefono += 1
    
    # Matrículas
    matricula = ws.cell(row, col_matricula).value
    if matricula and matricula != "0":
        # Asegurar formato texto
        cell = ws.cell(row, col_matricula)
        if cell.number_format != '@':
            cell.number_format = '@'
            cell.alignment = Alignment(horizontal='left')
            cambios_matricula += 1

print(f"\n📊 Cambios de formato:")
print(f"   Teléfonos → texto: {cambios_telefono}")
print(f"   Matrículas → texto: {cambios_matricula}")

if cambios_telefono > 0 or cambios_matricula > 0:
    print(f"\n💾 Guardando...")
    wb.save(archivo)
    print(f"✅ Archivo actualizado")
else:
    print(f"\n✅ Todos los campos ya tienen formato correcto")

# Mostrar algunos ejemplos
print(f"\n🔍 Ejemplos de teléfonos (primeros 5 socios):")
wb2 = load_workbook(archivo)
ws2 = wb2.active
for row in range(2, min(7, ws2.max_row + 1)):
    nombre = ws2.cell(row, col_nombre).value
    telefono = ws2.cell(row, col_telefono).value
    if nombre and telefono:
        print(f"   {nombre[:30]}: {telefono}")

print(f"\n🔍 Ejemplos de matrículas (REMIGIO):")
for row in range(2, ws2.max_row + 1):
    nombre = ws2.cell(row, col_nombre).value
    if nombre and "REMIGIO" in nombre.upper():
        matricula = ws2.cell(row, col_matricula).value
        marca = ws2.cell(row, headers.index("MARCA") + 1).value
        modelo = ws2.cell(row, headers.index("MODELO") + 1).value
        print(f"   {marca} {modelo}: {matricula}")
        break
