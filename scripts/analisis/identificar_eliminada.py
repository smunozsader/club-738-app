#!/usr/bin/env python3
import openpyxl

archivo_dic = '/Applications/club-738-web/2025-dic-usb-738/CLUB 738-31-DE-DICIEMBRE-2025_RELACION_SOCIOS_ARMAS.xlsx'
archivo_ene = '/Applications/club-738-web/data/socios/Copy of 2026.31.01_RELACION_SOCIOS_ARMAS_SEPARADO_verified.xlsx'

# Leer diciembre
wb_dic = openpyxl.load_workbook(archivo_dic)
ws_dic = wb_dic['CLUB 738. RELACION SOCIOS 31 DI']

armas_dic = {}
for i, row in enumerate(ws_dic.iter_rows(min_row=8, values_only=True), 8):
    if row[5] and 'TOTAL POR PERSONA' in str(row[5]).upper():
        continue
    if row[8]:  # Col 9 (CLASE)
        matricula = str(row[12]).strip() if row[12] else ''
        nombre = str(row[2]).strip() if row[2] else ''
        clase = str(row[8]).strip() if row[8] else ''
        marca = str(row[10]).strip() if row[10] else ''
        modelo = str(row[11]).strip() if row[11] else ''
        if matricula:
            armas_dic[matricula] = {
                'nombre': nombre,
                'clase': clase,
                'marca': marca,
                'modelo': modelo
            }

# Leer enero
wb_ene = openpyxl.load_workbook(archivo_ene)
ws_ene = wb_ene.active
headers_ene = [cell.value for cell in ws_ene[1]]

col_mat = headers_ene.index('MATRÍCULA') + 1
col_nom = headers_ene.index('NOMBRE DEL SOCIO') + 1
col_clase = headers_ene.index('CLASE') + 1

armas_ene = set()
for row in ws_ene.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    matricula = str(row[col_mat - 1]).strip() if row[col_mat - 1] else ''
    if matricula and matricula != 'None':
        armas_ene.add(matricula)

# Encontrar la eliminada
set_dic = set(armas_dic.keys())
eliminadas = set_dic - armas_ene

print('ARMA(S) ELIMINADA(S) entre diciembre 2025 y enero 2026:')
print('=' * 80)
if eliminadas:
    for mat in sorted(eliminadas):
        if mat and mat != 'None':
            data = armas_dic[mat]
            print(f'\nMatrícula: {mat}')
            print(f'Socio: {data["nombre"]}')
            print(f'Clase: {data["clase"]}')
            print(f'Marca: {data["marca"]}')
            print(f'Modelo: {data["modelo"]}')
else:
    print('NO HAY ARMAS ELIMINADAS')

print('\n' + '=' * 80)
print(f'Total eliminadas: {len(eliminadas)}')
