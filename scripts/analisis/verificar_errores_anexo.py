#!/usr/bin/env python3
"""
Verificar errores en Anexo A:
1. Email duplicado de Agustín Moreno
2. Teléfonos de los dos Arieles
"""
import openpyxl

anexo_a = "2025-dic-usb-738/CLUB 738-31-DE-DICIEMBRE-2025_ANEXOS A, B Y C.xlsx"

print("=" * 80)
print("VERIFICANDO ERRORES EN ANEXO A")
print("=" * 80)

wb = openpyxl.load_workbook(anexo_a)
ws = wb["Anexo A"]

# Buscar Ezequiel y Agustín
print("\n🔍 1. VERIFICANDO EMAILS DE EZEQUIEL Y AGUSTÍN:")
print("-" * 80)

for i in range(7, ws.max_row + 1):
    nombre = ws.cell(i, 2).value
    if nombre and ("EZEQUIEL GALVAN" in str(nombre).upper() or "AGUSTIN MORENO" in str(nombre).upper()):
        curp = ws.cell(i, 3).value
        telefono = ws.cell(i, 4).value
        email = ws.cell(i, 5).value
        num_socio = ws.cell(i, 6).value
        
        print(f"\nFila {i}:")
        print(f"  Nombre: {nombre}")
        print(f"  CURP: {curp}")
        print(f"  Teléfono: {telefono}")
        print(f"  Email: {email}")
        print(f"  No. Socio: {num_socio}")

# Buscar los dos Arieles
print("\n\n🔍 2. VERIFICANDO TELÉFONOS DE LOS DOS ARIELES:")
print("-" * 80)

for i in range(7, ws.max_row + 1):
    nombre = ws.cell(i, 2).value
    if nombre and "ARIEL" in str(nombre).upper():
        curp = ws.cell(i, 3).value
        telefono = ws.cell(i, 4).value
        email = ws.cell(i, 5).value
        num_socio = ws.cell(i, 6).value
        
        print(f"\nFila {i}:")
        print(f"  Nombre: {nombre}")
        print(f"  CURP: {curp}")
        print(f"  Teléfono: {telefono}")
        print(f"  Email: {email}")
        print(f"  No. Socio: {num_socio}")

wb.close()
