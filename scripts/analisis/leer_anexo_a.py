#!/usr/bin/env python3
"""
Leer Anexo A del Excel oficial - 76 socios
"""
import openpyxl

archivo_anexos = "2025-dic-usb-738/CLUB 738-31-DE-DICIEMBRE-2025_ANEXOS A, B Y C.xlsx"

print("=" * 80)
print("ANEXO A - LISTA DE 76 SOCIOS (DICIEMBRE 2025)")
print("=" * 80)

wb = openpyxl.load_workbook(archivo_anexos)

# Listar todas las hojas
print(f"\n📋 Hojas disponibles:")
for sheet_name in wb.sheetnames:
    print(f"   - {sheet_name}")

# Leer Anexo A
if "Anexo A" in wb.sheetnames:
    ws = wb["Anexo A"]
    print(f"\n✅ Leyendo hoja 'Anexo A'...")
    print(f"   Total filas: {ws.max_row}")
    print(f"   Total columnas: {ws.max_column}")
    
    # Mostrar primeras 20 filas para entender estructura
    print(f"\n📊 PRIMERAS 20 FILAS:")
    print("=" * 80)
    for i in range(1, min(21, ws.max_row + 1)):
        row_data = []
        for j in range(1, min(10, ws.max_column + 1)):
            val = ws.cell(i, j).value
            if val:
                row_data.append(str(val)[:40])
        if row_data:
            print(f"Fila {i:2d}: {' | '.join(row_data)}")
    
    # Buscar headers
    print(f"\n🔍 BUSCANDO HEADERS...")
    header_row = None
    for i in range(1, 15):
        for j in range(1, ws.max_column + 1):
            val = ws.cell(i, j).value
            if val and 'NOMBRE' in str(val).upper():
                header_row = i
                print(f"   Headers encontrados en fila {i}")
                # Mostrar todos los headers
                print(f"\n   Headers:")
                for k in range(1, ws.max_column + 1):
                    h = ws.cell(i, k).value
                    if h:
                        print(f"      Col {k}: {h}")
                break
        if header_row:
            break
    
    # Contar socios
    if header_row:
        print(f"\n👥 LISTA DE SOCIOS:")
        print("=" * 80)
        socios = []
        for i in range(header_row + 1, ws.max_row + 1):
            nombre = ws.cell(i, 2).value  # Asumir col 2 es nombre
            curp = ws.cell(i, 3).value    # Asumir col 3 es CURP
            if nombre or curp:
                socios.append({
                    'fila': i,
                    'nombre': nombre,
                    'curp': curp
                })
                if len(socios) <= 80:  # Mostrar primeros 80
                    print(f"{len(socios):2d}. {nombre} - {curp}")
        
        print(f"\n📊 TOTAL DE SOCIOS EN ANEXO A: {len(socios)}")
        
        # Guardar lista para comparación
        with open('lista_76_socios_anexo_a.txt', 'w') as f:
            f.write("ANEXO A - 76 SOCIOS (DICIEMBRE 2025)\n")
            f.write("=" * 80 + "\n\n")
            for i, s in enumerate(socios, 1):
                f.write(f"{i}. {s['nombre']} - {s['curp']}\n")
        print(f"\n💾 Lista guardada en: lista_76_socios_anexo_a.txt")

else:
    print(f"\n❌ No se encontró hoja 'Anexo A'")

wb.close()
