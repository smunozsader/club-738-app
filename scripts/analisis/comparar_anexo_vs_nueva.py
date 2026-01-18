#!/usr/bin/env python3
"""
Comparar 76 socios del Anexo A vs Nueva Fuente de Verdad (66 socios)
"""
import openpyxl

# Archivos
anexo_a = "2025-dic-usb-738/CLUB 738-31-DE-DICIEMBRE-2025_ANEXOS A, B Y C.xlsx"
nueva_fuente = "data/socios/2026_ENERO_RELACION_SOCIOS_ARMAS_MASTER.xlsx"

print("=" * 80)
print("COMPARACIÓN: ANEXO A (76 socios) vs NUEVA FUENTE (66 socios)")
print("=" * 80)

# Leer Anexo A
wb_anexo = openpyxl.load_workbook(anexo_a)
ws_anexo = wb_anexo["Anexo A"]

socios_anexo = set()
curps_anexo = {}
for i in range(7, ws_anexo.max_row + 1):
    curp = ws_anexo.cell(i, 3).value
    nombre = ws_anexo.cell(i, 2).value
    if curp and curp != "None" and nombre and "TOTAL" not in str(nombre).upper():
        socios_anexo.add(curp)
        curps_anexo[curp] = nombre

print(f"\n📊 Socios en Anexo A: {len(socios_anexo)}")

# Leer Nueva Fuente de Verdad
wb_nueva = openpyxl.load_workbook(nueva_fuente)
ws_nueva = wb_nueva.active

socios_nueva = set()
curps_nueva = {}
for i in range(2, ws_nueva.max_row + 1):
    curp = ws_nueva.cell(i, 5).value
    nombre = ws_nueva.cell(i, 4).value
    if curp:
        socios_nueva.add(curp)
        if curp not in curps_nueva:
            curps_nueva[curp] = nombre

print(f"📊 Socios en Nueva Fuente: {len(socios_nueva)}")

# Encontrar faltantes
faltantes = socios_anexo - socios_nueva
print(f"\n❌ SOCIOS FALTANTES EN NUEVA FUENTE: {len(faltantes)}")
print("=" * 80)

if faltantes:
    for i, curp in enumerate(sorted(faltantes), 1):
        nombre = curps_anexo.get(curp, "???")
        print(f"{i:2d}. {nombre}")
        print(f"    CURP: {curp}")
        print()

# Encontrar extras (en nueva fuente pero no en anexo)
extras = socios_nueva - socios_anexo
if extras:
    print(f"\n⚠️  SOCIOS EN NUEVA FUENTE PERO NO EN ANEXO A: {len(extras)}")
    print("=" * 80)
    for curp in sorted(extras):
        nombre = curps_nueva.get(curp, "???")
        print(f"   {nombre} - {curp}")

# Guardar lista de faltantes
with open('socios_faltantes_10.txt', 'w') as f:
    f.write("SOCIOS FALTANTES EN NUEVA FUENTE DE VERDAD\n")
    f.write("=" * 80 + "\n\n")
    f.write(f"Total: {len(faltantes)}\n\n")
    for i, curp in enumerate(sorted(faltantes), 1):
        nombre = curps_anexo.get(curp, "???")
        f.write(f"{i:2d}. {nombre}\n")
        f.write(f"    CURP: {curp}\n\n")

print(f"\n💾 Lista guardada en: socios_faltantes_10.txt")

wb_anexo.close()
wb_nueva.close()
