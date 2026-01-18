#!/usr/bin/env python3
import openpyxl

archivo = '/Applications/club-738-web/2025-dic-usb-738/CLUB 738-31-DE-DICIEMBRE-2025_RELACION_SOCIOS_ARMAS.xlsx'

wb = openpyxl.load_workbook(archivo)
print(f'Hojas disponibles: {wb.sheetnames}\n')

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'=== HOJA: {sheet_name} ===')
    print(f'Filas totales: {ws.max_row}')
    print(f'Columnas totales: {ws.max_column}\n')
    
    print('Primeras 15 filas:')
    for i in range(1, min(16, ws.max_row + 1)):
        row = list(ws.iter_rows(min_row=i, max_row=i, values_only=True))[0]
        # Mostrar solo celdas no vacías
        non_empty = [f'Col{j+1}:{v}' for j, v in enumerate(row) if v is not None]
        if non_empty:
            print(f'Fila {i}: {" | ".join(non_empty[:5])}...')
    print('\n')
