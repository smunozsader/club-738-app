#!/usr/bin/env python3
"""
Comparar archivo de diciembre 2025 (USB a 32 ZM) vs enero 2026 (actual)
"""
import openpyxl
from collections import Counter

# Archivos a comparar
archivo_dic = '/Applications/club-738-web/2025-dic-usb-738/CLUB 738-31-DE-DICIEMBRE-2025_RELACION_SOCIOS_ARMAS.xlsx'
archivo_ene = '/Applications/club-738-web/data/socios/Copy of 2026.31.01_RELACION_SOCIOS_ARMAS_SEPARADO_verified.xlsx'

print("=" * 80)
print("COMPARACIÓN DICIEMBRE 2025 (USB 32 ZM) vs ENERO 2026 (ACTUAL)")
print("=" * 80)

# Análisis DICIEMBRE 2025
print("\n📅 DICIEMBRE 2025 - Archivo entregado a 32 ZM")
print("-" * 80)
wb_dic = openpyxl.load_workbook(archivo_dic)
ws_dic = wb_dic.active

print(f"Archivo: {archivo_dic.split('/')[-1]}")
print(f"Hoja activa: {ws_dic.title}")

# Contar filas (saltar encabezado)
total_dic = 0
cortas_dic = 0
largas_dic = 0
socios_dic = {}
armas_dic = []  # Para guardar matrículas

# Detectar columnas
headers_dic = [cell.value for cell in ws_dic[1]]
print(f"\nColumnas detectadas ({len(headers_dic)}):")
for i, h in enumerate(headers_dic, 1):
    if h:
        print(f"  Col {i}: {h}")

# Buscar índices de columnas clave
try:
    col_clase = headers_dic.index('CLASE') + 1 if 'CLASE' in headers_dic else None
    col_email = headers_dic.index('EMAIL') + 1 if 'EMAIL' in headers_dic else None
    col_matricula = headers_dic.index('MATRÍCULA') + 1 if 'MATRÍCULA' in headers_dic else None
    col_nombre = headers_dic.index('NOMBRE DEL SOCIO') + 1 if 'NOMBRE DEL SOCIO' in headers_dic else None
except ValueError as e:
    print(f"⚠️ Error buscando columnas: {e}")
    col_clase = col_email = col_matricula = col_nombre = None

print(f"\nÍndices de columnas:")
print(f"  CLASE: {col_clase}")
print(f"  EMAIL: {col_email}")
print(f"  MATRÍCULA: {col_matricula}")
print(f"  NOMBRE: {col_nombre}")

# Contar armas
for row in ws_dic.iter_rows(min_row=2, values_only=True):
    if not any(row):  # Fila vacía
        continue
    
    if col_clase and row[col_clase - 1]:
        clase = str(row[col_clase - 1]).strip().upper()
        matricula = str(row[col_matricula - 1]).strip() if col_matricula and row[col_matricula - 1] else ""
        email = str(row[col_email - 1]).strip().lower() if col_email and row[col_email - 1] else ""
        nombre = str(row[col_nombre - 1]).strip() if col_nombre and row[col_nombre - 1] else ""
        
        total_dic += 1
        armas_dic.append(matricula)
        
        # Contar por socio
        if email and email != 'nan':
            if email not in socios_dic:
                socios_dic[email] = {'nombre': nombre, 'total': 0, 'cortas': 0, 'largas': 0}
            socios_dic[email]['total'] += 1
            
            # Clasificar
            if 'PISTOLA' in clase or 'REVÓLVER' in clase or 'REVOLVER' in clase or 'KIT' in clase:
                cortas_dic += 1
                socios_dic[email]['cortas'] += 1
            else:
                largas_dic += 1
                socios_dic[email]['largas'] += 1

print(f"\n📊 TOTALES DICIEMBRE 2025:")
print(f"  Total armas: {total_dic}")
print(f"  Cortas (pistolas/revólveres/kits): {cortas_dic}")
print(f"  Largas (rifles/escopetas): {largas_dic}")
print(f"  Total socios con armas: {len(socios_dic)}")

# Análisis ENERO 2026
print("\n" + "=" * 80)
print("📅 ENERO 2026 - Archivo actual (post-sincronización)")
print("-" * 80)
wb_ene = openpyxl.load_workbook(archivo_ene)
ws_ene = wb_ene.active

print(f"Archivo: {archivo_ene.split('/')[-1]}")
print(f"Hoja activa: {ws_ene.title}")

total_ene = 0
cortas_ene = 0
largas_ene = 0
socios_ene = {}
armas_ene = []

headers_ene = [cell.value for cell in ws_ene[1]]

# Buscar índices
try:
    col_clase_ene = headers_ene.index('CLASE') + 1 if 'CLASE' in headers_ene else None
    col_email_ene = headers_ene.index('EMAIL') + 1 if 'EMAIL' in headers_ene else None
    col_matricula_ene = headers_ene.index('MATRÍCULA') + 1 if 'MATRÍCULA' in headers_ene else None
    col_nombre_ene = headers_ene.index('NOMBRE DEL SOCIO') + 1 if 'NOMBRE DEL SOCIO' in headers_ene else None
except ValueError as e:
    print(f"⚠️ Error: {e}")

# Contar armas
for row in ws_ene.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    
    if col_clase_ene and row[col_clase_ene - 1]:
        clase = str(row[col_clase_ene - 1]).strip().upper()
        matricula = str(row[col_matricula_ene - 1]).strip() if col_matricula_ene and row[col_matricula_ene - 1] else ""
        email = str(row[col_email_ene - 1]).strip().lower() if col_email_ene and row[col_email_ene - 1] else ""
        nombre = str(row[col_nombre_ene - 1]).strip() if col_nombre_ene and row[col_nombre_ene - 1] else ""
        
        total_ene += 1
        armas_ene.append(matricula)
        
        if email and email != 'nan':
            if email not in socios_ene:
                socios_ene[email] = {'nombre': nombre, 'total': 0, 'cortas': 0, 'largas': 0}
            socios_ene[email]['total'] += 1
            
            if 'PISTOLA' in clase or 'REVÓLVER' in clase or 'REVOLVER' in clase or 'KIT' in clase:
                cortas_ene += 1
                socios_ene[email]['cortas'] += 1
            else:
                largas_ene += 1
                socios_ene[email]['largas'] += 1

print(f"\n📊 TOTALES ENERO 2026:")
print(f"  Total armas: {total_ene}")
print(f"  Cortas: {cortas_ene}")
print(f"  Largas: {largas_ene}")
print(f"  Total socios: {len(socios_ene)}")

# DIFERENCIAS
print("\n" + "=" * 80)
print("🔍 DIFERENCIAS DICIEMBRE → ENERO")
print("=" * 80)

diff_total = total_ene - total_dic
diff_cortas = cortas_ene - cortas_dic
diff_largas = largas_ene - largas_dic

print(f"\n📈 Cambios en totales:")
print(f"  Total: {total_dic} → {total_ene} ({diff_total:+d})")
print(f"  Cortas: {cortas_dic} → {cortas_ene} ({diff_cortas:+d})")
print(f"  Largas: {largas_dic} → {largas_ene} ({diff_largas:+d})")

# Identificar armas nuevas
set_dic = set(armas_dic)
set_ene = set(armas_ene)
nuevas = set_ene - set_dic
eliminadas = set_dic - set_ene

print(f"\n🆕 Armas NUEVAS (no estaban en diciembre): {len(nuevas)}")
if nuevas:
    for mat in sorted(nuevas):
        if mat and mat != 'nan':
            print(f"  • {mat}")

print(f"\n❌ Armas ELIMINADAS (estaban en diciembre, no en enero): {len(eliminadas)}")
if eliminadas:
    for mat in sorted(eliminadas):
        if mat and mat != 'nan':
            print(f"  • {mat}")

# Socios con cambios
print("\n" + "=" * 80)
print("👥 SOCIOS CON CAMBIOS EN ARSENAL")
print("=" * 80)

todos_emails = set(socios_dic.keys()) | set(socios_ene.keys())

cambios = []
for email in sorted(todos_emails):
    dic_data = socios_dic.get(email, {'nombre': '', 'total': 0, 'cortas': 0, 'largas': 0})
    ene_data = socios_ene.get(email, {'nombre': '', 'total': 0, 'cortas': 0, 'largas': 0})
    
    if dic_data['total'] != ene_data['total']:
        cambios.append({
            'email': email,
            'nombre': ene_data['nombre'] or dic_data['nombre'],
            'dic': dic_data['total'],
            'ene': ene_data['total'],
            'diff': ene_data['total'] - dic_data['total']
        })

if cambios:
    for c in sorted(cambios, key=lambda x: abs(x['diff']), reverse=True):
        print(f"\n{c['nombre']}")
        print(f"  Email: {c['email']}")
        print(f"  Diciembre: {c['dic']} armas")
        print(f"  Enero: {c['ene']} armas")
        print(f"  Cambio: {c['diff']:+d}")
else:
    print("✅ No hay cambios en arsenales por socio")

print("\n" + "=" * 80)
