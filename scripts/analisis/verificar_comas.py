#!/usr/bin/env python3
"""
Buscar y corregir modelos con comas
"""

from openpyxl import load_workbook

archivo = 'socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx'

print(f"📂 Abriendo: {archivo}")
wb = load_workbook(archivo)
ws = wb.active

headers = [cell.value for cell in ws[1]]
col_modelo = headers.index("MODELO") + 1
col_nombre = headers.index("NOMBRE SOCIO") + 1
col_matricula = headers.index("MATRÍCULA") + 1

print(f"\n🔍 Buscando modelos con comas...")
modelos_con_comas = []

for row in range(2, ws.max_row + 1):
    modelo = ws.cell(row, col_modelo).value
    if modelo and ',' in str(modelo):
        nombre = ws.cell(row, col_nombre).value
        modelos_con_comas.append({
            'fila': row,
            'nombre': nombre,
            'modelo_original': modelo
        })

if modelos_con_comas:
    print(f"\n⚠️  Encontrados {len(modelos_con_comas)} modelos con comas:")
    for item in modelos_con_comas:
        print(f"   Fila {item['fila']}: {item['nombre']}")
        print(f"      Modelo: {item['modelo_original']}")
        
        # Corregir
        modelo_nuevo = str(item['modelo_original']).replace(',', '')
        ws.cell(item['fila'], col_modelo).value = modelo_nuevo
        print(f"      → Corregido: {modelo_nuevo}")
    
    # Guardar
    print(f"\n💾 Guardando correcciones...")
    wb.save(archivo)
    print(f"✅ Archivo actualizado")
else:
    print(f"✅ No se encontraron modelos con comas")

# Verificar matrículas con comas
print(f"\n🔍 Verificando matrículas con comas...")
matriculas_con_comas = []

for row in range(2, ws.max_row + 1):
    matricula = ws.cell(row, col_matricula).value
    if matricula and ',' in str(matricula):
        nombre = ws.cell(row, col_nombre).value
        matriculas_con_comas.append({
            'fila': row,
            'nombre': nombre,
            'matricula': matricula
        })

if matriculas_con_comas:
    print(f"⚠️  Encontradas {len(matriculas_con_comas)} matrículas con comas:")
    for item in matriculas_con_comas:
        print(f"   Fila {item['fila']}: {item['nombre']}")
        print(f"      Matrícula: {item['matricula']}")
else:
    print(f"✅ No se encontraron matrículas con comas")
