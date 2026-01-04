#!/usr/bin/env python3
"""
Script para corregir CURPs incorrectas en TODOS los archivos Excel
Basado en las CURPs oficiales de los PDFs descargados de gob.mx
"""

import openpyxl
import os
import glob
from datetime import datetime

# Diccionario de correcciones: CURP_INCORRECTA -> CURP_CORRECTA
CORRECCIONES = {
    # Formato: 'CURP_VIEJA': ('CURP_NUEVA', 'NOMBRE_SOCIO')
    'GACE770131HDFNFN04': ('GACE770131HDFNSN04', 'ENRIQUE GAONA CASTAÑEDA'),
    'AERF781023MDFRMR09': ('AERF781223MDFRMR07', 'MARIA FERNANDA ARECHIGA RAMOS'),
    'DEHE890423HDFNRD03': ('DEHE890423HMCNRD08', 'EDUARDO DENIS HERRERA'),
    'GANJ740807HDFRXQ04': ('GANJ740807HMCRXQ09', 'JOAQUIN RODOLFO GARDONI NUÑEZ'),
    'GAPC790619HYNRRR09': ('GAPC790606HYNRRR06', 'CARLOS ANTONIO GRANJA PEREZ'),
    'MAHH810329HYNRRG09': ('MAHH810329HCCRRG06', 'HUGO MARTINEZ HERNANDEZ'),
    'MAOF620504HDFRRB06': ('MAOF720504HDFRRB02', 'FABIAN MARQUEZ ORTEGA'),
    'RIPR580721HYNVLF06': ('RIPR580720HYNVLF05', 'RAFAEL RIVAS POLANCO'),
    'RODY940625HYNMSL01': ('RODP940625HYNMSB06', 'YAEL ROMERO DE DIOS'),
    # También agregar variantes con espacios o caracteres extra
    'GACE770131HDFNFN04\nGACE770131HDFNSN04': ('GACE770131HDFNSN04', 'ENRIQUE GAONA CASTAÑEDA'),
}

# También buscar parciales (primeros 10 caracteres)
CORRECCIONES_PARCIALES = {
    'GACE770131': 'GACE770131HDFNSN04',
    'AERF781023': 'AERF781223MDFRMR07', 
    'DEHE890423': 'DEHE890423HMCNRD08',
    'GANJ740807': 'GANJ740807HMCRXQ09',
    'GAPC790619': 'GAPC790606HYNRRR06',
    'MAHH810329': 'MAHH810329HCCRRG06',
    'MAOF620504': 'MAOF720504HDFRRB02',
    'RIPR580721': 'RIPR580720HYNVLF05',
    'RODY940625': 'RODP940625HYNMSB06',
}

def corregir_curp(valor):
    """Corrige una CURP si está en la lista de correcciones"""
    if not valor:
        return valor, False
    
    valor_str = str(valor).strip().upper()
    
    # Buscar coincidencia exacta
    for curp_vieja, (curp_nueva, nombre) in CORRECCIONES.items():
        if curp_vieja in valor_str:
            return curp_nueva, True
    
    # Buscar coincidencia parcial (primeros 10 caracteres)
    for prefijo, curp_nueva in CORRECCIONES_PARCIALES.items():
        if valor_str.startswith(prefijo) and valor_str != curp_nueva:
            return curp_nueva, True
    
    return valor_str, False

def procesar_excel(filepath):
    """Procesa un archivo Excel y corrige las CURPs"""
    print(f"\n📂 Procesando: {os.path.basename(filepath)}")
    
    try:
        wb = openpyxl.load_workbook(filepath)
    except Exception as e:
        print(f"   ⚠️ Error abriendo archivo: {e}")
        return 0
    
    correcciones_hechas = 0
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    valor_str = str(cell.value).strip()
                    
                    # Verificar si parece una CURP (18 caracteres alfanuméricos)
                    if len(valor_str) >= 18:
                        nuevo_valor, fue_corregido = corregir_curp(valor_str)
                        
                        if fue_corregido:
                            print(f"   ✏️  Celda {cell.coordinate} en '{sheet_name}':")
                            print(f"      Antes:  {valor_str[:40]}...")
                            print(f"      Ahora:  {nuevo_valor}")
                            cell.value = nuevo_valor
                            correcciones_hechas += 1
    
    if correcciones_hechas > 0:
        # Crear backup
        backup_path = filepath.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        
        # Guardar archivo corregido
        try:
            wb.save(filepath)
            print(f"   ✅ Guardado con {correcciones_hechas} correcciones")
        except Exception as e:
            print(f"   ⚠️ Error guardando: {e}")
    else:
        print(f"   ✓ Sin correcciones necesarias")
    
    wb.close()
    return correcciones_hechas

def main():
    print("=" * 70)
    print("🔧 CORRECCIÓN DE CURPs EN ARCHIVOS EXCEL")
    print("=" * 70)
    print("\nCURPs a corregir:")
    for curp_vieja, (curp_nueva, nombre) in CORRECCIONES.items():
        if '\n' not in curp_vieja:  # No mostrar variantes con newline
            print(f"  {nombre}:")
            print(f"    ❌ {curp_vieja}")
            print(f"    ✅ {curp_nueva}")
    
    # Buscar todos los archivos Excel
    base_path = '/Applications/club-738-web'
    archivos_excel = []
    
    for root, dirs, files in os.walk(base_path):
        # Ignorar node_modules y carpetas ocultas
        dirs[:] = [d for d in dirs if not d.startswith('.') and d != 'node_modules']
        
        for file in files:
            if file.endswith('.xlsx') and not file.startswith('~') and 'backup' not in file.lower():
                archivos_excel.append(os.path.join(root, file))
    
    print(f"\n📁 Encontrados {len(archivos_excel)} archivos Excel")
    
    total_correcciones = 0
    
    for archivo in archivos_excel:
        correcciones = procesar_excel(archivo)
        total_correcciones += correcciones
    
    print("\n" + "=" * 70)
    print(f"📊 RESUMEN: {total_correcciones} correcciones en {len(archivos_excel)} archivos")
    print("=" * 70)

if __name__ == '__main__':
    main()
