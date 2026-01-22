#!/usr/bin/env python3

import openpyxl
from datetime import datetime

archivo = '/Applications/club-738-web/socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx'
wb = openpyxl.load_workbook(archivo)
ws = wb.active

# Obtener headers
headers = [cell.value for cell in ws[1]]

# Obtener todas las filas
filas = list(ws.iter_rows(min_row=2, values_only=True))

print("=" * 100)
print("BUSCANDO SOCIO MÁS NUEVO POR FECHA DE ALTA")
print("=" * 100)

# Encontrar índice de columna de fecha de alta
fecha_col_idx = None
for i, header in enumerate(headers):
    if header and 'fecha' in str(header).lower() and 'alta' in str(header).lower():
        fecha_col_idx = i
        print(f"\nColumna de fecha de alta encontrada: {header} (índice {i})\n")
        break

if fecha_col_idx is None:
    print("\nNo se encontró columna de fecha de alta. Headers disponibles:")
    for i, h in enumerate(headers):
        print(f"  [{i}] {h}")
else:
    # Ordenar filas por fecha de alta (más reciente primero)
    filas_con_fecha = [(fila, fila[fecha_col_idx]) for fila in filas if fila[fecha_col_idx]]
    filas_ordenadas = sorted(filas_con_fecha, key=lambda x: x[1], reverse=True)
    
    print("TOP 5 SOCIOS MÁS NUEVOS:")
    print("-" * 100)
    
    for idx, (fila, fecha) in enumerate(filas_ordenadas[:5], 1):
        credencial = fila[0]
        nombre = fila[1] if len(fila) > 1 else "?"
        email = fila[2] if len(fila) > 2 else "?"
        
        print(f"\n{idx}. Credencial: {credencial}")
        print(f"   Nombre: {nombre}")
        print(f"   Email: {email}")
        print(f"   Fecha de Alta: {fecha}")
        print("-" * 100)
    
    # Mostrar detalle completo del más nuevo
    fila_mas_nueva = filas_ordenadas[0][0]
    
    print("\n" + "=" * 100)
    print("DETALLE COMPLETO DEL SOCIO MÁS NUEVO")
    print("=" * 100)
    
    for i, header in enumerate(headers):
        if header and i < len(fila_mas_nueva):
            valor = fila_mas_nueva[i]
            print(f"{header:35} | {valor}")
