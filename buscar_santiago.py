#!/usr/bin/env python3
import openpyxl

EXCEL_PATH = '/Applications/club-738-web/socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx'
SOCIO_BUSCADO = 'SANTIAGO ALEJANDRO QUINTAL PAREDES'
EMAIL_BUSCADO = 'squintal158@gmail.com'

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

print(f"BUSCANDO: {SOCIO_BUSCADO}\n")
print(f"EMAIL: {EMAIL_BUSCADO}\n")

filas_encontradas = []

for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
    if row and len(row) > 6:
        nombre = str(row[3]).strip().upper() if row[3] else ''
        email = str(row[6]).strip().lower() if row[6] else ''
        
        # Buscar por email exacto (método más seguro)
        if email == EMAIL_BUSCADO.lower():
            filas_encontradas.append((idx, row))
            print(f"Fila {idx}: {row[3]} - {row[6]}")

if not filas_encontradas:
    print(f"⚠️  No encontrado por email. Buscando por nombre parcial...")
    for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and len(row) > 3:
            nombre = str(row[3]).strip().upper() if row[3] else ''
            if 'SANTIAGO ALEJANDRO' in nombre or 'QUINTAL PAREDES' in nombre:
                filas_encontradas.append((idx, row))
                print(f"Fila {idx}: {row}")

print(f"\n📊 Total de filas encontradas: {len(filas_encontradas)}")

if filas_encontradas:
    print(f"\nARMAS DE {SOCIO_BUSCADO}:")
    for idx, row in filas_encontradas:
        print(f"\nFila {idx}:")
        print(f"  Nombre: {row[3]}")
        print(f"  Email: {row[6]}")
        if len(row) > 18:
            print(f"  Clase: {row[13]}")
            print(f"  Calibre: {row[14]}")
            print(f"  Marca: {row[15]}")
            print(f"  Modelo: {row[16]}")
            print(f"  Matrícula: {row[17]}")
            print(f"  Folio: {row[18]}")
