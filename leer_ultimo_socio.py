#!/usr/bin/env python3

import openpyxl
import json

archivo = '/Applications/club-738-web/socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx'
wb = openpyxl.load_workbook(archivo)
ws = wb.active

# Obtener headers
headers = [cell.value for cell in ws[1]]

# Obtener todas las filas
filas = list(ws.iter_rows(min_row=2, values_only=True))

print("=" * 80)
print("SOCIO NUEVO - ÚLTIMO REGISTRO EN FUENTE DE VERDAD")
print("=" * 80)
print(f"\nTotal de socios: {len(filas)}\n")

# Última fila
ultima = filas[-1]

# Mostrar todos los campos
print("DATOS DEL SOCIO NUEVO:")
print("-" * 80)
for i, header in enumerate(headers):
    if header and i < len(ultima):
        valor = ultima[i]
        print(f"{header:30} | {valor}")

# Extraer datos específicos para Firestore
print("\n" + "=" * 80)
print("BÚSQUEDA EN FIRESTORE:")
print("=" * 80)

# Mapear columnas (basado en estructura típica Excel)
credencial = ultima[0]
nombre = ultima[1]
email = ultima[2] if len(ultima) > 2 else None
curp = ultima[3] if len(ultima) > 3 else None
telefono = ultima[4] if len(ultima) > 4 else None

print(f"\nCredencial: {credencial}")
print(f"Nombre: {nombre}")
print(f"Email: {email}")
print(f"CURP: {curp}")
print(f"Teléfono: {telefono}")

if email:
    print(f"\n→ Buscar en Firestore: socios/{email.lower()}")
