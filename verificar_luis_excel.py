#!/usr/bin/env python3

import openpyxl

archivo = '/Applications/club-738-web/socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx'
wb = openpyxl.load_workbook(archivo)
ws = wb.active

# Obtener headers
headers = [cell.value for cell in ws[1]]

# Obtener todas las filas
filas = list(ws.iter_rows(min_row=2, values_only=True))

print("=" * 100)
print("BÚSQUEDA EN EXCEL: Credencial 236 - LUIS FERNANDO GUILLERMO GAMBOA")
print("=" * 100)

# Buscar por credencial 236
credencial_col = headers.index('No. CREDENCIAL') if 'No. CREDENCIAL' in headers else 2

encontrado = False

for fila in filas:
    if fila[credencial_col] == 236:
        encontrado = True
        
        print(f"\n✅ ENCONTRADO EN EXCEL\n")
        
        # Mostrar todos los datos
        for i, header in enumerate(headers):
            if header and i < len(fila):
                valor = fila[i]
                print(f"{header:35} | {valor}")
        
        # Verificar campos de arma
        print("\n" + "=" * 100)
        print("VERIFICACIÓN DE ARMAS:")
        print("=" * 100)
        
        clase_col = headers.index('CLASE') if 'CLASE' in headers else None
        clase = fila[clase_col] if clase_col and clase_col < len(fila) else None
        
        if clase is None or clase == 0 or clase == '0' or str(clase).strip() == '':
            print("\n✅ SIN ARMAS REGISTRADAS")
            print(f"   Campo CLASE: {clase}")
        else:
            print(f"\n❌ CON ARMAS REGISTRADAS")
            print(f"   CLASE: {clase}")
            
            # Mostrar detalles de arma si existen
            if 'CALIBRE' in headers:
                calibre_idx = headers.index('CALIBRE')
                print(f"   CALIBRE: {fila[calibre_idx]}")
            if 'MARCA' in headers:
                marca_idx = headers.index('MARCA')
                print(f"   MARCA: {fila[marca_idx]}")
            if 'MODELO' in headers:
                modelo_idx = headers.index('MODELO')
                print(f"   MODELO: {fila[modelo_idx]}")
            if 'MATRÍCULA' in headers:
                matricula_idx = headers.index('MATRÍCULA')
                print(f"   MATRÍCULA: {fila[matricula_idx]}")
        
        break

if not encontrado:
    print(f"\n❌ NO ENCONTRADO EN EXCEL")
    print(f"\nLa credencial 236 no existe en la fuente de verdad")

print("\n" + "=" * 100)
