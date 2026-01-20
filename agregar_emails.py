#!/usr/bin/env python3
"""Agregar emails faltantes a Gardoni y Arechiga en Excel"""

import openpyxl

EXCEL_PATH = "/Applications/club-738-web/socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx"

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

gardoni_email = "jrgardoni@gmail.com"
arechiga_email = "arechiga@jogarplastics.com"

print("Agregando emails faltantes...")

# Buscar y agregar emails
for row_num in range(2, ws.max_row + 1):
    nombre = ws[f'D{row_num}'].value
    email = ws[f'G{row_num}'].value
    
    if nombre:
        nombre_str = str(nombre).upper()
        
        if "GARDONI" in nombre_str and (not email or str(email).lower() == "nan"):
            ws[f'G{row_num}'] = gardoni_email
            print(f"✅ Fila {row_num}: Email agregado a Gardoni")
        
        elif "ARECHIGA" in nombre_str and (not email or str(email).lower() == "nan"):
            ws[f'G{row_num}'] = arechiga_email
            print(f"✅ Fila {row_num}: Email agregado a Arechiga")

wb.save(EXCEL_PATH)
print(f"\n✅ Excel actualizado correctamente")
