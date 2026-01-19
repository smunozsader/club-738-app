#!/usr/bin/env python3
import openpyxl
from openpyxl.utils import get_column_letter

print("\n📋 FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx (CORREGIDA)\n")

try:
    wb = openpyxl.load_workbook('socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx')
    ws = wb.active
    
    # Leer encabezados
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    
    print("=" * 200)
    print(f"{'ENCABEZADOS':<50}")
    print("=" * 200)
    
    # Mostrar encabezados con números de columna
    for i, h in enumerate(headers, 1):
        print(f"  Col {i:2d}: {h}")
    
    print("\n" + "=" * 200)
    print(f"{'PRIMEROS 15 REGISTROS (CORREGIDOS)':<50}")
    print("=" * 200 + "\n")
    
    # Mostrar primeros 15 registros
    for row_idx in range(2, min(17, ws.max_row + 1)):
        print(f"\n[Row {row_idx}]")
        
        # Mostrar columnas importantes
        cols_mostrar = [0, 3, 4, 5, 6, 7, 10, 11, 12, 13, 14, 15, 16, 17, 18]  # Índices 0-based
        
        for col_idx in cols_mostrar:
            if col_idx < len(headers):
                cell_value = ws.cell(row_idx, col_idx + 1).value
                header = headers[col_idx] or f"Col{col_idx + 1}"
                
                # Resaltar columnas corregidas
                if col_idx == 11:  # ESTADO
                    print(f"  ✅ {header:30s} = '{cell_value}'")
                elif col_idx == 12:  # CP
                    print(f"  ✅ {header:30s} = '{cell_value}'")
                else:
                    print(f"     {header:30s} = '{cell_value}'")
    
    print("\n" + "=" * 200)
    print(f"TOTAL DE REGISTROS: {ws.max_row - 1}")
    print("=" * 200 + "\n")
    
except Exception as e:
    print(f"❌ Error: {e}\n")
    import traceback
    traceback.print_exc()
