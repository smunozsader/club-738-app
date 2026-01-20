#!/usr/bin/env python3
import openpyxl
import os

excel_path = "socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx"

wb = openpyxl.load_workbook(excel_path)
ws = wb.active

print("📋 BUSCANDO A RICARDO EN LA COLUMNA NOMBRE SOCIO:\n")

ricardo_rows = []
for row in range(2, ws.max_row + 1):
    nombre = ws.cell(row, 4).value  # Columna 4: NOMBRE SOCIO
    if nombre and "RICARDO" in str(nombre).upper():
        print(f"Fila {row}:")
        print(f"  CREDENCIAL: {ws.cell(row, 3).value}")
        print(f"  NOMBRE: {ws.cell(row, 4).value}")
        print(f"  EMAIL: {ws.cell(row, 7).value}")
        print(f"  CLASE: {ws.cell(row, 14).value}")
        print(f"  CALIBRE: {ws.cell(row, 15).value}")
        print(f"  MARCA: {ws.cell(row, 16).value}")
        print(f"  MODELO: {ws.cell(row, 17).value}")
        print(f"  MATRÍCULA: {ws.cell(row, 18).value}")
        print(f"  FOLIO: {ws.cell(row, 19).value}")
        print()
        ricardo_rows.append(row)

print(f"✓ Total filas encontradas: {len(ricardo_rows)}")
