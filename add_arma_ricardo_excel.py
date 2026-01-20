#!/usr/bin/env python3
import openpyxl
from datetime import datetime
import os

excel_path = "socios/FUENTE_DE_VERDAD_CLUB_738_ENERO_2026.xlsx"

wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# Datos de la nueva arma
new_arma = {
    "registro": 284,  # Próximo número
    "domicilio": None,  # Basado en fila anterior
    "credencial": 230,
    "nombre": "RICARDO ANTONIO SOBERANIS GAMBOA",
    "curp": None,
    "telefono": None,
    "email": "rsoberanis11@hotmail.com",
    "fecha_alta": None,
    "calle": None,
    "colonia": None,
    "ciudad": None,
    "estado": None,
    "cp": None,
    "clase": "PISTOLA",
    "calibre": ".40 S&W",
    "marca": "CZ",
    "modelo": "P-10 C",
    "matricula": "EP29710",
    "folio": "A3912487"
}

# Copiar datos de la fila anterior para llenar vacíos
for row in range(280, 283):
    if row == 282:  # Última fila de Ricardo
        prev_registro = ws.cell(row, 1).value
        prev_domicilio = ws.cell(row, 2).value
        prev_curp = ws.cell(row, 5).value
        prev_telefono = ws.cell(row, 6).value
        prev_fecha_alta = ws.cell(row, 8).value
        prev_calle = ws.cell(row, 9).value
        prev_colonia = ws.cell(row, 10).value
        prev_ciudad = ws.cell(row, 11).value
        prev_estado = ws.cell(row, 12).value
        prev_cp = ws.cell(row, 13).value
        break

# Insertar nueva fila (fila 283)
new_row = 283
ws.insert_rows(new_row, 1)

# Establecer valores
ws.cell(new_row, 1).value = prev_registro  # No. REGISTRO
ws.cell(new_row, 2).value = prev_domicilio  # DOMICILIO CLUB
ws.cell(new_row, 3).value = 230  # No. CREDENCIAL
ws.cell(new_row, 4).value = "RICARDO ANTONIO SOBERANIS GAMBOA"  # NOMBRE SOCIO
ws.cell(new_row, 5).value = prev_curp  # CURP
ws.cell(new_row, 6).value = prev_telefono  # TELEFONO
ws.cell(new_row, 7).value = "rsoberanis11@hotmail.com"  # EMAIL
ws.cell(new_row, 8).value = prev_fecha_alta  # FECHA ALTA
ws.cell(new_row, 9).value = prev_calle  # CALLE
ws.cell(new_row, 10).value = prev_colonia  # COLONIA
ws.cell(new_row, 11).value = prev_ciudad  # CIUDAD
ws.cell(new_row, 12).value = prev_estado  # ESTADO
ws.cell(new_row, 13).value = prev_cp  # CP
ws.cell(new_row, 14).value = "PISTOLA"  # CLASE
ws.cell(new_row, 15).value = ".40 S&W"  # CALIBRE
ws.cell(new_row, 16).value = "CZ"  # MARCA
ws.cell(new_row, 17).value = "P-10 C"  # MODELO
ws.cell(new_row, 18).value = "EP29710"  # MATRÍCULA
ws.cell(new_row, 19).value = "A3912487"  # FOLIO

# Guardar
wb.save(excel_path)

print("✓ Nueva arma agregada al Excel")
print(f"   Fila: {new_row}")
print(f"   CLASE: PISTOLA")
print(f"   CALIBRE: .40 S&W")
print(f"   MARCA: CZ")
print(f"   MODELO: P-10 C")
print(f"   MATRÍCULA: EP29710")
print(f"   FOLIO: A3912487")
print(f"\n✓ Excel guardado")
